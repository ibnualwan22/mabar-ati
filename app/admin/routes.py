import io
from . import admin_bp
from flask import render_template, redirect, url_for, flash, request, jsonify, abort, Response, send_file, current_app
from app.models import Absen, ActivityLog, Rombongan, Tarif, Santri, Pendaftaran, Izin, Partisipan, Transaksi, User, Edisi, Bus, Role, Wisuda
from app.admin.forms import ChangePasswordForm, EditIzinForm, GrupWaliURLForm, ImportPartisipanForm, ImportWisudaForm, KonfirmasiSetoranForm, PengeluaranBusForm, PetugasLapanganForm, RombonganForm, SantriEditForm, PendaftaranForm, PendaftaranEditForm, IzinForm, PartisipanForm, PartisipanEditForm, LoginForm, SantriManualForm, TransaksiForm, UserForm, UserEditForm, EdisiForm, BusForm, WisudaForm
from app import db, login_manager
import json, requests
from collections import defaultdict
from sqlalchemy import update, text, or_, and_, func
from datetime import date, datetime
from flask_login import login_user, logout_user, current_user, login_required
from functools import wraps
from sqlalchemy.orm import joinedload, selectinload # <-- Tambahkan import ini di atas
from datetime import date, datetime # Pastikan datetime diimpor
from io import BytesIO

from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.pdfgen import canvas
import locale
import pandas as pd




def check_and_update_expired_izin():
    """Mencari izin yang kadaluwarsa, mengubah statusnya, dan mengembalikan status santri."""
    today = date.today()
    # Cari izin yang statusnya 'Aktif' dan sudah lewat tanggal
    expired_izins = Izin.query.filter(Izin.status == 'Aktif', Izin.tanggal_berakhir < today).all()
    
    if not expired_izins:
        return

    for izin in expired_izins:
        santri = izin.santri
        if santri and santri.status_santri == 'Izin':
            santri.status_santri = 'Aktif'
            izin.status = 'Selesai' # <-- Ubah status, bukan hapus
            log_activity('Update Otomatis', 'Perizinan', f"Izin untuk '{santri.nama}' telah berakhir.")

    db.session.commit()
    print(f"Update otomatis: {len(expired_izins)} status izin telah diperbarui.")

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# Di app/admin/routes.py
@admin_bp.before_request
def require_login():
    # Cek jika endpoint yang diminta BUKAN halaman login dan user BELUM login
    if request.endpoint and 'admin.' in request.endpoint and request.endpoint != 'admin.login':
        if not current_user.is_authenticated:
            return redirect(url_for('admin.login', next=request.url))

@admin_bp.route('/login', methods=['GET', 'POST'])
def login():
    # Jika user sudah login, arahkan ke dashboard
    if current_user.is_authenticated:
        return redirect(url_for('admin.dashboard'))
    
    form = LoginForm()
    if form.validate_on_submit():
        # Cari user di database berdasarkan username
        user = User.query.filter_by(username=form.username.data).first()
        
        # Cek apakah user ada dan passwordnya cocok
        if user and user.check_password(form.password.data):
            login_user(user, remember=form.remember_me.data)
            # Arahkan ke halaman yang tadinya ingin diakses, atau ke dashboard
            next_page = request.args.get('next')
            return redirect(next_page) if next_page else redirect(url_for('admin.dashboard'))
        else:
            flash('Username atau password salah.', 'danger')
            
    return render_template('login.html', form=form)

# Di app/admin/routes.py
@admin_bp.route('/logout')
def logout():
    logout_user()
    return redirect(url_for('admin.login'))

def role_required(*roles): # Terima beberapa nama peran
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not current_user.is_authenticated:
                return redirect(url_for('admin.login', next=request.url))
            if current_user.role.name not in roles:
                abort(403) # Forbidden - Akses Ditolak
            return f(*args, **kwargs)
        return decorated_function
    return decorator

@admin_bp.route('/users')
@login_required
@role_required('Korpus', 'Korpuspi')
def manajemen_user():
    users = User.query.all()
    return render_template('manajemen_user.html', users=users)

def setup_user_form_choices(form):
    """Fungsi helper untuk mengisi pilihan dropdown form user."""
    if current_user.role.name == 'Korda':
        # Korda hanya bisa membuat Sarpras
        form.role.choices = [(r.id, r.name) for r in Role.query.filter_by(name='Sarpras')]
        # Korda hanya bisa menugaskan ke rombongannya sendiri
        form.managed_rombongan_single.choices = [(r.id, r.nama_rombongan) for r in current_user.active_managed_rombongan]
    else: # Korpus
        form.role.choices = [(r.id, r.name) for r in Role.query.all()]
        form.managed_rombongan_single.choices = [(r.id, r.nama_rombongan) for r in Rombongan.query.all()]

@admin_bp.route('/users/tambah', methods=['GET', 'POST'])
@login_required
@role_required('Korpus', 'Korpuspi')
def tambah_user():
    form = UserForm()
    setup_user_form_choices(form) # Panggil helper di sini
    if form.validate_on_submit():
        # Buat objek user terlebih dahulu
        new_user = User(
            username=form.username.data,
            role=form.role.data
        )
        # Set password-nya
        new_user.set_password(form.password.data)
        
        # Tentukan rombongan yang dikelola berdasarkan peran
        if form.role.data.name == 'Korwil':
            new_user.managed_rombongan = form.managed_rombongan_multi.data
        elif form.role.data.name == 'Korda':
            # Pastikan ada data sebelum dimasukkan ke list
            if form.managed_rombongan_single.data:
                new_user.managed_rombongan = [form.managed_rombongan_single.data]
            else:
                new_user.managed_rombongan = []
        else: # Untuk peran lain, kosongkan
            new_user.managed_rombongan = []
            
        # Simpan user baru ke database
        db.session.add(new_user)
        log_activity('Tambah', 'User', f"Menambahkan user baru: '{form.username.data}' dengan peran '{form.role.data.name}'")
        db.session.commit()
        
        flash('User baru berhasil ditambahkan.', 'success')
        return redirect(url_for('admin.manajemen_user'))
        
    return render_template('form_user.html', form=form, title="Tambah User Baru")


@admin_bp.route('/users/edit/<int:user_id>', methods=['GET', 'POST'])
@login_required
@role_required('Korpus', 'Korpuspi')
def edit_user(user_id):
    user = User.query.get_or_404(user_id)
    # Saat GET, form diisi dengan data awal dari 'user'
    form = UserEditForm(obj=user)
    setup_user_form_choices(form) # Panggil helper di sini juga

    # Saat POST, kita proses datanya
    if form.validate_on_submit():
        user.username = form.username.data
        user.role = form.role.data
        
        # Tentukan rombongan yang dikelola berdasarkan peran
        if form.role.data.name == 'Korwil':
            user.managed_rombongan = form.managed_rombongan_multi.data
        elif form.role.data.name == 'Korda':
            if form.managed_rombongan_single.data:
                user.managed_rombongan = [form.managed_rombongan_single.data]
            else:
                user.managed_rombongan = []
        else:
            user.managed_rombongan = []

        # Jika ada password baru yang diisi, update passwordnya
        if form.password.data:
            user.set_password(form.password.data)
            
        log_activity('Edit', 'User', f"Mengubah data user: '{user.username}'")
        db.session.commit()
        
        flash('Data user berhasil diperbarui.', 'success')
        return redirect(url_for('admin.manajemen_user'))

    # Saat GET, isi field rombongan secara manual karena 'obj' tidak menanganinya dengan baik untuk field ini
    if user.role.name == 'Korwil':
        form.managed_rombongan_multi.data = user.managed_rombongan
    elif user.role.name == 'Korda' and user.managed_rombongan:
        form.managed_rombongan_single.data = user.managed_rombongan[0]

    return render_template('form_user.html', form=form, title="Edit User")

@admin_bp.route('/users/hapus/<int:user_id>', methods=['POST'])
@login_required
@role_required('Korpus', 'Korpuspi')
def hapus_user(user_id):
    user = User.query.get_or_404(user_id)
    # Jangan biarkan korpus menghapus dirinya sendiri
    if user == current_user:
        flash('Anda tidak bisa menghapus akun Anda sendiri.', 'danger')
        return redirect(url_for('admin.manajemen_user'))
    username_dihapus = user.username
    log_activity('Hapus', 'User', f"Menghapus user: '{username_dihapus}'")
    db.session.delete(user)
    db.session.commit()
    flash('User berhasil dihapus.', 'info')
    return redirect(url_for('admin.manajemen_user'))


@role_required('Korpus', 'Korda', 'Korwil', 'Keamanan', 'PJ Acara', 'Bendahara Pusat', 'Sekretaris', 'Korpuspi')
@admin_bp.route('/dashboard')
@login_required
def dashboard():
    active_edisi = get_active_edisi()
    if not active_edisi:
        return render_template('dashboard.html', active_edisi=None)

    # Tentukan apakah user bisa melihat semua data atau hanya wilayahnya
    is_full_access = current_user.role.name in ['Korpus', 'Korpuspi']
    is_regional_access = current_user.role.name in ['Korda', 'Korwil']
    
    # Ambil semua rombongan yang dikelola user untuk filter dropdown
    if is_full_access:
        all_managed_rombongan = Rombongan.query.filter(
            Rombongan.edisi_id == active_edisi.id
        ).order_by(Rombongan.nama_rombongan).all()
        managed_rombongan_ids = {r.id for r in all_managed_rombongan}
    else:
        all_managed_rombongan = current_user.active_managed_rombongan
        managed_rombongan_ids = {r.id for r in all_managed_rombongan}

    # Kumpulkan wilayah yang dikelola user untuk filter regional
    managed_regions = set()
    if is_regional_access:
        for r in current_user.active_managed_rombongan:
            if r.cakupan_wilayah:
                for w in r.cakupan_wilayah:
                    managed_regions.add(w.get('label'))

    # --- LOGIKA FILTER ---
    selected_rombongan_id = request.args.get('rombongan_id', type=int)
    
    # Siapkan query dasar untuk pendaftaran
    pendaftaran_query = Pendaftaran.query.options(
        joinedload(Pendaftaran.santri),
        joinedload(Pendaftaran.rombongan_pulang).joinedload(Rombongan.tarifs),
        joinedload(Pendaftaran.rombongan_kembali).joinedload(Rombongan.tarifs)
    ).filter(Pendaftaran.edisi_id == active_edisi.id)

    # Terapkan filter berdasarkan hak akses dan pilihan dropdown
    if is_regional_access:
        # Untuk regional access, ambil semua peserta yang:
        # 1. Ikut rombongan yang dikelola (termasuk lintas wilayah)
        # 2. ATAU berasal dari wilayah yang dikelola
        pendaftaran_query = pendaftaran_query.filter(
            or_(
                Pendaftaran.rombongan_pulang_id.in_(managed_rombongan_ids),
                Pendaftaran.rombongan_kembali_id.in_(managed_rombongan_ids)
            )
        )
    elif is_full_access and selected_rombongan_id:
        pendaftaran_query = pendaftaran_query.filter(
            or_(
                Pendaftaran.rombongan_pulang_id == selected_rombongan_id,
                Pendaftaran.rombongan_kembali_id == selected_rombongan_id
            )
        )
    elif not is_full_access and not is_regional_access:
        # Untuk role lain, hanya tampilkan data peserta terdaftar tanpa filter lainnya
        pass

    pendaftar_list = pendaftaran_query.all()
    registered_santri_ids = {p.santri_id for p in pendaftar_list}

    # --- PERHITUNGAN STATISTIK KARTU RINGKASAN ---
    stats = {}
    
    # Base queries untuk berbagai statistik
    santri_q = Santri.query
    izin_q = Izin.query.filter(Izin.edisi_id == active_edisi.id)
    partisipan_q = Partisipan.query.filter(Partisipan.edisi_id == active_edisi.id)
    wisuda_q = Wisuda.query.filter(Wisuda.edisi_id == active_edisi.id)

    # Apply regional filter untuk santri belum daftar - hanya untuk santri dari wilayah yang dikelola
    # Ini tetap berdasarkan wilayah asal santri, bukan rombongan
    if is_regional_access and managed_regions:
        santri_q = santri_q.filter(Santri.kabupaten.in_(managed_regions))
        izin_q = izin_q.join(Santri).filter(Santri.kabupaten.in_(managed_regions))
        partisipan_q = partisipan_q.join(Santri).filter(Santri.kabupaten.in_(managed_regions))
        wisuda_q = wisuda_q.join(Santri, Santri.nis == Wisuda.santri_nis).filter(Santri.kabupaten.in_(managed_regions))

    # Hitung statistik dasar
    stats['total_peserta'] = len(pendaftar_list)
    
    # Untuk role selain Korpus/Korpuspi, hanya hitung santri belum daftar dari wilayah yang dikelola
    if not is_full_access and not is_regional_access:
        stats['santri_belum_terdaftar'] = 0  # Tidak ditampilkan untuk role lain
        stats['total_izin'] = 0
        stats['total_partisipan'] = 0
        stats['total_wisuda'] = 0
    else:
        stats['santri_belum_terdaftar'] = santri_q.filter(
            Santri.status_santri == 'Aktif',
            ~Santri.id.in_(registered_santri_ids)
        ).count()
        stats['total_izin'] = izin_q.count()
        stats['total_partisipan'] = partisipan_q.count()
        stats['total_wisuda'] = wisuda_q.count()

    # Hitung statistik gelombang pulang dan total kembali - DIPERBAIKI
    stats['pulang_gelombang_1'] = 0
    stats['pulang_gelombang_2'] = 0
    stats['total_kembali'] = 0

    for p in pendaftar_list:
        # STATISTIK PESERTA PULANG
        if p.status_pulang != 'Tidak Ikut':
            # Filter untuk regional access: hanya hitung jika ikut rombongan pulang yang dikelola
            should_count_pulang = True
            if is_regional_access:
                should_count_pulang = p.rombongan_pulang_id in managed_rombongan_ids
            
            if should_count_pulang:
                # Tentukan gelombang berdasarkan prioritas
                gelombang = None
                if hasattr(p, 'gelombang_pulang') and p.gelombang_pulang:
                    gelombang = p.gelombang_pulang
                elif p.rombongan_pulang and hasattr(p.rombongan_pulang, 'gelombang'):
                    gelombang = p.rombongan_pulang.gelombang
                
                # Hitung statistik berdasarkan gelombang
                if gelombang == 1:
                    stats['pulang_gelombang_1'] += 1
                elif gelombang == 2:
                    stats['pulang_gelombang_2'] += 1

        # STATISTIK PESERTA KEMBALI - LOGIKA DIPERBAIKI
        if p.status_kembali != 'Tidak Ikut':
            should_count_kembali = True
            
            if is_regional_access:
                # Jika ada rombongan kembali eksplisit, gunakan itu
                if p.rombongan_kembali_id:
                    should_count_kembali = p.rombongan_kembali_id in managed_rombongan_ids
                # Jika tidak ada rombongan kembali eksplisit, gunakan rombongan pulang
                # TAPI hanya jika rombongan pulang juga dikelola
                elif p.rombongan_pulang_id:
                    should_count_kembali = p.rombongan_pulang_id in managed_rombongan_ids
                else:
                    should_count_kembali = False
            
            if should_count_kembali:
                stats['total_kembali'] += 1

    # --- PERHITUNGAN DATA GRAFIK (hanya untuk yang punya akses penuh atau regional) ---
    chart_data_alokasi = {'bus': 0, 'korda': 0, 'pondok': 0}
    chart_data_status = {'lunas': 0, 'belum_lunas': 0}
    
    if is_full_access or is_regional_access:
        FEE_PONDOK = 10000

        for p in pendaftar_list:
            # Kalkulasi biaya pulang - hanya jika rombongan pulang sesuai filter
            if (p.status_pulang != 'Tidak Ikut' and p.rombongan_pulang and p.titik_turun):
                # Untuk full access: hitung semua atau hanya selected_rombongan jika ada
                should_count_pulang = False
                if is_full_access:
                    if selected_rombongan_id:
                        should_count_pulang = p.rombongan_pulang_id == selected_rombongan_id
                    else:
                        should_count_pulang = True
                elif is_regional_access:
                    should_count_pulang = p.rombongan_pulang_id in managed_rombongan_ids
                
                if should_count_pulang:
                    tarif_pulang = next((t for t in p.rombongan_pulang.tarifs 
                                      if t.titik_turun == p.titik_turun), None)
                    if tarif_pulang:
                        total_biaya_pulang = tarif_pulang.harga_bus + tarif_pulang.fee_korda + FEE_PONDOK
                        chart_data_alokasi['bus'] += tarif_pulang.harga_bus
                        chart_data_alokasi['korda'] += tarif_pulang.fee_korda
                        chart_data_alokasi['pondok'] += FEE_PONDOK
                        if p.status_pulang == 'Lunas':
                            chart_data_status['lunas'] += total_biaya_pulang
            
            # Kalkulasi biaya kembali - hanya jika rombongan kembali sesuai filter
            if p.status_kembali != 'Tidak Ikut':
                rombongan_utk_kembali = p.rombongan_kembali or p.rombongan_pulang
                titik_jemput = p.titik_jemput_kembali or p.titik_turun
                
                if rombongan_utk_kembali and titik_jemput:
                    # Untuk full access: hitung semua atau hanya selected_rombongan jika ada
                    should_count_kembali = False
                    if is_full_access:
                        if selected_rombongan_id:
                            should_count_kembali = rombongan_utk_kembali.id == selected_rombongan_id
                        else:
                            should_count_kembali = True
                    elif is_regional_access:
                        should_count_kembali = rombongan_utk_kembali.id in managed_rombongan_ids
                    
                    if should_count_kembali:
                        tarif_kembali = next((t for t in rombongan_utk_kembali.tarifs 
                                           if t.titik_turun == titik_jemput), None)
                        if tarif_kembali:
                            total_biaya_kembali = tarif_kembali.harga_bus + tarif_kembali.fee_korda + FEE_PONDOK
                            chart_data_alokasi['bus'] += tarif_kembali.harga_bus
                            chart_data_alokasi['korda'] += tarif_kembali.fee_korda
                            chart_data_alokasi['pondok'] += FEE_PONDOK
                            if p.status_kembali == 'Lunas':
                                chart_data_status['lunas'] += total_biaya_kembali

        total_pemasukan = sum(chart_data_alokasi.values())
        chart_data_status['belum_lunas'] = total_pemasukan - chart_data_status['lunas']

    return render_template('dashboard.html', 
                           active_edisi=active_edisi,
                           stats=stats,
                           semua_rombongan=all_managed_rombongan,
                           selected_rombongan_id=selected_rombongan_id,
                           chart_data_alokasi=chart_data_alokasi,
                           chart_data_status=chart_data_status,
                           managed_regions=list(managed_regions) if managed_regions else [],
                           is_full_access=is_full_access,
                           is_regional_access=is_regional_access,
                           show_filter=is_full_access or is_regional_access
                           )
@admin_bp.route('/rombongan')
@login_required
@role_required('Korpus', 'Korwil', 'Korda', 'Bendahara Pusat', 'Sekretaris', 'Korpuspi')
def manajemen_rombongan():
    active_edisi = get_active_edisi()
    semua_rombongan = []  # Inisialisasi daftar kosong di awal
    search_query = None   # <--- Tambahkan inisialisasi awal di sini

    if active_edisi:
        # Query hanya dijalankan jika ada edisi yang aktif
        query = Rombongan.query.filter_by(edisi=active_edisi)
        
        if current_user.role.name in ['Korwil', 'Korda']:
            managed_rombongan_ids = [r.id for r in current_user.active_managed_rombongan]
            if not managed_rombongan_ids:
                query = query.filter(db.false())
            else:
                query = query.filter(Rombongan.id.in_(managed_rombongan_ids))
        
        search_query = request.args.get('q')
        
        if search_query:
            query = query.filter(Rombongan.nama_rombongan.ilike(f'%{search_query}%'))

        semua_rombongan = query.order_by(Rombongan.nama_rombongan).all()

    return render_template('manajemen_rombongan.html', semua_rombongan=semua_rombongan)



# --- KODE BARU DIMULAI DI SINI ---

@admin_bp.route('/rombongan/tambah', methods=['GET', 'POST'])
@login_required
@role_required('Korpus', 'Sekretaris', 'Korpuspi')
def tambah_rombongan():
    active_edisi = get_active_edisi()
    if not active_edisi:
        flash("Tidak bisa menambah rombongan karena tidak ada edisi yang aktif.", "danger")
        return redirect(url_for('admin.manajemen_rombongan'))
    
    form = RombonganForm()
    if form.validate_on_submit():
        new_rombongan = Rombongan(
            edisi=active_edisi,
            nama_rombongan=form.nama_rombongan.data,
            penanggung_jawab_putra=form.penanggung_jawab_putra.data,
            kontak_person_putra=form.kontak_person_putra.data,
            penanggung_jawab_putri=form.penanggung_jawab_putri.data,
            kontak_person_putri=form.kontak_person_putri.data,
            nomor_rekening=form.nomor_rekening.data,
            cakupan_wilayah=json.loads(form.cakupan_wilayah.data or '[]'),
            jadwal_pulang=form.jadwal_pulang.data,
            batas_pembayaran_pulang=form.batas_pembayaran_pulang.data,
            jadwal_berangkat=form.jadwal_berangkat.data,
            batas_pembayaran_berangkat=form.batas_pembayaran_berangkat.data,
            titik_jemput_berangkat=form.titik_jemput_berangkat.data
        )
        # Hapus tarif lama (jika ada, seharusnya tidak ada untuk form tambah) dan isi dengan yang baru
        new_rombongan.tarifs = []
        for tarif_data in form.tarifs.data:
            if tarif_data['titik_turun'] and tarif_data['harga_bus'] is not None:
                new_rombongan.tarifs.append(Tarif(**tarif_data))
        
        db.session.add(new_rombongan)
        log_activity('Tambah', 'Rombongan', f"Menambahkan rombongan baru: '{new_rombongan.nama_rombongan}'")
        db.session.commit()
        flash('Rombongan baru berhasil ditambahkan.', 'success')
        # Arahkan ke halaman edit agar bisa langsung tambah bus
        return redirect(url_for('admin.edit_rombongan', id=new_rombongan.id))

    return render_template('form_rombongan.html', form=form, title="Tambah Rombongan Baru")


@admin_bp.route('/rombongan/edit/<int:id>', methods=['GET', 'POST'])
@login_required
@role_required('Korpus', 'Sekretaris', 'Korpuspi')
def edit_rombongan(id):
    """Route untuk edit rombongan dengan manajemen bus"""
    active_edisi = get_active_edisi()
    rombongan = Rombongan.query.get_or_404(id)

    # Validasi akses rombongan
    if active_edisi and rombongan.edisi_id != active_edisi.id:
        flash("Anda tidak bisa mengedit rombongan dari edisi yang sudah selesai.", "danger")
        return redirect(url_for('admin.manajemen_rombongan'))
    
    if current_user.role.name == 'Korda' and rombongan not in current_user.active_managed_rombongan:
        abort(403)

    # Buat form untuk rombongan dan bus
    form = RombonganForm(obj=rombongan)
    bus_form = BusForm()

    if form.validate_on_submit():
        try:
            # Update semua field dari form ke objek rombongan
            rombongan.nama_rombongan = form.nama_rombongan.data
            rombongan.penanggung_jawab_putra = form.penanggung_jawab_putra.data
            rombongan.kontak_person_putra = form.kontak_person_putra.data
            rombongan.penanggung_jawab_putri = form.penanggung_jawab_putri.data
            rombongan.kontak_person_putri = form.kontak_person_putri.data
            rombongan.nomor_rekening = form.nomor_rekening.data

            
            # Handle cakupan_wilayah - parse JSON string
            try:
                cakupan_wilayah_data = form.cakupan_wilayah.data
                if cakupan_wilayah_data:
                    rombongan.cakupan_wilayah = json.loads(form.cakupan_wilayah.data)
                else:
                    rombongan.cakupan_wilayah = []
            except json.JSONDecodeError:
                flash("Format cakupan wilayah tidak valid. Gunakan format JSON yang benar.", "danger")
                return render_template('edit_rombongan.html', 
                                     form=form, 
                                     bus_form=bus_form,
                                     title="Edit Rombongan", 
                                     rombongan=rombongan)
            
            rombongan.jadwal_pulang = form.jadwal_pulang.data
            rombongan.batas_pembayaran_pulang = form.batas_pembayaran_pulang.data
            rombongan.jadwal_berangkat = form.jadwal_berangkat.data
            rombongan.batas_pembayaran_berangkat = form.batas_pembayaran_berangkat.data
            rombongan.titik_jemput_berangkat = form.titik_jemput_berangkat.data
            
            # Update tarif - hapus tarif lama dan ganti dengan yang baru
            for tarif in rombongan.tarifs:
                db.session.delete(tarif)

            for tarif_data in form.tarifs.data:
                if tarif_data['titik_turun'] and tarif_data['harga_bus'] is not None:
                    new_tarif = Tarif(
                        titik_turun=tarif_data['titik_turun'],
                        harga_bus=tarif_data['harga_bus'],
                        fee_korda=tarif_data.get('fee_korda', 0),
                        rombongan=rombongan
                    )
                    rombongan.tarifs.append(new_tarif)
            
            # Log aktivitas
            log_activity('Edit', 'Rombongan', f"Mengubah detail rombongan: '{rombongan.nama_rombongan}'")
            
            db.session.commit()
            flash('Data rombongan berhasil diperbarui!', 'success')
            return redirect(url_for('admin.edit_rombongan', id=rombongan.id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Terjadi kesalahan saat menyimpan data: {str(e)}', 'danger')

    # Isi data untuk request GET (pertama kali halaman dibuka)
    if request.method == 'GET':
        # Set cakupan_wilayah sebagai JSON string untuk tampilan di form
        form.cakupan_wilayah.data = json.dumps(rombongan.cakupan_wilayah or [], ensure_ascii=False, indent=2)
        
        # Clear existing tarif entries and add current ones
        while len(form.tarifs) > 0:
            form.tarifs.pop_entry()
        
        for tarif in rombongan.tarifs:
            tarif_form = form.tarifs.append_entry()
            tarif_form.titik_turun.data = tarif.titik_turun
            tarif_form.harga_bus.data = tarif.harga_bus
            tarif_form.fee_korda.data = tarif.fee_korda

        # Jika tidak ada tarif, tambahkan satu baris kosong
        if not rombongan.tarifs:
            form.tarifs.append_entry()

    return render_template('edit_rombongan.html', 
                         form=form, 
                         bus_form=bus_form,
                         title="Edit Rombongan", 
                         rombongan=rombongan)

@admin_bp.route('/rombongan/hapus/<int:id>', methods=['POST'])
@login_required
@role_required('Korpus', 'Sekretaris', 'Korpuspi') # Hanya Korpus yang bisa buat rombongan baru
def hapus_rombongan(id):
    rombongan = Rombongan.query.get_or_404(id)
    nama_rombongan = rombongan.nama_rombongan
    log_activity('Hapus', 'Rombongan', f"Menghapus rombongan: '{nama_rombongan}'")
    db.session.delete(rombongan)
    db.session.commit()
    flash('Rombongan berhasil dihapus.', 'info')
    return redirect(url_for('admin.dashboard'))

@admin_bp.route('/api/search-wilayah')
def search_wilayah_proxy():
    # Ambil query pencarian dari parameter URL (contoh: /api/search-wilayah?q=tegal)
    query = request.args.get('q', '')
    if not query:
        return jsonify({"results": []})

    try:
        # Panggil API sebenarnya dari backend
        api_url = f"https://backapp.amtsilatipusat.com/api/regencies?name={query}"
        response = requests.get(api_url, timeout=5)
        response.raise_for_status()  # Cek jika ada error HTTP
        
        # Kembalikan hasil JSON ke frontend kita
        return jsonify(response.json())
        
    except requests.exceptions.RequestException as e:
        # Tangani jika ada error koneksi ke API
        print(f"Error fetching API: {e}")
        return jsonify({"error": "Gagal mengambil data dari API"}), 500
    
    
@admin_bp.errorhandler(403)
def forbidden(error):
    return render_template('errors/403.html'), 403

@admin_bp.route('/santri')
@login_required
@role_required('Korpus', 'Korwil', 'Korda', 'Keamanan', 'Sekretaris', 'Korpuspi')
def manajemen_santri():
    active_edisi = get_active_edisi()
    page = request.args.get('page', 1, type=int)
    per_page = 20

    # Ambil data untuk dropdown filter
    all_rombongan = Rombongan.query.filter_by(edisi_id=active_edisi.id).order_by(Rombongan.nama_rombongan).all()
    all_kabupaten = [k[0] for k in db.session.query(Santri.kabupaten).filter(Santri.kabupaten.isnot(None)).distinct().order_by(Santri.kabupaten)]
    all_provinsi = [p[0] for p in db.session.query(Santri.provinsi).filter(Santri.provinsi.isnot(None)).distinct().order_by(Santri.provinsi)]

    # Ambil parameter filter dari URL
    nama_filter = request.args.get('nama', '')
    alamat_filter = request.args.getlist('alamat')
    keterangan_filter = request.args.get('keterangan', '')
    status_filter = request.args.get('status', '')
    rombongan_id_filter = request.args.get('rombongan_id', '')
    provinsi_filter = request.args.get('provinsi', '')

    # Query dasar
    query = Santri.query

    # Terapkan filter
    if nama_filter:
        query = query.filter(Santri.nama.ilike(f'%{nama_filter}%'))
    if alamat_filter:
        query = query.filter(Santri.kabupaten.in_(alamat_filter))
    if provinsi_filter:
        query = query.filter(Santri.provinsi == provinsi_filter)
    if status_filter:
        query = query.filter(Santri.status_santri == status_filter)
    if keterangan_filter:
        if keterangan_filter == 'Pengurus':
            query = query.filter(Santri.nama_jabatan.isnot(None))
        elif keterangan_filter == 'Santri':
            query = query.filter(Santri.nama_jabatan.is_(None))
    
    if rombongan_id_filter:
        if rombongan_id_filter == 'belum_terdaftar':
            query = query.filter(~Santri.pendaftarans.any(Pendaftaran.edisi_id == active_edisi.id))
        else:
            rombongan_id_int = int(rombongan_id_filter)
            query = query.join(Santri.pendaftarans).filter(
                Pendaftaran.edisi_id == active_edisi.id,
                or_(
                    Pendaftaran.rombongan_pulang_id == rombongan_id_int,
                    Pendaftaran.rombongan_kembali_id == rombongan_id_int
                )
            )

    # Lakukan paginasi
    pagination = query.order_by(Santri.nama).paginate(page=page, per_page=per_page, error_out=False)
    
    # Ambil data pendaftaran yang relevan untuk halaman saat ini agar efisien
    santri_ids_on_page = [s.id for s in pagination.items]
    pendaftaran_terkait = {}
    if santri_ids_on_page:
        pendaftarans = Pendaftaran.query.options(
            joinedload(Pendaftaran.rombongan_pulang),
            joinedload(Pendaftaran.rombongan_kembali)
        ).filter(
            Pendaftaran.edisi_id == active_edisi.id,
            Pendaftaran.santri_id.in_(santri_ids_on_page)
        ).all()
        pendaftaran_terkait = {p.santri_id: p for p in pendaftarans}

    return render_template('manajemen_santri.html',
                           pagination=pagination,
                           all_rombongan=all_rombongan,
                           all_kabupaten=all_kabupaten,
                           all_provinsi=all_provinsi,
                           pendaftaran_terkait=pendaftaran_terkait)

@admin_bp.route('/api/search-student')
def search_student_proxy():
    query = request.args.get('q', '')
    if not query:
        return jsonify({"data": [], "success": False})

    try:
        # Panggil API sebenarnya dari backend. Tambahkan limit untuk mendapatkan lebih banyak hasil.
        api_url = f"https://sigap.amtsilatipusat.com/api/student?search={query}&limit=20"
        response = requests.get(api_url, timeout=10)
        response.raise_for_status()
        return jsonify(response.json())
        
    except requests.exceptions.RequestException as e:
        print(f"Error fetching student API: {e}")
        return jsonify({"error": "Gagal mengambil data dari API"}), 500


@admin_bp.route('/santri/impor', methods=['POST'])
@login_required
@role_required('Korpus', 'Sekretaris', 'Korpuspi') # Menambahkan Korpuspi untuk konsistensi
def impor_santri():
    data = request.get_json()
    if not data or not data.get('id'):
        return jsonify({"success": False, "message": "Data tidak valid dari API."}), 400

    # Cek apakah santri dengan api_student_id ini sudah ada di database lokal
    santri = Santri.query.filter_by(api_student_id=str(data['id'])).first()

    # Logika baru: Jika santri sudah ada, perbarui datanya. Jika tidak, buat baru.
    if santri:
        # --- BLOK UNTUK MEMPERBARUI DATA ---
        action = 'Edit'
        message = f"Data {santri.nama} berhasil diperbarui."
        
        # Timpa field yang ada dengan data baru dari API
        santri.nis = data.get('nis', santri.nis)
        santri.nama = data.get('name', santri.nama)
        santri.kabupaten = data.get('regency', santri.kabupaten)
        santri.asrama = data.get('activeDormitory', santri.asrama)
        santri.no_hp_wali = data.get('parrentPhone', santri.no_hp_wali)
        santri.jenis_kelamin = data.get('gender') or santri.jenis_kelamin or 'PUTRA'
        
    else:
        # --- BLOK UNTUK MEMBUAT DATA BARU (LOGIKA LAMA) ---
        action = 'Tambah'
        message = f"{data.get('name', 'Santri baru')} berhasil diimpor."
        
        santri = Santri(
            api_student_id=str(data['id']),
            nis=data.get('nis', 'N/A'),
            nama=data.get('name', 'Tanpa Nama'),
            kabupaten=data.get('regency'),
            asrama=data.get('activeDormitory'),
            no_hp_wali=data.get('parrentPhone'),
            jenis_kelamin=data.get('gender') or 'PUTRA'
        )
        db.session.add(santri)

    try:
        db.session.commit()
        log_activity(action, 'Santri', f"Impor/Update individual: '{santri.nama}' (ID API: {santri.api_student_id})")
        return jsonify({"success": True, "message": message})
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "message": f"Terjadi kesalahan database: {e}"}), 500

@admin_bp.route('/santri/impor-semua', methods=['POST'])
@login_required
@role_required('Korpus', 'Sekretaris')
def impor_semua_santri():
    try:
        api_url = "https://sigap.amtsilatipusat.com/api/student?limit=2000"
        response = requests.get(api_url, timeout=60)
        response.raise_for_status()
        santri_from_api = response.json().get('data', [])

        # 1. Ambil semua santri yang ada berdasarkan NIS sebagai patokan
        existing_santri_map_by_nis = {s.nis: s for s in Santri.query.all() if s.nis}
        
        updated_count = 0
        new_count = 0
        
        to_create_mappings = []
        
        for data in santri_from_api:
            api_nis = data.get('nis')
            if not api_nis:
                continue # Lewati data dari API yang tidak memiliki NIS

            leadership_data = data.get('leadership')
            
            # 2. Cek apakah santri sudah ada di database lokal berdasarkan NIS
            if api_nis in existing_santri_map_by_nis:
                # Jika ADA, update datanya
                santri_to_update = existing_santri_map_by_nis[api_nis]
                
                santri_to_update.api_student_id = data.get('id') # Sinkronkan ulang ID API
                santri_to_update.nama = data.get('name', 'Tanpa Nama')
                santri_to_update.kabupaten = data.get('regency')
                santri_to_update.provinsi = data.get('provinnce') # <-- Tambahkan baris ini
                santri_to_update.asrama = data.get('activeDormitory')
                santri_to_update.no_hp_wali = data.get('parrentPhone')
                santri_to_update.jenis_kelamin = data.get('gender')
                santri_to_update.kelas_formal = data.get('formalClass')
                santri_to_update.kelas_ngaji = data.get('activeClass')
                santri_to_update.nama_jabatan = leadership_data.get('name') if leadership_data else None
                santri_to_update.status_jabatan = leadership_data.get('status') if leadership_data else None
                
                updated_count += 1
            else:
                # Jika TIDAK ADA, siapkan untuk dibuat sebagai data baru
                to_create_mappings.append({
                    'api_student_id': data.get('id'),
                    'nis': api_nis,
                    'nama': data.get('name', 'Tanpa Nama'),
                    'kabupaten': data.get('regency'),
                    'provinsi': data.get('provinnce'),
                    'asrama': data.get('activeDormitory'),
                    'no_hp_wali': data.get('parrentPhone'),
                    'jenis_kelamin': data.get('gender'),
                    'kelas_formal': data.get('formalClass'),
                    'kelas_ngaji': data.get('activeClass'),
                    'nama_jabatan': leadership_data.get('name') if leadership_data else None,
                    'status_jabatan': leadership_data.get('status') if leadership_data else None,
                })

        # Proses pembuatan data baru secara massal
        if to_create_mappings:
            db.session.bulk_insert_mappings(Santri, to_create_mappings)
            new_count = len(to_create_mappings)

        db.session.commit()
        log_activity('Sinkronisasi', 'Santri', f"Sinkronisasi massal: {new_count} baru, {updated_count} diperbarui.")
        flash(f"Proses selesai! Berhasil mengimpor {new_count} santri baru dan memperbarui {updated_count} data santri.", "success")

    except requests.exceptions.RequestException as e:
        flash(f"Gagal mengambil data dari API Induk: {e}", "danger")
    except Exception as e:
        db.session.rollback()
        flash(f"Terjadi error saat memproses data. Cek terminal untuk detail.", "danger")
        print("=====================================")
        print(f"Error detail impor santri: {e}")
        print("=====================================")

    return redirect(url_for('admin.manajemen_santri'))
                            
@admin_bp.route('/santri/edit/<int:id>', methods=['GET', 'POST'])
def edit_santri(id):
    # Fungsi ini sekarang hanya untuk mengedit data minor, pendaftaran dipindah
    # Anda bisa kembangkan ini nanti jika perlu edit data snapshot
    santri = Santri.query.get_or_404(id)
    flash("Fungsi edit santri akan dikembangkan lebih lanjut.", "info")
    return redirect(url_for('admin.manajemen_santri'))

@admin_bp.route('/santri/hapus/<int:id>', methods=['POST'])
@login_required
@role_required('Korpus', 'Korpuspi', 'Sekretaris') # Hanya Korpus yang bisa buat rombongan baru
def hapus_santri(id):
    santri = Santri.query.get_or_404(id)
    nama_santri = santri.nama
    log_activity('Hapus', 'Santri', f"Menghapus data santri: '{nama_santri}' (ID: {santri.id})")
    db.session.delete(santri)
    db.session.commit()
    flash(f"Santri '{nama_santri}' berhasil dihapus dari sistem.", "info")
    return redirect(url_for('admin.manajemen_santri'))

@admin_bp.route('/pendaftaran-rombongan', methods=['GET', 'POST'])
@login_required
@role_required('Korpus', 'Korda', 'Korpuspi')
def pendaftaran_rombongan():
    form = PendaftaranForm()
    active_edisi = get_active_edisi()
    if not active_edisi:
        flash("Tidak ada edisi yang aktif.", "warning")
        return redirect(url_for('admin.dashboard'))

    # --- BAGIAN PERBAIKAN ---
    # 1. Isi pilihan Rombongan di luar kondisi GET/POST agar selalu terisi
    if current_user.role.name == 'Korda':
        form.rombongan.choices = [(r.id, r.nama_rombongan) for r in current_user.active_managed_rombongan]
    else:
        form.rombongan.choices = [(r.id, r.nama_rombongan) for r in Rombongan.query.filter_by(edisi_id=active_edisi.id).order_by(Rombongan.nama_rombongan).all()]

    # 2. Jika ada data rombongan yang dikirim (baik GET maupun POST), isi pilihan Titik Turun dan Bus
    selected_rombongan_id = form.rombongan.data
    if request.method == 'POST' or request.args.get('rombongan_id'):
        if not selected_rombongan_id:
             selected_rombongan_id = request.args.get('rombongan_id') # Ambil dari URL jika ada
        
        if selected_rombongan_id:
            rombongan = Rombongan.query.get(int(selected_rombongan_id))
            if rombongan:
                form.titik_turun.choices = [(t.titik_turun, t.titik_turun) for t in rombongan.tarifs]
                
    # --- AKHIR PERBAIKAN ---

    if form.validate_on_submit():
        # Ambil daftar NIS dari form (sekarang berupa string dipisah koma)
        nis_list_str = form.santri_list.data
        nis_list = [nis.strip() for nis in nis_list_str.split(',') if nis.strip()]

        if not nis_list:
            flash('Tidak ada santri yang dipilih.', 'danger')
            return redirect(url_for('admin.pendaftaran_rombongan'))

        santri_terpilih = Santri.query.filter(Santri.nis.in_(nis_list)).all()
        rombongan_id = form.rombongan.data
        rombongan_obj = Rombongan.query.get(rombongan_id)
        
        berhasil_didaftarkan = 0
        nama_santri_berhasil = []  # TAMBAHAN: List untuk menyimpan nama santri yang berhasil
        sudah_terdaftar_sebelumnya = []
        gagal_karena_izin = []
        pendaftaran_baru_list = []

        for santri in santri_terpilih:
            existing_pendaftaran = Pendaftaran.query.filter_by(santri_id=santri.id, edisi_id=active_edisi.id).first()
            if existing_pendaftaran:
                sudah_terdaftar_sebelumnya.append(santri.nama)
                continue
            if santri.status_santri == 'Izin':
                gagal_karena_izin.append(santri.nama)
                continue

            tarif = Tarif.query.filter_by(rombongan_id=rombongan_id, titik_turun=form.titik_turun.data).first()
            total_biaya = 0
            if tarif:
                biaya_per_perjalanan = tarif.harga_bus + tarif.fee_korda + 10000
                if form.status_pulang.data != 'Tidak Ikut': total_biaya += biaya_per_perjalanan
                if form.status_kembali.data != 'Tidak Ikut': total_biaya += biaya_per_perjalanan
            
            # Buat dictionary untuk menampung data pendaftaran
            pendaftaran_data = {
                'edisi_id': active_edisi.id,
                'santri_id': santri.id,
                'rombongan_pulang_id': rombongan_id,
                'status_pulang': form.status_pulang.data,
                'metode_pembayaran_pulang': form.metode_pembayaran_pulang.data,
                'titik_turun': form.titik_turun.data,
                'status_kembali': form.status_kembali.data,
                'metode_pembayaran_kembali': form.metode_pembayaran_kembali.data,
                'gelombang_pulang': form.gelombang_pulang.data,
                'total_biaya': total_biaya
            }

            # Salin data pulang ke kembali HANYA JIKA santri ikut perjalanan kembali
            if form.status_kembali.data != 'Tidak Ikut':
                pendaftaran_data['rombongan_kembali_id'] = rombongan_id
                pendaftaran_data['titik_jemput_kembali'] = form.titik_turun.data

            new_pendaftaran = Pendaftaran(**pendaftaran_data)
            pendaftaran_baru_list.append(new_pendaftaran)
            berhasil_didaftarkan += 1
            nama_santri_berhasil.append(santri.nama)  # TAMBAHAN: Simpan nama santri yang berhasil
        
        if pendaftaran_baru_list:
            db.session.add_all(pendaftaran_baru_list)
            db.session.commit()
            # MODIFIKASI: Log activity dengan semua nama santri untuk transparansi
            if berhasil_didaftarkan == 1:
                log_activity('Tambah', 'Pendaftaran', f"Mendaftarkan santri {nama_santri_berhasil[0]} ke rombongan '{rombongan_obj.nama_rombongan}'")
            else:
                nama_str_log = ', '.join(nama_santri_berhasil)
                log_activity('Tambah', 'Pendaftaran', f"Mendaftarkan {berhasil_didaftarkan} santri ke rombongan '{rombongan_obj.nama_rombongan}': {nama_str_log}")

        # MODIFIKASI: Flash message dengan semua nama santri
        if berhasil_didaftarkan > 0:
            if berhasil_didaftarkan == 1:
                flash(f'Santri {nama_santri_berhasil[0]} udah berhasil didaftarin yaa.', 'success')
            else:
                # Tampilkan semua nama dengan format yang rapi
                nama_str = ', '.join(nama_santri_berhasil[:-1]) + f' dan {nama_santri_berhasil[-1]}'
                flash(f'{berhasil_didaftarkan} santri berhasil didaftarkan: {nama_str}.', 'success')
        
        if sudah_terdaftar_sebelumnya: 
            flash(f'Santri berikut dilewati karena sudah terdaftar: {", ".join(sudah_terdaftar_sebelumnya)}', 'warning')
        if gagal_karena_izin: 
            flash(f'Santri berikut dilewati karena berstatus Izin: {", ".join(gagal_karena_izin)}', 'danger')

        return redirect(url_for('admin.pendaftaran_rombongan', rombongan_id=rombongan_id))

    rombongan_id_url = request.args.get('rombongan_id')
    if rombongan_id_url:
        form.rombongan.data = int(rombongan_id_url)

    return render_template('pendaftaran_rombongan.html', form=form, title="Form Pendaftaran Rombongan")

# Di dalam file app/admin/routes.py


@admin_bp.route('/pendaftaran/<int:pendaftaran_id>/edit', methods=['GET', 'POST'])
@login_required
@role_required('Korpus', 'Korwil', 'Korda', 'Bendahara Pusat', 'Sekretaris', 'Korpuspi')
def edit_pendaftaran(pendaftaran_id):
    pendaftaran = Pendaftaran.query.get_or_404(pendaftaran_id)
    
    # --- PERBAIKAN UTAMA ADA DI SINI ---
    # Kita pisahkan inisialisasi form menggunakan if/else standar
    if request.method == 'POST':
        # Saat submit, form diisi dari data request yang masuk
        form = PendaftaranEditForm(request.form)
    else:
        # Saat halaman dibuka (GET), form diisi dari objek pendaftaran dari database
        form = PendaftaranEditForm(obj=pendaftaran)
    # --- AKHIR PERBAIKAN ---

    # --- PENENTUAN HAK AKSES ---
    # (Logika ini tetap sama)
    can_edit_pulang = current_user.role.name in ['Korpus', 'Korwil', 'Bendahara Pusat', 'Sekretaris', 'Korpuspi']
    can_edit_kembali = current_user.role.name in ['Korpus', 'Korwil', 'Bendahara Pusat', 'Sekretaris', 'Korpuspi']

    if current_user.role.name == 'Korda':
        managed_ids = {r.id for r in current_user.active_managed_rombongan}
        if pendaftaran.rombongan_pulang_id in managed_ids:
            can_edit_pulang = True
        if pendaftaran.rombongan_pulang_id in managed_ids or (pendaftaran.rombongan_kembali_id and pendaftaran.rombongan_kembali_id in managed_ids):
            can_edit_kembali = True
            
    if not can_edit_pulang and not can_edit_kembali:
        flash('Anda tidak memiliki hak akses untuk mengedit pendaftaran ini.', 'danger')
        return redirect(url_for('admin.daftar_peserta_global'))

    # --- PENGISIAN DROPDOWN SEBELUM VALIDASI (KUNCI PERBAIKAN) ---
    # 1. Mengisi pilihan untuk dropdown bagian PULANG
    if pendaftaran.rombongan_pulang:
        form.titik_turun.choices = [(t.titik_turun, t.titik_turun) for t in pendaftaran.rombongan_pulang.tarifs]
        form.bus_pulang.choices = [("", "-- Pilih Bus --")] + [(b.id, f"{b.nama_armada} - {b.nomor_lambung or b.plat_nomor}") for b in pendaftaran.rombongan_pulang.buses]
    
    # 2. Mengisi pilihan untuk dropdown bagian KEMBALI
    rombongan_kembali_obj = form.rombongan_kembali.data or pendaftaran.rombongan_kembali or pendaftaran.rombongan_pulang
    if rombongan_kembali_obj:
        form.titik_jemput_kembali.choices = [(t.titik_turun, t.titik_turun) for t in rombongan_kembali_obj.tarifs]
        form.bus_kembali.choices = [("", "-- Pilih Bus --")] + [(b.id, f"{b.nama_armada} - {b.nomor_lambung or b.plat_nomor}") for b in rombongan_kembali_obj.buses]
    # --- AKHIR BAGIAN PENTING ---

    if form.validate_on_submit():
        # Hanya simpan data yang berhak diubah oleh user
        if can_edit_pulang:
            pendaftaran.status_pulang = form.status_pulang.data
            pendaftaran.metode_pembayaran_pulang = form.metode_pembayaran_pulang.data
            pendaftaran.titik_turun = form.titik_turun.data
            pendaftaran.gelombang_pulang = form.gelombang_pulang.data # <-- Tambahkan ini
            pendaftaran.bus_pulang_id = form.bus_pulang.data or None

        if can_edit_kembali:
            pendaftaran.rombongan_kembali = form.rombongan_kembali.data
            pendaftaran.status_kembali = form.status_kembali.data
            pendaftaran.metode_pembayaran_kembali = form.metode_pembayaran_kembali.data
            pendaftaran.titik_jemput_kembali = form.titik_jemput_kembali.data
            pendaftaran.bus_kembali_id = form.bus_kembali.data or None
        
        # Hitung ulang total biaya jika diperlukan (opsional, tapi disarankan)
        # Anda bisa menambahkan fungsi kalkulasi biaya di sini jika ada

        try:
            db.session.commit()
            flash('Perubahan data pendaftaran berhasil disimpan.', 'success')
            return redirect(url_for('admin.daftar_peserta_global', rombongan_id=pendaftaran.rombongan_pulang_id))
        except Exception as e:
            db.session.rollback()
            flash(f'Terjadi kesalahan saat menyimpan: {e}', 'danger')

    # Isi form dengan data dari DB saat request GET pertama kali
    if request.method == 'GET':
        form.gelombang_pulang.data = pendaftaran.gelombang_pulang # <-- Tambahkan ini
        form.bus_pulang.data = pendaftaran.bus_pulang_id
        form.bus_kembali.data = pendaftaran.bus_kembali_id
        # Field lain sudah terisi otomatis dari `obj=pendaftaran`

    # Siapkan data untuk JavaScript
    initial_data_dict = {
        'rombonganPulangId': pendaftaran.rombongan_pulang_id,
        'rombonganKembaliId': pendaftaran.rombongan_kembali_id,
        'busPulangId': pendaftaran.bus_pulang_id,
        'busKembaliId': pendaftaran.bus_kembali_id,
        'titikTurun': pendaftaran.titik_turun or '',
        'titikJemputKembali': pendaftaran.titik_jemput_kembali or ''
    }
    initial_data_json = json.dumps(initial_data_dict)

    return render_template('edit_pendaftaran.html', 
                           form=form, 
                           pendaftaran=pendaftaran, 
                           title="Edit Detail Pendaftaran",
                           can_edit_pulang=can_edit_pulang, 
                           can_edit_kembali=can_edit_kembali,
                           initial_data_json=initial_data_json)

@admin_bp.route('/pendaftaran/hapus/<int:pendaftaran_id>', methods=['POST'])
@login_required
@role_required('Korpus', 'Korda', 'Korpuspi')
def hapus_pendaftaran(pendaftaran_id):
    pendaftaran = Pendaftaran.query.get_or_404(pendaftaran_id)
    # Verifikasi Kepemilikan (menggunakan rombongan pulang sebagai acuan)
    if current_user.role.name == 'Korda':
        # Cek apakah Korda mengelola salah satu dari rombongan pendaftaran
        managed_ids = {r.id for r in current_user.active_managed_rombongan}
        if pendaftaran.rombongan_pulang_id not in managed_ids and pendaftaran.rombongan_kembali_id not in managed_ids:
            abort(403)

    # Simpan ID rombongan untuk redirect, utamakan rombongan pulang
    rombongan_id = pendaftaran.rombongan_pulang_id or pendaftaran.rombongan_kembali_id
    nama_santri = pendaftaran.santri.nama
    
    db.session.delete(pendaftaran)
    log_activity('Hapus', 'Pendaftaran', f"Menghapus pendaftaran untuk santri: '{nama_santri}'")
    db.session.commit()
    
    flash(f"Pendaftaran untuk '{nama_santri}' telah berhasil dihapus.", "info")
    return redirect(url_for('admin.daftar_peserta_global', rombongan_id=rombongan_id))

# Route API helper BARU untuk mengambil detail rombongan
@admin_bp.route('/api/rombongan-detail/<int:id>')
def rombongan_detail(id):
    rombongan = Rombongan.query.get_or_404(id)
    
    tarifs = [{'titik_turun': t.titik_turun, 'harga': t.harga_bus + t.fee_korda + 10000} for t in rombongan.tarifs]

    buses_data = []
    for bus in rombongan.buses:
        terisi_pulang = Pendaftaran.query.filter_by(bus_pulang_id=bus.id).count()
        terisi_kembali = Pendaftaran.query.filter_by(bus_kembali_id=bus.id).count()
        buses_data.append({
            'id': bus.id, 'nama_armada': bus.nama_armada,
            'nomor_lambung': bus.nomor_lambung, 'plat_nomor': bus.plat_nomor,
            'kuota': bus.kuota,
            'sisa_kuota_pulang': bus.kuota - terisi_pulang,
            'sisa_kuota_kembali': bus.kuota - terisi_kembali,
            'gmaps_share_url': bus.gmaps_share_url # <-- Pastikan ini ada

        })
    semua_pendaftar_objek = list(rombongan.pendaftar_pulang) + list(rombongan.pendaftar_kembali)
    unique_pendaftar_objek = {p.id: p for p in semua_pendaftar_objek}.values()

    pendaftar_data = [{'nama': p.santri.nama, 'nis': p.santri.nis, 'titik_turun': p.titik_turun, 'total_biaya': p.total_biaya} for p in unique_pendaftar_objek]

    return jsonify({
        'cakupan_wilayah': rombongan.cakupan_wilayah, 'tarifs': tarifs,
        'buses': buses_data, 'pendaftar': pendaftar_data
    })

# Di dalam app/admin/routes.py

@admin_bp.route('/api/search-santri')
@login_required
def api_search_santri():
    """
    API untuk mencari semua santri aktif, 
    dengan tambahan informasi apakah mereka sudah terdaftar atau belum.
    """
    active_edisi = get_active_edisi()
    if not active_edisi:
        return jsonify({'results': []})

    q = request.args.get('q', '')
    query_id = request.args.get('q_id')

    # Query dasar untuk semua santri aktif, tanpa filter status pendaftaran
    base_query = Santri.query.filter(Santri.status_santri.in_(['Aktif', 'Partisipan', 'Wisuda', 'Izin']))

    if q:
        base_query = base_query.filter(Santri.nama.ilike(f'%{q}%'))
    
    if query_id:
        base_query = Santri.query.filter_by(id=query_id)
    
    santri_list = base_query.limit(20).all()

    # Ambil ID santri yang ditemukan untuk cek pendaftaran
    santri_ids_on_page = [s.id for s in santri_list]

    # Buat map pendaftaran untuk efisiensi
    pendaftaran_map = {}
    if santri_ids_on_page:
        pendaftarans = Pendaftaran.query.options(
            joinedload(Pendaftaran.rombongan_pulang) # Eager load untuk ambil nama rombongan
        ).filter(
            Pendaftaran.edisi_id == active_edisi.id,
            Pendaftaran.santri_id.in_(santri_ids_on_page)
        ).all()
        pendaftaran_map = {p.santri_id: p for p in pendaftarans}

    # Format hasil JSON dengan menambahkan status pendaftaran
    results = []
    for s in santri_list:
        pendaftaran_info = pendaftaran_map.get(s.id)
        is_registered = pendaftaran_info is not None
        
        # Ambil nama rombongan jika sudah terdaftar
        rombongan_nama = ""
        if is_registered and pendaftaran_info.rombongan_pulang:
            rombongan_nama = pendaftaran_info.rombongan_pulang.nama_rombongan
        elif is_registered and pendaftaran_info.rombongan_kembali:
             rombongan_nama = pendaftaran_info.rombongan_kembali.nama_rombongan
        
        results.append({
            'id': s.id,
            'nis': s.nis,
            'nama': s.nama,
            'asrama': s.asrama,
            'kabupaten': s.kabupaten,
            'status_santri': s.status_santri,
            'is_registered': is_registered,
            'rombongan_nama': rombongan_nama
        })
    
    return jsonify({'results': results})
# Di dalam app/admin/routes.py

@admin_bp.route('/rombongan/<int:rombongan_id>/peserta')
@login_required
@role_required('Korpus', 'Korda', 'Korwil', 'Bendahara Pusat', 'Sekretaris', 'Korpuspi')
def daftar_peserta(rombongan_id):
    rombongan = Rombongan.query.get_or_404(rombongan_id)
    
    # Ambil semua pendaftar yang terkait dengan rombongan ini (baik pulang maupun kembali)
    pendaftar = Pendaftaran.query.filter(
        or_(
            Pendaftaran.rombongan_pulang_id == rombongan_id,
            Pendaftaran.rombongan_kembali_id == rombongan_id
        )
    ).options(
        joinedload(Pendaftaran.santri),
        joinedload(Pendaftaran.bus_pulang),
        joinedload(Pendaftaran.bus_kembali),
        joinedload(Pendaftaran.rombongan_pulang), # Pastikan rombongan pulang di-load
        joinedload(Pendaftaran.rombongan_kembali)
    ).all()

    # --- AWAL BLOK PERBAIKAN LOGIKA STATISTIK ---
    
    # Gunakan nama variabel yang lebih jelas
    stats = {
        'total_peserta': len(set(p.santri_id for p in pendaftar)), # Hitung santri unik
        'peserta_pulang': 0,    # Ganti dari 'sudah_bus_pulang'
        'peserta_kembali': 0,   # Ganti dari 'sudah_bus_kembali'
        'lunas_pulang': 0,
        'lunas_kembali': 0,
    }

    for p in pendaftar:
        # Cek partisipasi pulang (hanya jika rombongan pulangnya adalah rombongan ini)
        if p.rombongan_pulang_id == rombongan_id and p.status_pulang != 'Tidak Ikut':
            stats['peserta_pulang'] += 1
            if p.status_pulang == 'Lunas':
                stats['lunas_pulang'] += 1

        # Cek partisipasi kembali
        # Santri dihitung ikut kembali dengan rombongan ini JIKA:
        # 1. rombongan_kembali_id mereka adalah rombongan ini, ATAU
        # 2. rombongan_kembali_id mereka KOSONG, TAPI rombongan_pulang_id mereka adalah rombongan ini
        is_kembali_with_this_rombongan = (p.rombongan_kembali_id == rombongan_id) or \
                                         (p.rombongan_kembali_id is None and p.rombongan_pulang_id == rombongan_id)

        if is_kembali_with_this_rombongan and p.status_kembali != 'Tidak Ikut':
            stats['peserta_kembali'] += 1
            if p.status_kembali == 'Lunas':
                stats['lunas_kembali'] += 1

    # --- AKHIR BLOK PERBAIKAN ---

    return render_template('daftar_peserta.html', 
                           rombongan=rombongan, 
                           pendaftar=pendaftar,
                           stats=stats) # Kirim statistik yang sudah benar ke template

# Di dalam file app/admin/routes.py

# Di dalam app/admin/routes.py

@admin_bp.route('/peserta-global')
@login_required
@role_required('Korpus', 'Korwil', 'Korda', 'Bendahara Pusat', 'Sekretaris', 'Korpuspi')
def daftar_peserta_global():
    active_edisi = get_active_edisi()
    if not active_edisi:
        flash("Tidak ada edisi aktif.", "warning")
        return render_template('daftar_peserta_global.html', pagination=None, stats={}, all_rombongan=[])

    page = request.args.get('page', 1, type=int)
    per_page = 100
    nama_filter = request.args.get('nama', '')
    rombongan_id_filter = request.args.get('rombongan_id', type=int)
    status_bayar_filter = request.args.get('status_bayar', '')
    jenis_kelamin_filter = request.args.get('jenis_kelamin', '')
    perjalanan_filter = request.args.get('perjalanan', '')
    gelombang_filter = request.args.get('gelombang', type=int) # <-- BARIS BARU


    query = Pendaftaran.query.filter(Pendaftaran.edisi_id == active_edisi.id)

    if current_user.role.name in ['Korda', 'Korwil']:
        managed_rombongan_ids = [r.id for r in current_user.active_managed_rombongan]
        if managed_rombongan_ids:
            query = query.filter(or_(Pendaftaran.rombongan_pulang_id.in_(managed_rombongan_ids), Pendaftaran.rombongan_kembali_id.in_(managed_rombongan_ids)))
        else:
            query = query.filter(db.false())
    
    # Get all pendaftaran for stats calculation
    all_pendaftar_query = query
    pendaftar_for_stats = all_pendaftar_query.all()

    # ADDED: Create tarif map for biaya calculation
    tarif_map = {}
    for p in pendaftar_for_stats:
        # Tarif for pulang
        if p.rombongan_pulang_id and p.titik_turun:
            tarif_key = f"{p.rombongan_pulang_id}_{p.titik_turun}"
            if tarif_key not in tarif_map:
                tarif = Tarif.query.filter_by(
                    rombongan_id=p.rombongan_pulang_id, 
                    titik_turun=p.titik_turun
                ).first()
                if tarif:
                    tarif_map[tarif_key] = tarif
        
        # Tarif for kembali
        rombongan_kembali_id = p.rombongan_kembali_id or p.rombongan_pulang_id
        titik_jemput_kembali = p.titik_jemput_kembali or p.titik_turun
        if rombongan_kembali_id and titik_jemput_kembali:
            tarif_key = f"{rombongan_kembali_id}_{titik_jemput_kembali}"
            if tarif_key not in tarif_map:
                tarif = Tarif.query.filter_by(
                    rombongan_id=rombongan_kembali_id, 
                    titik_turun=titik_jemput_kembali
                ).first()
                if tarif:
                    tarif_map[tarif_key] = tarif

    # ADDED: Calculate biaya for each pendaftar and update stats
    total_biaya_lunas = 0
    total_biaya_belum_lunas = 0
    total_biaya_cash = 0
    total_biaya_transfer = 0
    count_cash = 0
    count_transfer = 0

    for p in pendaftar_for_stats:
        # Calculate biaya pulang
        biaya_pulang = 0
        if p.status_pulang and p.status_pulang != 'Tidak Ikut':
            if p.rombongan_pulang_id and p.titik_turun:
                tarif_key = f"{p.rombongan_pulang_id}_{p.titik_turun}"
                tarif = tarif_map.get(tarif_key)
                if tarif:
                    biaya_pulang = tarif.harga_bus + tarif.fee_korda + 10000
        
        # Calculate biaya kembali
        biaya_kembali = 0
        if p.status_kembali and p.status_kembali != 'Tidak Ikut':
            rombongan_kembali_id = p.rombongan_kembali_id or p.rombongan_pulang_id
            titik_jemput_kembali = p.titik_jemput_kembali or p.titik_turun
            if rombongan_kembali_id and titik_jemput_kembali:
                tarif_key = f"{rombongan_kembali_id}_{titik_jemput_kembali}"
                tarif = tarif_map.get(tarif_key)
                if tarif:
                    biaya_kembali = tarif.harga_bus + tarif.fee_korda + 10000
        
        # Set calculated biaya to the object for template use
        p.biaya_pulang_calculated = biaya_pulang
        p.biaya_kembali_calculated = biaya_kembali
        p.total_biaya_calculated = biaya_pulang + biaya_kembali
        
        # Update stats with calculated biaya
        if p.status_pulang == 'Lunas' and p.status_kembali == 'Lunas':
            total_biaya_lunas += p.total_biaya_calculated
        else:
            total_biaya_belum_lunas += p.total_biaya_calculated
        
        # Calculate payment method totals
        if ('Cash' in (str(p.metode_pembayaran_pulang or '') + str(p.metode_pembayaran_kembali or ''))):
            total_biaya_cash += p.total_biaya_calculated
            count_cash += 1
        if ('Transfer' in (str(p.metode_pembayaran_pulang or '') + str(p.metode_pembayaran_kembali or ''))):
            total_biaya_transfer += p.total_biaya_calculated
            count_transfer += 1

    # MODIFIED: Enhanced stats with biaya information
    stats = {
        'total_peserta': len(pendaftar_for_stats),
        'peserta_pulang': sum(1 for p in pendaftar_for_stats if p.status_pulang != 'Tidak Ikut'),
        'peserta_kembali': sum(1 for p in pendaftar_for_stats if p.status_kembali != 'Tidak Ikut'),
        'lunas_pulang': sum(1 for p in pendaftar_for_stats if p.status_pulang == 'Lunas'),
        'lunas_kembali': sum(1 for p in pendaftar_for_stats if p.status_kembali == 'Lunas'),
        'total_biaya_lunas': total_biaya_lunas,
        'total_biaya_belum_lunas': total_biaya_belum_lunas,
        'total_biaya_cash': total_biaya_cash,
        'total_biaya_transfer': total_biaya_transfer,
        'count_cash': count_cash,
        'count_transfer': count_transfer
    }

    # Apply filters after stats calculation
    query = query.join(Santri)

    if nama_filter:
        query = query.filter(Santri.nama.ilike(f'%{nama_filter}%'))
    if rombongan_id_filter:
        query = query.filter(or_(Pendaftaran.rombongan_pulang_id == rombongan_id_filter, Pendaftaran.rombongan_kembali_id == rombongan_id_filter))
    
    if jenis_kelamin_filter:
        query = query.filter(Santri.jenis_kelamin == jenis_kelamin_filter)
    if gelombang_filter:
        query = query.filter(Pendaftaran.gelombang_pulang == gelombang_filter)
        
    if perjalanan_filter == 'ikut_pulang':
        query = query.filter(Pendaftaran.status_pulang != 'Tidak Ikut')
    elif perjalanan_filter == 'tidak_ikut_pulang':
        query = query.filter(Pendaftaran.status_pulang == 'Tidak Ikut')
    elif perjalanan_filter == 'ikut_kembali':
        query = query.filter(Pendaftaran.status_kembali != 'Tidak Ikut')
    elif perjalanan_filter == 'tidak_ikut_kembali':
        query = query.filter(Pendaftaran.status_kembali == 'Tidak Ikut')

    if status_bayar_filter:
        if status_bayar_filter == 'Lunas':
            query = query.filter(and_(Pendaftaran.status_pulang == 'Lunas', Pendaftaran.status_kembali == 'Lunas'))
        else:
            query = query.filter(or_(Pendaftaran.status_pulang == 'Belum Bayar', Pendaftaran.status_kembali == 'Belum Bayar'))

    # Get paginated results
    pagination = query.order_by(Santri.nama).paginate(page=page, per_page=per_page, error_out=False)
    
    # ADDED: Calculate biaya for paginated results (for display in table)
    for pendaftar in pagination.items:
        # Calculate biaya pulang
        biaya_pulang = 0
        if pendaftar.status_pulang and pendaftar.status_pulang != 'Tidak Ikut':
            if pendaftar.rombongan_pulang_id and pendaftar.titik_turun:
                tarif_key = f"{pendaftar.rombongan_pulang_id}_{pendaftar.titik_turun}"
                tarif = tarif_map.get(tarif_key)
                if tarif:
                    biaya_pulang = tarif.harga_bus + tarif.fee_korda + 10000
        
        # Calculate biaya kembali
        biaya_kembali = 0
        if pendaftar.status_kembali and pendaftar.status_kembali != 'Tidak Ikut':
            rombongan_kembali_id = pendaftar.rombongan_kembali_id or pendaftar.rombongan_pulang_id
            titik_jemput_kembali = pendaftar.titik_jemput_kembali or pendaftar.titik_turun
            if rombongan_kembali_id and titik_jemput_kembali:
                tarif_key = f"{rombongan_kembali_id}_{titik_jemput_kembali}"
                tarif = tarif_map.get(tarif_key)
                if tarif:
                    biaya_kembali = tarif.harga_bus + tarif.fee_korda + 10000
        
        # Set calculated biaya to the object for template display
        pendaftar.biaya_pulang_calculated = biaya_pulang
        pendaftar.biaya_kembali_calculated = biaya_kembali
        pendaftar.total_biaya_calculated = biaya_pulang + biaya_kembali

    all_rombongan = Rombongan.query.filter_by(edisi=active_edisi).order_by(Rombongan.nama_rombongan).all()

    return render_template('daftar_peserta_global.html',
                           pagination=pagination,
                           stats=stats,
                           all_rombongan=all_rombongan)

# Di dalam app/admin/routes.py

# Di dalam app/admin/routes.py

# Export untuk Data Global (Pulang + Kembali)
@admin_bp.route('/export-peserta-global')
@login_required
@role_required('Korpus', 'Korwil', 'Korda', 'Bendahara Pusat', 'Sekretaris', 'Korpuspi')
def export_peserta_global():
    active_edisi = get_active_edisi()
    query = Pendaftaran.query.filter(Pendaftaran.edisi_id == active_edisi.id)

    if current_user.role.name in ['Korda', 'Korwil']:
        managed_rombongan_ids = [r.id for r in current_user.active_managed_rombongan]
        if managed_rombongan_ids:
            query = query.filter(or_(Pendaftaran.rombongan_pulang_id.in_(managed_rombongan_ids), Pendaftaran.rombongan_kembali_id.in_(managed_rombongan_ids)))
        else:
            query = query.filter(db.false())

    # Ambil semua filter dari URL, termasuk filter gelombang
    nama_filter = request.args.get('nama', '')
    rombongan_id_filter = request.args.get('rombongan_id', type=int)
    status_bayar_filter = request.args.get('status_bayar', '')
    jenis_kelamin_filter = request.args.get('jenis_kelamin', '')
    gelombang_filter = request.args.get('gelombang', type=int)  # Filter gelombang

    query = query.join(Santri)
    
    # Terapkan semua filter yang ada
    if nama_filter:
        query = query.filter(Santri.nama.ilike(f'%{nama_filter}%'))
    if rombongan_id_filter:
        query = query.filter(or_(Pendaftaran.rombongan_pulang_id == rombongan_id_filter, Pendaftaran.rombongan_kembali_id == rombongan_id_filter))
    if jenis_kelamin_filter:
        query = query.filter(Santri.jenis_kelamin == jenis_kelamin_filter)
    if status_bayar_filter:
        if status_bayar_filter == 'Lunas':
            query = query.filter(and_(Pendaftaran.status_pulang == 'Lunas', Pendaftaran.status_kembali == 'Lunas'))
        else:
            query = query.filter(or_(Pendaftaran.status_pulang == 'Belum Bayar', Pendaftaran.status_kembali == 'Belum Bayar'))
    
    # Filter gelombang
    if gelombang_filter:
        query = query.filter(Pendaftaran.gelombang_pulang == gelombang_filter)
    
    semua_pendaftar = query.order_by(Santri.nama).all()

    # Create tarif map for calculation
    tarif_map = {}
    for p in semua_pendaftar:
        # Tarif for pulang
        if p.rombongan_pulang_id and p.titik_turun:
            tarif_key = f"{p.rombongan_pulang_id}_{p.titik_turun}"
            if tarif_key not in tarif_map:
                tarif = Tarif.query.filter_by(
                    rombongan_id=p.rombongan_pulang_id, 
                    titik_turun=p.titik_turun
                ).first()
                if tarif:
                    tarif_map[tarif_key] = tarif
        
        # Tarif for kembali
        rombongan_kembali_id = p.rombongan_kembali_id or p.rombongan_pulang_id
        titik_jemput_kembali = p.titik_jemput_kembali or p.titik_turun
        if rombongan_kembali_id and titik_jemput_kembali:
            tarif_key = f"{rombongan_kembali_id}_{titik_jemput_kembali}"
            if tarif_key not in tarif_map:
                tarif = Tarif.query.filter_by(
                    rombongan_id=rombongan_kembali_id, 
                    titik_turun=titik_jemput_kembali
                ).first()
                if tarif:
                    tarif_map[tarif_key] = tarif

    # Calculate biaya using tarif data
    total_lunas = 0
    total_belum_lunas = 0
    cash_data = []
    transfer_data = []

    for p in semua_pendaftar:
        # Calculate biaya pulang
        biaya_pulang = 0
        if p.status_pulang and p.status_pulang != 'Tidak Ikut':
            if p.rombongan_pulang_id and p.titik_turun:
                tarif_key = f"{p.rombongan_pulang_id}_{p.titik_turun}"
                tarif = tarif_map.get(tarif_key)
                if tarif:
                    biaya_pulang = tarif.harga_bus + tarif.fee_korda + 10000
        
        # Calculate biaya kembali
        biaya_kembali = 0
        if p.status_kembali and p.status_kembali != 'Tidak Ikut':
            rombongan_kembali_id = p.rombongan_kembali_id or p.rombongan_pulang_id
            titik_jemput_kembali = p.titik_jemput_kembali or p.titik_turun
            if rombongan_kembali_id and titik_jemput_kembali:
                tarif_key = f"{rombongan_kembali_id}_{titik_jemput_kembali}"
                tarif = tarif_map.get(tarif_key)
                if tarif:
                    biaya_kembali = tarif.harga_bus + tarif.fee_korda + 10000
        
        # Set calculated biaya to the object for later use
        p.biaya_pulang_calculated = biaya_pulang
        p.biaya_kembali_calculated = biaya_kembali
        p.total_biaya_calculated = biaya_pulang + biaya_kembali
        
        # Calculate totals
        if p.status_pulang == 'Lunas' and p.status_kembali == 'Lunas':
            total_lunas += p.total_biaya_calculated
        else:
            total_belum_lunas += p.total_biaya_calculated
        
        # Track payment methods
        if ('Cash' in (str(p.metode_pembayaran_pulang or '') + str(p.metode_pembayaran_kembali or ''))):
            cash_data.append(p)
        if ('Transfer' in (str(p.metode_pembayaran_pulang or '') + str(p.metode_pembayaran_kembali or ''))):
            transfer_data.append(p)
    
    total_cash = sum(p.total_biaya_calculated for p in cash_data)
    total_transfer = sum(p.total_biaya_calculated for p in transfer_data)
    jumlah_orang_cash = len(cash_data)
    jumlah_orang_transfer = len(transfer_data)

    data_for_excel = []
    for index, p in enumerate(semua_pendaftar, 1):
        # Tentukan tanggal pulang berdasarkan gelombang
        tanggal_pulang = "26 September" if p.gelombang_pulang == 1 else "27 September" if p.gelombang_pulang == 2 else "-"
        
        data_for_excel.append({
            'No.': index,
            'Nama Santri': p.santri.nama if p.santri else 'N/A',
            'Jenis Kelamin': p.santri.jenis_kelamin if p.santri else 'N/A',
            'Tanggal Pulang': tanggal_pulang if p.status_pulang != 'Tidak Ikut' else 'Tidak Ikut',
            'Asrama': p.santri.asrama if p.santri else 'N/A',
            'Kabupaten': p.santri.kabupaten if p.santri else 'N/A',
            'Rombongan Pulang': p.rombongan_pulang.nama_rombongan if p.rombongan_pulang else 'N/A',
            'Rombongan Kembali': p.rombongan_kembali.nama_rombongan if p.rombongan_kembali else (p.rombongan_pulang.nama_rombongan if p.rombongan_pulang else 'N/A'),
            'Titik Turun': p.titik_turun or '-',
            'Titik Jemput': p.titik_jemput_kembali or p.titik_turun or '-',
            'Bus Pulang': p.bus_pulang.nama_armada if p.bus_pulang else '-',
            'Bus Kembali': p.bus_kembali.nama_armada if p.bus_kembali else '-',
            'Status Pulang': p.status_pulang,
            'Metode Bayar Pulang': p.metode_pembayaran_pulang or '-',
            'Status Kembali': p.status_kembali,
            'Metode Bayar Kembali': p.metode_pembayaran_kembali or '-',
            'Biaya Pulang': f"Rp. {p.biaya_pulang_calculated:,.0f}" if p.biaya_pulang_calculated else "Rp. 0",
            'Biaya Kembali': f"Rp. {p.biaya_kembali_calculated:,.0f}" if p.biaya_kembali_calculated else "Rp. 0",
            'Total Biaya': f"Rp. {p.total_biaya_calculated:,.0f}" if p.total_biaya_calculated else "Rp. 0"
        })

    if not data_for_excel:
        flash('Tidak ada data untuk diekspor sesuai filter yang dipilih.', 'warning')
        return redirect(url_for('admin.daftar_peserta_global', **request.args))

    # Create Excel with styling
    df = pd.DataFrame(data_for_excel)
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Write main data
        df.to_excel(writer, index=False, sheet_name='Data Peserta Global', startrow=0)
        
        # Get workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Data Peserta Global']
        
        # Define styles
        from openpyxl.styles import PatternFill, Font, Alignment
        
        # Header styling - Blue background with white text
        header_fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        # Apply header styling
        for cell in worksheet[1]:  # First row (header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Add summary data at the bottom
        summary_start_row = len(df) + 3
        
        summary_data = [
            ['RINGKASAN PEMBAYARAN', ''],
            ['Total Lunas', f"Rp. {total_lunas:,.0f}"],
            ['Total Belum Lunas', f"Rp. {total_belum_lunas:,.0f}"],
            ['Total Pembayaran Cash', f"Rp. {total_cash:,.0f}"],
            ['Jumlah Orang Cash', jumlah_orang_cash],
            ['Total Pembayaran Transfer', f"Rp. {total_transfer:,.0f}"],
            ['Jumlah Orang Transfer', jumlah_orang_transfer],
            ['Total Keseluruhan', f"Rp. {(total_lunas + total_belum_lunas):,.0f}"]
        ]
        
        for idx, (label, value) in enumerate(summary_data):
            row_num = summary_start_row + idx
            worksheet[f'A{row_num}'] = label
            worksheet[f'B{row_num}'] = value
            
            # Style summary header
            if idx == 0:
                worksheet[f'A{row_num}'].fill = header_fill
                worksheet[f'A{row_num}'].font = header_font
                worksheet[f'B{row_num}'].fill = header_fill
                worksheet[f'B{row_num}'].font = header_font
            else:
                # Style other summary rows
                worksheet[f'A{row_num}'].font = Font(bold=True)

    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='Data_Peserta_Global.xlsx'
    )

# Export untuk Data Pulang
@admin_bp.route('/export-peserta-pulang')
@login_required
@role_required('Korpus', 'Korwil', 'Korda', 'Bendahara Pusat', 'Sekretaris', 'Korpuspi')
def export_peserta_pulang():
    active_edisi = get_active_edisi()
    query = Pendaftaran.query.filter(Pendaftaran.edisi_id == active_edisi.id)
    
    # Filter hanya yang ikut pulang
    query = query.filter(Pendaftaran.status_pulang != 'Tidak Ikut')

    if current_user.role.name in ['Korda', 'Korwil']:
        managed_rombongan_ids = [r.id for r in current_user.active_managed_rombongan]
        if managed_rombongan_ids:
            query = query.filter(Pendaftaran.rombongan_pulang_id.in_(managed_rombongan_ids))
        else:
            query = query.filter(db.false())

    # Apply other filters
    query = apply_filters(query, exclude_perjalanan=True)
    semua_pendaftar = query.order_by(Santri.nama).all()

    # ADDED: Create tarif map for pulang
    tarif_map = {}
    for p in semua_pendaftar:
        if p.rombongan_pulang_id and p.titik_turun:
            tarif_key = f"{p.rombongan_pulang_id}_{p.titik_turun}"
            if tarif_key not in tarif_map:
                tarif = Tarif.query.filter_by(
                    rombongan_id=p.rombongan_pulang_id, 
                    titik_turun=p.titik_turun
                ).first()
                if tarif:
                    tarif_map[tarif_key] = tarif

    # MODIFIED: Kalkulasi untuk perjalanan pulang menggunakan tarif sebenarnya
    total_lunas = 0
    total_belum_lunas = 0
    total_cash = 0
    total_transfer = 0
    jumlah_orang_cash = 0
    jumlah_orang_transfer = 0
    
    for p in semua_pendaftar:
        # Calculate biaya pulang using tarif
        biaya_pulang = 0
        if p.status_pulang and p.status_pulang != 'Tidak Ikut':
            if p.rombongan_pulang_id and p.titik_turun:
                tarif_key = f"{p.rombongan_pulang_id}_{p.titik_turun}"
                tarif = tarif_map.get(tarif_key)
                if tarif:
                    biaya_pulang = tarif.harga_bus + tarif.fee_korda + 10000
        
        # Set calculated biaya to the object
        p.biaya_pulang_calculated = biaya_pulang
        
        # Kalkulasi berdasarkan status pembayaran
        if p.status_pulang == 'Lunas':
            total_lunas += biaya_pulang
        else:
            total_belum_lunas += biaya_pulang
        
        # Kalkulasi berdasarkan metode pembayaran
        if p.metode_pembayaran_pulang == 'Cash':
            total_cash += biaya_pulang
            jumlah_orang_cash += 1
        elif p.metode_pembayaran_pulang == 'Transfer':
            total_transfer += biaya_pulang
            jumlah_orang_transfer += 1

    data_for_excel = []
    for index, p in enumerate(semua_pendaftar, 1):
        data_for_excel.append({
            'No.': index,
            'Nama Santri': p.santri.nama if p.santri else 'N/A',
            'Jenis Kelamin': p.santri.jenis_kelamin if p.santri else 'N/A',
            'Asrama': p.santri.asrama if p.santri else 'N/A',
            'Kabupaten': p.santri.kabupaten if p.santri else 'N/A',
            'Rombongan Pulang': p.rombongan_pulang.nama_rombongan if p.rombongan_pulang else 'N/A',
            'Titik Turun': p.titik_turun or '-',
            'Bus Pulang': p.bus_pulang.nama_armada if p.bus_pulang else '-',
            'Status Pulang': p.status_pulang,
            'Metode Bayar': p.metode_pembayaran_pulang or '-',
            'Biaya Pulang': f"Rp. {p.biaya_pulang_calculated:,.0f}"
        })

    summary_data = {
        'total_lunas': total_lunas,
        'total_belum_lunas': total_belum_lunas,
        'total_cash': total_cash,
        'total_transfer': total_transfer,
        'jumlah_orang_cash': jumlah_orang_cash,
        'jumlah_orang_transfer': jumlah_orang_transfer
    }

    return create_excel_file(data_for_excel, summary_data, 'Data_Peserta_Pulang.xlsx', 'PULANG')

# Export untuk Data Kembali
@admin_bp.route('/export-peserta-kembali')
@login_required
@role_required('Korpus', 'Korwil', 'Korda', 'Bendahara Pusat', 'Sekretaris', 'Korpuspi')
def export_peserta_kembali():
    active_edisi = get_active_edisi()
    query = Pendaftaran.query.filter(Pendaftaran.edisi_id == active_edisi.id)
    
    # Filter hanya yang ikut kembali
    query = query.filter(Pendaftaran.status_kembali != 'Tidak Ikut')

    if current_user.role.name in ['Korda', 'Korwil']:
        managed_rombongan_ids = [r.id for r in current_user.active_managed_rombongan]
        if managed_rombongan_ids:
            query = query.filter(Pendaftaran.rombongan_kembali_id.in_(managed_rombongan_ids))
        else:
            query = query.filter(db.false())

    # Apply other filters
    query = apply_filters(query, exclude_perjalanan=True)
    semua_pendaftar = query.order_by(Santri.nama).all()

    # ADDED: Create tarif map for kembali
    tarif_map = {}
    for p in semua_pendaftar:
        rombongan_kembali_id = p.rombongan_kembali_id or p.rombongan_pulang_id
        titik_jemput_kembali = p.titik_jemput_kembali or p.titik_turun
        if rombongan_kembali_id and titik_jemput_kembali:
            tarif_key = f"{rombongan_kembali_id}_{titik_jemput_kembali}"
            if tarif_key not in tarif_map:
                tarif = Tarif.query.filter_by(
                    rombongan_id=rombongan_kembali_id, 
                    titik_turun=titik_jemput_kembali
                ).first()
                if tarif:
                    tarif_map[tarif_key] = tarif

    # MODIFIED: Kalkulasi untuk perjalanan kembali menggunakan tarif sebenarnya
    total_lunas = 0
    total_belum_lunas = 0
    total_cash = 0
    total_transfer = 0
    jumlah_orang_cash = 0
    jumlah_orang_transfer = 0
    
    for p in semua_pendaftar:
        # Calculate biaya kembali using tarif
        biaya_kembali = 0
        if p.status_kembali and p.status_kembali != 'Tidak Ikut':
            rombongan_kembali_id = p.rombongan_kembali_id or p.rombongan_pulang_id
            titik_jemput_kembali = p.titik_jemput_kembali or p.titik_turun
            if rombongan_kembali_id and titik_jemput_kembali:
                tarif_key = f"{rombongan_kembali_id}_{titik_jemput_kembali}"
                tarif = tarif_map.get(tarif_key)
                if tarif:
                    biaya_kembali = tarif.harga_bus + tarif.fee_korda + 10000
        
        # Set calculated biaya to the object
        p.biaya_kembali_calculated = biaya_kembali
        
        # Kalkulasi berdasarkan status pembayaran
        if p.status_kembali == 'Lunas':
            total_lunas += biaya_kembali
        else:
            total_belum_lunas += biaya_kembali
        
        # Kalkulasi berdasarkan metode pembayaran
        if p.metode_pembayaran_kembali == 'Cash':
            total_cash += biaya_kembali
            jumlah_orang_cash += 1
        elif p.metode_pembayaran_kembali == 'Transfer':
            total_transfer += biaya_kembali
            jumlah_orang_transfer += 1

    data_for_excel = []
    for index, p in enumerate(semua_pendaftar, 1):
        data_for_excel.append({
            'No.': index,
            'Nama Santri': p.santri.nama if p.santri else 'N/A',
            'Jenis Kelamin': p.santri.jenis_kelamin if p.santri else 'N/A',
            'Asrama': p.santri.asrama if p.santri else 'N/A',
            'Kabupaten': p.santri.kabupaten if p.santri else 'N/A',
            'Rombongan Kembali': p.rombongan_kembali.nama_rombongan if p.rombongan_kembali else (p.rombongan_pulang.nama_rombongan if p.rombongan_pulang else 'N/A'),
            'Titik Jemput': p.titik_jemput_kembali or p.titik_turun or '-',
            'Bus Kembali': p.bus_kembali.nama_armada if p.bus_kembali else '-',
            'Status Kembali': p.status_kembali,
            'Metode Bayar': p.metode_pembayaran_kembali or '-',
            'Biaya Kembali': f"Rp. {p.biaya_kembali_calculated:,.0f}"
        })

    summary_data = {
        'total_lunas': total_lunas,
        'total_belum_lunas': total_belum_lunas,
        'total_cash': total_cash,
        'total_transfer': total_transfer,
        'jumlah_orang_cash': jumlah_orang_cash,
        'jumlah_orang_transfer': jumlah_orang_transfer
    }

    return create_excel_file(data_for_excel, summary_data, 'Data_Peserta_Kembali.xlsx', 'KEMBALI')

# Helper function untuk apply filters
def apply_filters(query, exclude_perjalanan=False):
    nama_filter = request.args.get('nama', '')
    rombongan_id_filter = request.args.get('rombongan_id', type=int)
    status_bayar_filter = request.args.get('status_bayar', '')
    jenis_kelamin_filter = request.args.get('jenis_kelamin', '')
    
    # Gabungkan dengan tabel Santri untuk filter
    query = query.join(Santri)
    
    if nama_filter:
        query = query.filter(Santri.nama.ilike(f'%{nama_filter}%'))
    if rombongan_id_filter:
        query = query.filter(or_(Pendaftaran.rombongan_pulang_id == rombongan_id_filter, Pendaftaran.rombongan_kembali_id == rombongan_id_filter))
    if jenis_kelamin_filter:
        query = query.filter(Santri.jenis_kelamin == jenis_kelamin_filter)
    
    if not exclude_perjalanan:
        perjalanan_filter = request.args.get('perjalanan', '')
        if perjalanan_filter == 'ikut_pulang':
            query = query.filter(Pendaftaran.status_pulang != 'Tidak Ikut')
        elif perjalanan_filter == 'tidak_ikut_pulang':
            query = query.filter(Pendaftaran.status_pulang == 'Tidak Ikut')
        elif perjalanan_filter == 'ikut_kembali':
            query = query.filter(Pendaftaran.status_kembali != 'Tidak Ikut')
        elif perjalanan_filter == 'tidak_ikut_kembali':
            query = query.filter(Pendaftaran.status_kembali == 'Tidak Ikut')

    if status_bayar_filter:
        if status_bayar_filter == 'Lunas':
            query = query.filter(and_(Pendaftaran.status_pulang == 'Lunas', Pendaftaran.status_kembali == 'Lunas'))
        else:
            query = query.filter(or_(Pendaftaran.status_pulang == 'Belum Bayar', Pendaftaran.status_kembali == 'Belum Bayar'))
    
    return query

# Helper function untuk membuat file Excel (unchanged)
def create_excel_file(data_for_excel, summary_data, filename, report_type):
    if not data_for_excel:
        flash('Tidak ada data untuk diekspor sesuai filter yang dipilih.', 'warning')
        return redirect(url_for('admin.daftar_peserta_global', **request.args))

    df = pd.DataFrame(data_for_excel)
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=f'Data Peserta {report_type}', startrow=6)
        
        workbook = writer.book
        worksheet = writer.sheets[f'Data Peserta {report_type}']
        
        # Import untuk styling
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # Menambahkan header dan summary
        worksheet['A1'] = f'LAPORAN DATA PESERTA {report_type}'
        worksheet['A1'].font = Font(bold=True, size=16)
        
        worksheet['A3'] = 'RINGKASAN PEMBAYARAN'
        worksheet['A3'].font = Font(bold=True, size=14)
        
        worksheet['A4'] = f'Total Sudah Lunas: Rp. {summary_data["total_lunas"]:,.0f}'
        worksheet['A5'] = f'Total Belum Lunas: Rp. {summary_data["total_belum_lunas"]:,.0f}'
        
        worksheet['D4'] = f'Total Bayar Cash: Rp. {summary_data["total_cash"]:,.0f} ({summary_data["jumlah_orang_cash"]} orang)'
        worksheet['D5'] = f'Total Bayar Transfer: Rp. {summary_data["total_transfer"]:,.0f} ({summary_data["jumlah_orang_transfer"]} orang)'
        
        # Styling untuk summary
        for cell in ['A4', 'A5', 'D4', 'D5']:
            worksheet[cell].font = Font(bold=True)
        
        # Styling untuk header data (baris 7, karena data mulai dari baris 7)
        header_fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Aplikasikan styling ke header
        for col_num in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=7, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Auto-adjust column width
        for col_num in range(1, len(df.columns) + 1):
            column_letter = get_column_letter(col_num)
            max_length = 0
            column = worksheet[column_letter]
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Styling untuk data rows
        data_fill_even = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
        
        for row_num in range(8, len(df) + 8):  # Data mulai dari baris 8
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center')
                
                if row_num % 2 == 0:
                    cell.fill = data_fill_even

    output.seek(0)
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

# Ganti fungsi perizinan() di routes.py dengan ini

@admin_bp.route('/perizinan', methods=['GET', 'POST'])
@login_required
@role_required('Korpus', 'Keamanan', 'Korda', 'Korwil', 'Bendahara', 'Korpuspi', 'Sekretaris', 'PJ Acara')
def perizinan():
    form = IzinForm()
    active_edisi = get_active_edisi()
    
    if form.validate_on_submit():
        santri = Santri.query.get(form.santri.data)
        status_pengajuan = form.status_izin.data

        if not santri or not active_edisi:
            flash("Santri atau edisi aktif tidak ditemukan.", "danger")
            return redirect(url_for('admin.perizinan'))

        if status_pengajuan == 'Diterima':
            if not form.tanggal_berakhir.data:
                flash("Untuk izin yang diterima, tanggal berakhir wajib diisi.", 'danger')
                return redirect(url_for('admin.perizinan'))
            
            # Cek duplikat izin aktif
            existing_izin = Izin.query.filter_by(santri_id=santri.id, edisi_id=active_edisi.id, status='Aktif').first()
            if existing_izin:
                flash(f"{santri.nama} sudah memiliki izin aktif.", "warning")
                return redirect(url_for('admin.perizinan'))

            santri.status_santri = 'Izin'
            new_izin = Izin(
                santri_id=santri.id,
                edisi_id=active_edisi.id,
                status='Aktif',
                tanggal_pengajuan=form.tanggal_pengajuan.data,
                tanggal_berakhir=form.tanggal_berakhir.data,
                keterangan=form.keterangan.data
            )
            flash(f"Izin untuk {santri.nama} berhasil dicatat dan statusnya diubah menjadi Izin.", "success")

        elif status_pengajuan == 'Ditolak':
            # Santri status tetap 'Aktif'
            new_izin = Izin(
                santri_id=santri.id,
                edisi_id=active_edisi.id,
                status='Ditolak',
                tanggal_pengajuan=form.tanggal_pengajuan.data,
                keterangan=form.keterangan.data
                # tanggal_berakhir dikosongkan
            )
            flash(f"Pengajuan izin untuk {santri.nama} berhasil dicatat dengan status DITOLAK.", "info")

        db.session.add(new_izin)
        db.session.commit()
        return redirect(url_for('admin.perizinan'))

    # Logika untuk menampilkan data di tabel
    view_mode = request.args.get('view', 'aktif')
    query = Izin.query.filter_by(edisi_id=active_edisi.id).options(joinedload(Izin.santri))

    if view_mode == 'aktif':
        query = query.filter_by(status='Aktif')
    elif view_mode == 'riwayat':
        query = query.filter_by(status='Selesai')
    elif view_mode == 'ditolak':
        query = query.filter_by(status='Ditolak')

    # Filter pencarian
    nama_filter = request.args.get('nama', '')
    if nama_filter:
        query = query.join(Santri).filter(Santri.nama.ilike(f'%{nama_filter}%'))

    semua_izin = query.order_by(Izin.tanggal_pengajuan.desc()).all()
    
    return render_template('perizinan.html', 
                           form=form, 
                           semua_izin=semua_izin,
                           view_mode=view_mode)

@admin_bp.route('/export/perizinan/<jenis_kelamin>')
@login_required
@role_required('Korpus', 'Korda', 'Korwil', 'Keamanan', 'PJ Acara', 'Sekretaris', 'Korpuspi')
def export_perizinan(jenis_kelamin):
    """Export data perizinan ke Excel berdasarkan jenis kelamin"""
    
    # Validasi parameter jenis_kelamin
    if jenis_kelamin.upper() not in ['PUTRA', 'PUTRI']:
        flash("Jenis kelamin tidak valid!", "danger")
        return redirect(url_for('admin.perizinan'))
    
    active_edisi = get_active_edisi()
    if not active_edisi:
        flash("Tidak ada edisi aktif!", "danger")
        return redirect(url_for('admin.perizinan'))
    
    try:
        # Query semua izin (aktif dan riwayat) berdasarkan jenis kelamin
        query = db.session.query(Izin, Santri).join(Santri).filter(
            Izin.edisi == active_edisi,
            Santri.jenis_kelamin == jenis_kelamin.upper()
        ).order_by(Izin.tanggal_berakhir.desc())
        
        results = query.all()
        
        # Buat data untuk DataFrame
        data = []
        for i, (izin, santri) in enumerate(results, 1):
            data.append({
                'No': i,
                'Nama Santri': santri.nama,
                'Jenis Kelamin': 'Putra' if santri.jenis_kelamin == 'PUTRA' else 'Putri',
                'Asrama': santri.asrama or '-',
                'Kabupaten': santri.kabupaten or '-',
                'Izin Sampai': izin.tanggal_berakhir.strftime('%d/%m/%Y') if izin.tanggal_berakhir else '-',
                'Status': izin.status,
                'Keterangan': izin.keterangan or '-',
                'Dicatat Oleh': izin.user.username if hasattr(izin, 'user') and izin.user else '-'
            })
        
        # Buat DataFrame
        df = pd.DataFrame(data)
        
        # Buat Excel file dalam memory
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Sheet utama dengan data
            sheet_name = f'Perizinan {jenis_kelamin.title()}'
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Ambil workbook dan worksheet untuk formatting
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]
            
            # Format header
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Format header row
            for col_num, column in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Format data rows
            for row in range(2, len(df) + 2):
                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.border = border
                    cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Tambah sheet ringkasan
            summary_data = {
                'Informasi': [
                    'Total Izin',
                    'Izin Aktif', 
                    'Izin Selesai',
                    'Tanggal Export',
                    'Export Oleh',
                    'Edisi'
                ],
                'Value': [
                    len(df),
                    len(df[df['Status'] == 'Aktif']) if len(df) > 0 else 0,
                    len(df[df['Status'] == 'Selesai']) if len(df) > 0 else 0,
                    datetime.now().strftime('%d/%m/%Y %H:%M'),
                    current_user.username,
                    active_edisi.nama
                ]
            }
            
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Ringkasan', index=False)
            
            # Format sheet ringkasan
            summary_sheet = writer.sheets['Ringkasan']
            for col_num in range(1, 3):
                cell = summary_sheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Auto-adjust summary columns
            for column in summary_sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = max_length + 2
                summary_sheet.column_dimensions[column_letter].width = adjusted_width
                
            # Format data rows di summary
            for row in range(2, len(summary_df) + 2):
                for col in range(1, 3):
                    cell = summary_sheet.cell(row=row, column=col)
                    cell.border = border
                    cell.alignment = Alignment(horizontal='left', vertical='center')
        
        output.seek(0)
        
        # Generate filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'perizinan_{jenis_kelamin.lower()}_{active_edisi.nama}_{timestamp}.xlsx'
        
        # Log activity
        log_activity('Export', 'Perizinan', f"Export data perizinan {jenis_kelamin.title()} ke Excel: {len(df)} record")
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        flash(f"Terjadi kesalahan saat mengexport data: {str(e)}", "danger")
        return redirect(url_for('admin.perizinan'))

@admin_bp.route('/edit-izin/<int:izin_id>', methods=['GET', 'POST'])
@login_required
@role_required('Korpus', 'Keamanan')
def edit_izin(izin_id):
    """Edit data izin santri"""
    izin = Izin.query.get_or_404(izin_id)
    form = EditIzinForm()
    
    # Populate form dengan data existing saat GET request
    if request.method == 'GET':
        form.status_izin.data = izin.status
        form.tanggal_pengajuan.data = izin.tanggal_pengajuan
        form.tanggal_berakhir.data = izin.tanggal_berakhir
        form.keterangan.data = izin.keterangan
    
    if form.validate_on_submit():
        old_status = izin.status
        new_status = form.status_izin.data
        
        # Validasi tanggal berakhir untuk status 'Aktif'
        if new_status == 'Aktif' and not form.tanggal_berakhir.data:
            flash("Untuk izin yang diterima, tanggal berakhir wajib diisi.", 'danger')
            return redirect(url_for('admin.edit_izin', izin_id=izin_id))
        
        # Update data izin
        izin.status = new_status
        izin.tanggal_pengajuan = form.tanggal_pengajuan.data
        izin.tanggal_berakhir = form.tanggal_berakhir.data if new_status == 'Aktif' else None
        izin.keterangan = form.keterangan.data
        
        # Update status santri sesuai dengan status izin
        if new_status == 'Aktif':
            izin.santri.status_santri = 'Izin'
            status_message = "diubah menjadi DITERIMA"
        elif new_status == 'Ditolak':
            # Jika diubah ke ditolak, pastikan tidak ada izin aktif lainnya
            other_active_izin = Izin.query.filter_by(
                santri_id=izin.santri_id, 
                edisi_id=izin.edisi_id, 
                status='Aktif'
            ).filter(Izin.id != izin.id).first()
            
            if not other_active_izin:
                izin.santri.status_santri = 'Aktif'
            status_message = "diubah menjadi DITOLAK"
        
        try:
            db.session.commit()
            flash(f"Izin untuk {izin.santri.nama} berhasil {status_message}.", "success")
            return redirect(url_for('admin.perizinan'))
        except Exception as e:
            db.session.rollback()
            flash(f"Terjadi kesalahan: {str(e)}", "danger")
            return redirect(url_for('admin.edit_izin', izin_id=izin_id))
    
    return render_template('edit_izin.html', form=form, izin=izin)

@admin_bp.route('/delete-izin/<int:izin_id>', methods=['POST'])
@login_required
@role_required('Korpus', 'Keamanan')
def delete_izin(izin_id):
    """Hapus data izin (khusus untuk izin yang ditolak)"""
    izin = Izin.query.get_or_404(izin_id)
    
    # Hanya izin yang ditolak yang bisa dihapus
    if izin.status != 'Ditolak':
        flash("Hanya izin yang ditolak yang dapat dihapus. Gunakan 'Cabut' untuk izin aktif.", "warning")
        return redirect(url_for('admin.perizinan'))
    
    santri_nama = izin.santri.nama
    
    try:
        db.session.delete(izin)
        db.session.commit()
        flash(f"Record izin yang ditolak untuk {santri_nama} berhasil dihapus.", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Terjadi kesalahan saat menghapus: {str(e)}", "danger")
    
    return redirect(url_for('admin.perizinan'))

@admin_bp.route('/cabut-izin/<int:izin_id>', methods=['POST'])
@login_required
@role_required('Korpus', 'Keamanan')
def cabut_izin(izin_id):
    """Cabut izin aktif santri"""
    izin = Izin.query.get_or_404(izin_id)
    
    # Hanya izin aktif yang bisa dicabut
    if izin.status != 'Aktif':
        flash("Hanya izin aktif yang dapat dicabut.", "warning")
        return redirect(url_for('admin.perizinan'))
    
    try:
        # Update status izin menjadi 'Berakhir'
        izin.status = 'Berakhir'
        izin.tanggal_berakhir = datetime.utcnow().date()
        
        # Kembalikan status santri ke 'Aktif'
        izin.santri.status_santri = 'Aktif'
        
        db.session.commit()
        flash(f"Izin untuk {izin.santri.nama} berhasil dicabut. Status santri dikembalikan ke Aktif.", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Terjadi kesalahan saat mencabut izin: {str(e)}", "danger")
    
    return redirect(url_for('admin.perizinan'))

@admin_bp.route('/api/search-active-santri')
def search_active_santri_api():
    query = request.args.get('q', '')
    if len(query) < 3:
        return jsonify({'results': []})

    # Cari santri yang namanya cocok DAN statusnya 'Aktif'
    santri_results = Santri.query.filter(
        Santri.nama.ilike(f'%{query}%'),
        Santri.status_santri == 'Aktif'
    ).limit(20).all()

    results = [{'id': s.id, 'nama': s.nama} for s in santri_results]
    return jsonify({'results': results})



# Di dalam file app/admin/routes.py

from collections import defaultdict
from flask import request, make_response, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import io
from datetime import datetime

@admin_bp.route('/partisipan')
@role_required('Korpus', 'Korda', 'Korwil', 'Keamanan', 'PJ Acara', 'Bendahara Pusat', 'Sekretaris', 'Korpuspi')
@login_required
def data_partisipan():
    active_edisi = get_active_edisi()
    alamat_filter = request.args.get('alamat_filter', '').strip()
    
    semua_partisipan = []
    if active_edisi:
        # Mulai dengan query dasar untuk mengambil semua partisipan di edisi aktif
        query = Partisipan.query.filter_by(edisi=active_edisi)
        
        # Terapkan filter HANYA JIKA rolenya Korda atau Korwil
        if current_user.role.name in ['Korda', 'Korwil']:
            # 1. Kumpulkan semua cakupan wilayah yang dikelola oleh user saat ini
            managed_regions = {
                wilayah.get('label')
                for rombongan in current_user.active_managed_rombongan
                if rombongan.cakupan_wilayah
                for wilayah in rombongan.cakupan_wilayah
            }
            
            # 2. Filter query partisipan berdasarkan wilayah santri
            if managed_regions:
                # Lakukan join dengan tabel Santri dan filter berdasarkan kolom 'kabupaten'
                query = query.join(Santri).filter(Santri.kabupaten.in_(list(managed_regions)))
            else:
                # Jika Korda/Korwil tidak punya cakupan wilayah, jangan tampilkan data apapun
                query = query.filter(db.false())
        
        # Eksekusi query setelah semua filter yang diperlukan diterapkan
        semua_partisipan = query.all()
    
    # Dapatkan semua alamat unik dari partisipan (untuk dropdown filter)
    all_alamat = set()
    for p in semua_partisipan:
        if p.santri.kabupaten:
            all_alamat.add(p.santri.kabupaten)
    available_alamat = sorted(list(all_alamat))
    
    # Filter berdasarkan alamat jika ada
    if alamat_filter:
        semua_partisipan = [p for p in semua_partisipan if p.santri.kabupaten == alamat_filter]
    
    # Kelompokkan data berdasarkan kategori
    grouped_partisipan = defaultdict(list)
    for p in semua_partisipan:
        grouped_partisipan[p.kategori].append(p)
    
    # Hitung statistik
    total_partisipan = len(semua_partisipan)
    total_alamat = len(all_alamat)
    filtered_count = len(semua_partisipan) if alamat_filter else 0
    
    return render_template(
        'data_partisipan.html',
        grouped_partisipan=grouped_partisipan,
        available_alamat=available_alamat,
        total_partisipan=total_partisipan,
        total_alamat=total_alamat,
        filtered_count=filtered_count
    )


@admin_bp.route('/partisipan/export')
@role_required('Korpus', 'Korda', 'Korwil', 'Keamanan', 'PJ Acara', 'Bendahara Pusat', 'Sekretaris', 'Korpuspi')
@login_required
def export_partisipan():
    active_edisi = get_active_edisi()
    alamat_filter = request.args.get('alamat_filter', '').strip()
    
    if not active_edisi:
        flash('Tidak ada edisi aktif saat ini.', 'error')
        return redirect(url_for('admin.data_partisipan'))
    
    # Query data partisipan (sama dengan fungsi data_partisipan)
    query = Partisipan.query.filter_by(edisi=active_edisi)
    
    # Terapkan filter role
    if current_user.role.name in ['Korda', 'Korwil']:
        managed_regions = {
            wilayah.get('label')
            for rombongan in current_user.active_managed_rombongan
            if rombongan.cakupan_wilayah
            for wilayah in rombongan.cakupan_wilayah
        }
        
        if managed_regions:
            query = query.join(Santri).filter(Santri.kabupaten.in_(list(managed_regions)))
        else:
            query = query.filter(db.false())
    
    semua_partisipan = query.all()
    
    # Filter berdasarkan alamat jika ada
    if alamat_filter:
        semua_partisipan = [p for p in semua_partisipan if p.santri.kabupaten == alamat_filter]
    
    # Kelompokkan data berdasarkan kategori
    grouped_partisipan = defaultdict(list)
    for p in semua_partisipan:
        grouped_partisipan[p.kategori].append(p)
    
    # Buat workbook Excel
    wb = Workbook()
    
    # Hapus sheet default
    wb.remove(wb.active)
    
    # Style untuk header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Style untuk border
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'), 
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Style untuk kategori header
    kategori_font = Font(bold=True, color="FFFFFF")
    kategori_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    
    # Buat sheet ringkasan
    summary_sheet = wb.create_sheet("Ringkasan")
    
    # Header ringkasan
    summary_headers = ["Kategori", "Jumlah Partisipan"]
    for col, header in enumerate(summary_headers, 1):
        cell = summary_sheet.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Data ringkasan
    row = 2
    total_semua = 0
    for kategori, partisipan_list in sorted(grouped_partisipan.items()):
        summary_sheet.cell(row=row, column=1, value=kategori).border = thin_border
        jumlah = len(partisipan_list)
        summary_sheet.cell(row=row, column=2, value=jumlah).border = thin_border
        total_semua += jumlah
        row += 1
    
    # Total
    summary_sheet.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    summary_sheet.cell(row=row, column=1).border = thin_border
    summary_sheet.cell(row=row, column=2, value=total_semua).font = Font(bold=True)
    summary_sheet.cell(row=row, column=2).border = thin_border
    
    # Auto-width untuk kolom ringkasan
    for column in summary_sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        summary_sheet.column_dimensions[column_letter].width = adjusted_width
    
    # Buat sheet untuk setiap kategori
    for kategori, partisipan_list in sorted(grouped_partisipan.items()):
        # Buat nama sheet yang valid (maksimal 31 karakter, tanpa karakter khusus)
        sheet_name = kategori[:31].replace('/', '-').replace('\\', '-').replace('*', '-').replace('?', '-').replace(':', '-').replace('[', '-').replace(']', '-')
        ws = wb.create_sheet(sheet_name)
        
        # Header tabel
        headers = ["No", "Nama Santri", "Asrama", "Alamat", "Kelas Formal", "Kelas Ngaji"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Data partisipan (diurutkan berdasarkan nama)
        partisipan_sorted = sorted(partisipan_list, key=lambda x: x.santri.nama)
        
        for row, partisipan in enumerate(partisipan_sorted, 2):
            ws.cell(row=row, column=1, value=row-1).border = thin_border  # No
            ws.cell(row=row, column=2, value=partisipan.santri.nama).border = thin_border  # Nama
            ws.cell(row=row, column=3, value=partisipan.santri.asrama).border = thin_border  # Asrama
            ws.cell(row=row, column=4, value=partisipan.santri.kabupaten).border = thin_border  # Alamat
            ws.cell(row=row, column=5, value=partisipan.santri.kelas_formal or '-').border = thin_border  # Kelas Formal
            ws.cell(row=row, column=6, value=partisipan.santri.kelas_ngaji or '-').border = thin_border  # Kelas Ngaji
        
        # Auto-width untuk kolom
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Jika tidak ada data
    if not grouped_partisipan:
        ws = wb.create_sheet("Tidak Ada Data")
        ws.cell(row=1, column=1, value="Tidak ada data partisipan untuk diekspor.")
    
    # Simpan ke BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate nama file
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filter_text = f"_{alamat_filter}" if alamat_filter else ""
    filename = f"Data_Partisipan_{active_edisi.nama}{filter_text}_{timestamp}.xlsx"
    
    # Buat response
    response = make_response(output.getvalue())
    response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    
    return response

# Buat route placeholder untuk form tambah agar link-nya tidak error
@admin_bp.route('/partisipan/tambah', methods=['GET', 'POST'])
@login_required
@role_required('Korpus', 'PJ Acara', 'Sekretaris', 'Korpuspi')
def tambah_partisipan():
    manual_form = PartisipanForm()
    import_form = ImportPartisipanForm() # <-- Buat instance form impor
    active_edisi = get_active_edisi()

    if not active_edisi:
        flash("Tidak bisa menambah status partisipan karena tidak ada edisi yang aktif.", "danger")
        return redirect(url_for('admin.data_partisipan'))

    # --- Logika untuk Impor Massal ---
    if 'submit_import' in request.form and import_form.validate_on_submit():
        file = import_form.file.data
        kategori = import_form.kategori.data

        try:
            df = pd.read_excel(file)
            if 'NIS' not in df.columns:
                flash('File Excel harus memiliki kolom bernama "NIS".', 'danger')
                return redirect(url_for('admin.tambah_partisipan'))

            all_nis = df['NIS'].dropna().astype(str).tolist()
            santri_map = {s.nis: s for s in Santri.query.filter(Santri.nis.in_(all_nis)).all()}
            
            new_participants = []
            for nis in all_nis:
                if nis in santri_map:
                    santri = santri_map[nis]
                    # Cek duplikat dan status
                    is_already_participant = Partisipan.query.filter_by(santri_id=santri.id, edisi_id=active_edisi.id).first()
                    if santri.status_santri == 'Aktif' and not is_already_participant:
                        santri.status_santri = 'Partisipan'
                        new_participants.append(Partisipan(
                            santri_id=santri.id,
                            edisi_id=active_edisi.id,
                            kategori=kategori
                        ))

            if new_participants:
                db.session.add_all(new_participants)
                db.session.commit()
                flash(f'{len(new_participants)} santri berhasil diimpor sebagai partisipan.', 'success')
            else:
                flash('Tidak ada santri baru yang diimpor. Mungkin semua sudah terdata atau tidak aktif.', 'info')

        except Exception as e:
            flash(f'Terjadi error saat memproses file: {e}', 'danger')

        return redirect(url_for('admin.data_partisipan'))

    # --- Logika untuk Tambah Manual (sudah ada sebelumnya) ---
    if 'submit_manual' in request.form and manual_form.validate_on_submit():
        santri_ids = manual_form.santri_ids.data
        kategori = manual_form.kategori.data
        
        for santri_id in santri_ids:
            santri = Santri.query.get(santri_id)
            is_already_participant = Partisipan.query.filter_by(santri_id=santri.id, edisi_id=active_edisi.id).first()
            if santri and santri.status_santri == 'Aktif' and not is_already_participant:
                santri.status_santri = 'Partisipan'
                new_partisipan = Partisipan(
                    edisi=active_edisi,
                    santri=santri,
                    kategori=kategori
                )
                db.session.add(new_partisipan)
        
        db.session.commit()
        flash(f"{len(santri_ids)} santri telah ditetapkan sebagai partisipan.", "success")
        return redirect(url_for('admin.data_partisipan'))

    return render_template('tambah_partisipan.html', manual_form=manual_form, import_form=import_form)


# Tambahkan API Helper baru ini
@admin_bp.route('/api/search-santri-for-partisipan')
def search_santri_for_partisipan_api():
    query = request.args.get('q', '')
    if len(query) < 3:
        return jsonify({'results': []})

    # Cari santri yang namanya cocok DAN statusnya 'Aktif'
    santri_results = Santri.query.filter(
        Santri.nama.ilike(f'%{query}%'),
        Santri.status_santri == 'Aktif'
    ).limit(20).all()

    results = [
        {
            'id': s.id, 
            'nama': s.nama,
            'asrama': s.asrama,
            'kabupaten': s.kabupaten,
            'kelas_formal': s.kelas_formal,
            'kelas_ngaji': s.kelas_ngaji
        } for s in santri_results
    ]
    return jsonify({'results': results})

@admin_bp.route('/partisipan/edit/<int:partisipan_id>', methods=['GET', 'POST'])
@login_required
@role_required('Korpus', 'PJ Acara', 'Sekretaris', 'Korpuspi')
def edit_partisipan(partisipan_id):
    partisipan = Partisipan.query.get_or_404(partisipan_id)
    form = PartisipanEditForm(obj=partisipan)

    if form.validate_on_submit():
        partisipan.kategori = form.kategori.data
        log_activity('Edit', 'Partisipan', f"Mengubah kategori partisipan untuk '{partisipan.santri.nama}' menjadi '{form.kategori.data}'")
        db.session.commit()
        flash(f"Kategori untuk {partisipan.santri.nama} berhasil diubah.", "success")
        return redirect(url_for('admin.data_partisipan'))

    return render_template('edit_partisipan.html', form=form, partisipan=partisipan)


@admin_bp.route('/partisipan/hapus/<int:partisipan_id>', methods=['POST'])
@login_required
@role_required('Korpus', 'PJ Acara', 'Sekretaris', 'Korpuspi')
def hapus_partisipan(partisipan_id):
    partisipan = Partisipan.query.get_or_404(partisipan_id)
    santri = partisipan.santri
    nama_santri = santri.nama
    
    # Kembalikan status santri menjadi 'Aktif'
    santri.status_santri = 'Aktif'
    
    log_activity('Hapus', 'Partisipan', f"Menghapus status partisipan untuk '{nama_santri}'")
    db.session.delete(partisipan)
    
    db.session.commit()
    flash(f"Status partisipan untuk '{nama_santri}' telah berhasil dihapus.", "info")
    return redirect(url_for('admin.data_partisipan'))

# --- MANAJEMEN EDISI ---
@admin_bp.route('/edisi')
@login_required
@role_required('Korpus')
def manajemen_edisi():
    semua_edisi = Edisi.query.order_by(Edisi.tahun.desc()).all()
    return render_template('manajemen_edisi.html', semua_edisi=semua_edisi)

@admin_bp.route('/edisi/tambah', methods=['GET', 'POST'])
@login_required
@role_required('Korpus')
def tambah_edisi():
    form = EdisiForm()
    if form.validate_on_submit():
        # Jika edisi baru ini di-set sebagai aktif, nonaktifkan semua yang lain
        if form.is_active.data:
            Edisi.query.update({Edisi.is_active: False})
        
        new_edisi = Edisi(
            nama=form.nama.data,
            tahun=form.tahun.data,
            is_active=form.is_active.data,
            countdown_title=form.countdown_title.data,
            countdown_target_date=form.countdown_target_date.data
        )
        db.session.add(new_edisi)
        log_activity('Tambah', 'Edisi', f"Menambahkan edisi baru: '{form.nama.data}'")
        db.session.commit()
        flash('Edisi baru berhasil ditambahkan.', 'success')
        return redirect(url_for('admin.manajemen_edisi'))
    return render_template('form_edisi.html', form=form, title="Tambah Edisi Baru")

@admin_bp.route('/edisi/edit/<int:edisi_id>', methods=['GET', 'POST'])
@login_required
@role_required('Korpus')
def edit_edisi(edisi_id):
    edisi = Edisi.query.get_or_404(edisi_id)
    form = EdisiForm(obj=edisi)
    if form.validate_on_submit():
        # Jika edisi ini di-set sebagai aktif, nonaktifkan semua yang lain
        if form.is_active.data:
            Edisi.query.filter(Edisi.id != edisi_id).update({Edisi.is_active: False})

        edisi.nama = form.nama.data
        edisi.tahun = form.tahun.data
        edisi.is_active = form.is_active.data
        edisi.countdown_title = form.countdown_title.data
        edisi.countdown_target_date = form.countdown_target_date.data
        log_activity('Edit', 'Edisi', f"Mengubah data edisi: '{edisi.nama}'")
        db.session.commit()
        flash('Data edisi berhasil diperbarui.', 'success')
        return redirect(url_for('admin.manajemen_edisi'))
    return render_template('form_edisi.html', form=form, title="Edit Edisi")

@admin_bp.route('/edisi/hapus/<int:edisi_id>', methods=['POST'])
@login_required
@role_required('Korpus')
def hapus_edisi(edisi_id):
    edisi = Edisi.query.get_or_404(edisi_id)
    # Cek apakah edisi ini masih memiliki rombongan, jika iya jangan dihapus
    if edisi.rombongans:
        flash('Edisi ini tidak bisa dihapus karena masih memiliki data rombongan.', 'danger')
        return redirect(url_for('admin.manajemen_edisi'))
    
    nama_edisi = edisi.nama
    log_activity('Hapus', 'Edisi', f"Menghapus edisi: '{nama_edisi}'")
    db.session.delete(edisi)
    db.session.commit()
    flash('Edisi berhasil dihapus.', 'info')
    return redirect(url_for('admin.manajemen_edisi'))

def get_active_edisi():
    """Mencari dan mengembalikan edisi yang sedang aktif."""
    return Edisi.query.filter_by(is_active=True).first()

# --- CONTEXT PROCESSOR ---
# Ini membuat variabel 'active_edisi' tersedia di semua template secara otomatis
@admin_bp.context_processor
def inject_active_edisi():
    return dict(active_edisi=get_active_edisi())

@admin_bp.route('/riwayat')
@login_required
@role_required('Korpus')
def riwayat_edisi():
    # Ambil semua edisi yang statusnya tidak aktif
    arsip_edisi = Edisi.query.filter_by(is_active=False).order_by(Edisi.tahun.desc()).all()
    
    summary_arsip = []
    for edisi in arsip_edisi:
        # --- PERBAIKI QUERY DI SINI ---
        # JOIN secara eksplisit untuk menghindari ambiguitas
        # Kita bisa join pada salah satu (pulang/kembali) karena kita hanya butuh edisi_id
        base_query = Pendaftaran.query.join(
            Rombongan, Pendaftaran.rombongan_pulang_id == Rombongan.id
        ).filter(Rombongan.edisi_id == edisi.id)

        # Hitung total peserta yang unik
        total_peserta = base_query.distinct(Pendaftaran.santri_id).count()
        
        # Hitung total pemasukan
        total_pemasukan = base_query.with_entities(db.func.sum(Pendaftaran.total_biaya)).scalar() or 0
        
        summary_arsip.append({
            'edisi': edisi,
            'total_peserta': total_peserta,
            'total_pemasukan': total_pemasukan
        })

    return render_template('riwayat_edisi.html', summary_arsip=summary_arsip)


@admin_bp.route('/riwayat/<int:edisi_id>')
@login_required
@role_required('Korpus')
def detail_arsip(edisi_id):
    arsip_edisi = Edisi.query.get_or_404(edisi_id)
    
    # --- MULAI KALKULASI LENGKAP (SAMA SEPERTI DASHBOARD) ---
    stats = {}
    # JOIN eksplisit untuk menghindari ambiguitas
    base_pendaftaran_query = Pendaftaran.query.join(
        Rombongan, Pendaftaran.rombongan_pulang_id == Rombongan.id
    ).filter(Rombongan.edisi_id == arsip_edisi.id)
    
    pendaftar_list = base_pendaftaran_query.all()
    
    stats['total_peserta'] = len(pendaftar_list)
    stats['total_pemasukan'] = sum(p.total_biaya for p in pendaftar_list)
    
    # Kalkulasi Lunas (contoh sederhana)
    total_lunas = 0
    for p in pendaftar_list:
        if p.status_pulang == 'Lunas' and p.status_kembali in ['Lunas', 'Tidak Ikut']:
            total_lunas += p.total_biaya
        elif p.status_pulang in ['Lunas', 'Tidak Ikut'] and p.status_kembali == 'Lunas':
            total_lunas += p.total_biaya
    stats['total_lunas'] = total_lunas
    stats['total_belum_lunas'] = stats['total_pemasukan'] - stats['total_lunas']
    # Anda bisa tambahkan kalkulasi lain jika perlu...

    semua_rombongan = Rombongan.query.filter_by(edisi=arsip_edisi).order_by(Rombongan.nama_rombongan).all()
    # --- AKHIR KALKULASI ---
    
    return render_template('dashboard.html', 
                           stats=stats, 
                           semua_rombongan=semua_rombongan,
                           is_arsip=True,
                           arsip_edisi_nama=arsip_edisi.nama)

@admin_bp.route('/rombongan/<int:rombongan_id>/tambah-bus', methods=['POST'])
@login_required
@role_required('Korpus', 'Korda', 'Sekretaris', 'Korpuspi')
def tambah_bus(rombongan_id):
    rombongan = Rombongan.query.get_or_404(rombongan_id)
    # Verifikasi kepemilikan untuk Korda
    if current_user.role.name == 'Korda' and rombongan not in current_user.active_managed_rombongan:
        abort(403)
    
    form = BusForm()
    if form.validate_on_submit():
        new_bus = Bus(
            rombongan_id=rombongan.id,
            nama_armada=form.nama_armada.data,
            nomor_lambung=form.nomor_lambung.data,
            plat_nomor=form.plat_nomor.data,
            kuota=form.kuota.data,
            keterangan=form.keterangan.data
        )
        db.session.add(new_bus)
        log_activity('Tambah', 'Bus', f"Menambah bus '{form.nama_armada.data}' ke rombongan '{rombongan.nama_rombongan}'")
        db.session.commit()
        flash('Bus baru berhasil ditambahkan.', 'success')
    else:
        flash('Gagal menambahkan bus. Pastikan semua field terisi.', 'danger')
    
    return redirect(url_for('admin.manajemen_peserta_bus', id=rombongan_id))

@admin_bp.route('/bus/edit/<int:bus_id>', methods=['GET', 'POST'])
@login_required
@role_required('Korpus', 'Korda', 'Korpuspi')
def edit_bus(bus_id):
    bus = Bus.query.get_or_404(bus_id)
    # Keamanan: Pastikan user yang login berhak mengedit bus ini
    if current_user.role.name == 'Korda' and bus.rombongan not in current_user.active_managed_rombongan:
        abort(403)

    form = BusForm(obj=bus) # Isi form dengan data bus yang ada

    if form.validate_on_submit():
        # Simpan perubahan dari form ke objek bus
        bus.nama_armada = form.nama_armada.data
        bus.nomor_lambung = form.nomor_lambung.data
        bus.plat_nomor = form.plat_nomor.data
        bus.kuota = form.kuota.data
        bus.keterangan = form.keterangan.data
        
        db.session.commit()
        log_activity('Edit', 'Bus', f"Mengubah detail bus '{bus.nama_armada}' di rombongan '{bus.rombongan.nama_rombongan}'")
        flash('Detail bus berhasil diperbarui!', 'success')
        return redirect(url_for('admin.manajemen_peserta_bus', id=bus.rombongan_id))

    return render_template('edit_bus.html', form=form, bus=bus, title="Edit Detail Bus")

@admin_bp.route('/bus/hapus/<int:bus_id>', methods=['POST'])
@login_required
@role_required('Korpus', 'Korda', 'Sekretaris', 'Korpuspi')
def hapus_bus(bus_id):
    bus = Bus.query.get_or_404(bus_id)
    rombongan_id = bus.rombongan_id
    # Verifikasi kepemilikan untuk Korda
    if current_user.role.name == 'Korda' and bus.rombongan not in current_user.active_managed_rombongan:
        abort(403)

    nama_bus = f"{bus.nama_armada} ({bus.nomor_lambung or bus.plat_nomor})"
    db.session.delete(bus)
    log_activity('Hapus', 'Bus', f"Menghapus bus '{nama_bus}' dari rombongan '{bus.rombongan.nama_rombongan}'")
    db.session.commit()
    flash('Bus berhasil dihapus.', 'info')
    return redirect(url_for('admin.manajemen_peserta_bus', id=rombongan_id))

@admin_bp.route('/bus/<int:bus_id>/detail')
@login_required
@role_required('Korpus', 'Korwil', 'Korda', 'Bendahara Pusat', 'Sekretaris', 'Korpuspi')
def detail_bus(bus_id):
    bus = Bus.query.get_or_404(bus_id)
    
    # Verifikasi kepemilikan untuk Korda/Korwil
    if current_user.role.name in ['Korwil', 'Korda']:
        if bus.rombongan not in current_user.active_managed_rombongan:
            abort(403)

     # Ambil semua pendaftar untuk bus ini
    peserta_pulang = Pendaftaran.query.filter_by(bus_pulang_id=bus.id).all()
    peserta_kembali = Pendaftaran.query.filter_by(bus_kembali_id=bus.id).all()
    
    # --- PENGELOMPOKAN BERDASARKAN TITIK TURUN ---
    grouped_peserta_pulang = defaultdict(list)
    for p in peserta_pulang:
        grouped_peserta_pulang[p.titik_turun].append(p)
    # -----------------------------------------------

    grouped_peserta_kembali = defaultdict(list)
    for p in peserta_kembali:
        grouped_peserta_kembali[p.titik_turun].append(p)

    return render_template('detail_bus.html', 
                           bus=bus, 
                           grouped_peserta_pulang=grouped_peserta_pulang, # Kirim data terkelompok
                           grouped_peserta_kembali=grouped_peserta_kembali,
                           jumlah_peserta_pulang=len(peserta_pulang),
                           jumlah_peserta_kembali=len(peserta_kembali))

@admin_bp.route('/keuangan')
@login_required
@role_required('Korpus', 'Korwil', 'Korda', 'Bendahara Pusat', 'Sekretaris', 'Korpuspi')
def manajemen_keuangan():
    active_edisi = get_active_edisi()
    
    # Inisialisasi dictionary data yang lengkap dan benar
    data = {
        'global_total': 0, 'global_lunas': 0, 'global_belum_lunas': 0,
        'pulang_total': 0, 'pulang_lunas': 0, 'pulang_belum_lunas': 0,
        'kembali_total': 0, 'kembali_lunas': 0, 'kembali_belum_lunas': 0,
        'alokasi_bus_pulang': 0, 'alokasi_korda_pulang': 0, 'alokasi_pondok_pulang': 0,
        'alokasi_bus_kembali': 0, 'alokasi_korda_kembali': 0, 'alokasi_pondok_kembali': 0,
        'pulang_cash_count': 0, 'pulang_transfer_count': 0,
        'pulang_cash_lunas': 0, 'pulang_cash_belum_lunas': 0,
        'pulang_transfer_lunas': 0, 'pulang_transfer_belum_lunas': 0,
        'kembali_cash_count': 0, 'kembali_transfer_count': 0,
        'kembali_cash_lunas': 0, 'kembali_cash_belum_lunas': 0,
        'kembali_transfer_lunas': 0, 'kembali_transfer_belum_lunas': 0,
        'pulang_cash_lunas_rp': 0, 'pulang_cash_belum_lunas_rp': 0,
        'pulang_transfer_lunas_rp': 0, 'pulang_transfer_belum_lunas_rp': 0,
        'kembali_cash_lunas_rp': 0, 'kembali_cash_belum_lunas_rp': 0,
        'kembali_transfer_lunas_rp': 0, 'kembali_transfer_belum_lunas_rp': 0,
    }

    if not active_edisi:
        return render_template('manajemen_keuangan.html', data=data, active_edisi=None, rombongan_for_filter=[])

    # Query yang efisien untuk mengambil semua data yang dibutuhkan sekaligus
    pendaftaran_query = Pendaftaran.query.options(
        selectinload(Pendaftaran.rombongan_pulang).selectinload(Rombongan.tarifs),
        selectinload(Pendaftaran.rombongan_kembali).selectinload(Rombongan.tarifs)
    ).filter(Pendaftaran.edisi_id == active_edisi.id)

    # Filter berdasarkan role dan managed rombongan
    managed_rombongan_ids = {r.id for r in current_user.active_managed_rombongan}
    if current_user.role.name in ['Korwil', 'Korda'] and managed_rombongan_ids:
        pendaftaran_query = pendaftaran_query.filter(
            or_(
                Pendaftaran.rombongan_pulang_id.in_(managed_rombongan_ids),
                Pendaftaran.rombongan_kembali_id.in_(managed_rombongan_ids)
            )
        )
    
    all_pendaftaran = pendaftaran_query.all()

    for p in all_pendaftaran:
        # --- LOGIKA KALKULASI PULANG YANG DIPERBAIKI ---
        if p.status_pulang != 'Tidak Ikut' and p.rombongan_pulang:
            # Cek apakah user berhak mengakses rombongan pulang
            can_access_pulang = (
                current_user.role.name in ['Korpus', 'Korpuspi', 'Bendahara Pusat', 'Sekretaris'] or
                p.rombongan_pulang_id in managed_rombongan_ids
            )
            
            if can_access_pulang:
                tarif_pulang = next((t for t in p.rombongan_pulang.tarifs if t.titik_turun == p.titik_turun), None)
                if tarif_pulang:
                    biaya_pulang = tarif_pulang.harga_bus + tarif_pulang.fee_korda + 10000
                    data['pulang_total'] += biaya_pulang

                    # Alokasi fee untuk SEMUA STATUS (lunas dan belum lunas)
                    data['alokasi_bus_pulang'] += tarif_pulang.harga_bus
                    data['alokasi_korda_pulang'] += tarif_pulang.fee_korda
                    data['alokasi_pondok_pulang'] += 10000

                    # Proses pembayaran Cash
                    if p.metode_pembayaran_pulang == 'Cash':
                        data['pulang_cash_count'] += 1
                        if p.status_pulang == 'Lunas':
                            data['pulang_lunas'] += biaya_pulang
                            data['pulang_cash_lunas'] += 1
                            data['pulang_cash_lunas_rp'] += biaya_pulang
                        else: # Belum Bayar
                            data['pulang_belum_lunas'] += biaya_pulang
                            data['pulang_cash_belum_lunas'] += 1
                            data['pulang_cash_belum_lunas_rp'] += biaya_pulang
                    
                    # Proses pembayaran Transfer
                    elif p.metode_pembayaran_pulang == 'Transfer':
                        data['pulang_transfer_count'] += 1
                        if p.status_pulang == 'Lunas':
                            data['pulang_lunas'] += biaya_pulang
                            data['pulang_transfer_lunas'] += 1
                            data['pulang_transfer_lunas_rp'] += biaya_pulang
                        else: # Belum Bayar
                            data['pulang_belum_lunas'] += biaya_pulang
                            data['pulang_transfer_belum_lunas'] += 1
                            data['pulang_transfer_belum_lunas_rp'] += biaya_pulang
        
        # --- LOGIKA KALKULASI KEMBALI YANG DIPERBAIKI ---
        # Prioritas: rombongan_kembali jika ada, kalau tidak ada pakai rombongan_pulang
        rombongan_kembali = p.rombongan_kembali if p.rombongan_kembali else p.rombongan_pulang
        
        if p.status_kembali != 'Tidak Ikut' and rombongan_kembali:
            # Cek apakah user berhak mengakses rombongan kembali (lintas rombongan)
            can_access_kembali = (
                current_user.role.name in ['Korpus', 'Korpuspi', 'Bendahara Pusat', 'Sekretaris'] or
                rombongan_kembali.id in managed_rombongan_ids
            )
            
            if can_access_kembali:
                # Prioritas: titik_jemput_kembali jika ada, kalau tidak ada pakai titik_turun
                titik_jemput = p.titik_jemput_kembali if p.titik_jemput_kembali else p.titik_turun
                tarif_kembali = next((t for t in rombongan_kembali.tarifs if t.titik_turun == titik_jemput), None)
                
                if tarif_kembali:
                    biaya_kembali = tarif_kembali.harga_bus + tarif_kembali.fee_korda + 10000
                    data['kembali_total'] += biaya_kembali
                    
                    # Alokasi fee untuk SEMUA STATUS (lunas dan belum lunas)
                    data['alokasi_bus_kembali'] += tarif_kembali.harga_bus
                    data['alokasi_korda_kembali'] += tarif_kembali.fee_korda
                    data['alokasi_pondok_kembali'] += 10000
                    
                    # Proses pembayaran Cash
                    if p.metode_pembayaran_kembali == 'Cash':
                        data['kembali_cash_count'] += 1
                        if p.status_kembali == 'Lunas':
                            data['kembali_lunas'] += biaya_kembali
                            data['kembali_cash_lunas'] += 1
                            data['kembali_cash_lunas_rp'] += biaya_kembali
                        else: # Belum Bayar
                            data['kembali_belum_lunas'] += biaya_kembali
                            data['kembali_cash_belum_lunas'] += 1
                            data['kembali_cash_belum_lunas_rp'] += biaya_kembali

                    # Proses pembayaran Transfer
                    elif p.metode_pembayaran_kembali == 'Transfer':
                        data['kembali_transfer_count'] += 1
                        if p.status_kembali == 'Lunas':
                            data['kembali_lunas'] += biaya_kembali
                            data['kembali_transfer_lunas'] += 1
                            data['kembali_transfer_lunas_rp'] += biaya_kembali
                        else: # Belum Bayar
                            data['kembali_belum_lunas'] += biaya_kembali
                            data['kembali_transfer_belum_lunas'] += 1
                            data['kembali_transfer_belum_lunas_rp'] += biaya_kembali

    # Kalkulasi total global
    data['global_total'] = data['pulang_total'] + data['kembali_total']
    data['global_lunas'] = data['pulang_lunas'] + data['kembali_lunas']
    data['global_belum_lunas'] = data['global_total'] - data['global_lunas']
    
    # Daftar rombongan untuk filter
    rombongan_for_filter = []
    if current_user.role.name in ['Korwil', 'Korda']:
        rombongan_for_filter = sorted([r for r in current_user.active_managed_rombongan], key=lambda x: x.nama_rombongan)
    else:
        rombongan_for_filter = Rombongan.query.filter_by(edisi_id=active_edisi.id).order_by(Rombongan.nama_rombongan).all()

    return render_template('manajemen_keuangan.html', data=data, active_edisi=active_edisi, rombongan_for_filter=rombongan_for_filter)

@admin_bp.route('/keuangan/export-pdf')
@login_required
@role_required('Korpus', 'Korwil', 'Korda', 'Sekretaris', 'Korpuspi')
def export_keuangan_pdf():
    active_edisi = get_active_edisi()
    
    # 1. Inisialisasi dictionary data yang lengkap
    data = {
        'global_total': 0, 'global_lunas': 0, 'global_belum_lunas': 0,
        'pulang_total': 0, 'pulang_lunas': 0, 'pulang_belum_lunas': 0,
        'kembali_total': 0, 'kembali_lunas': 0, 'kembali_belum_lunas': 0,
        'alokasi_bus_pulang': 0, 'alokasi_korda_pulang': 0, 'alokasi_pondok_pulang': 0,
        'alokasi_bus_kembali': 0, 'alokasi_korda_kembali': 0, 'alokasi_pondok_kembali': 0,
        'pulang_cash_lunas_rp': 0, 'pulang_cash_belum_lunas_rp': 0,
        'pulang_transfer_lunas_rp': 0, 'pulang_transfer_belum_lunas_rp': 0,
        'kembali_cash_lunas_rp': 0, 'kembali_cash_belum_lunas_rp': 0,
        'kembali_transfer_lunas_rp': 0, 'kembali_transfer_belum_lunas_rp': 0,
    }

    if active_edisi:
        # 2. Query data pendaftaran secara efisien dengan filter Korda/Korwil
        pendaftaran_query = Pendaftaran.query.options(
            selectinload(Pendaftaran.rombongan_pulang).selectinload(Rombongan.tarifs),
            selectinload(Pendaftaran.rombongan_kembali).selectinload(Rombongan.tarifs)
        ).filter(Pendaftaran.edisi_id == active_edisi.id)

        managed_rombongan_ids = []
        if current_user.role.name in ['Korwil', 'Korda']:
            managed_rombongan_ids = [r.id for r in current_user.active_managed_rombongan]
            if not managed_rombongan_ids:
                pendaftaran_query = pendaftaran_query.filter(db.false())
            else:
                pendaftaran_query = pendaftaran_query.filter(
                    or_(
                        Pendaftaran.rombongan_pulang_id.in_(managed_rombongan_ids),
                        Pendaftaran.rombongan_kembali_id.in_(managed_rombongan_ids)
                    )
                )
        
        all_pendaftaran = pendaftaran_query.all()

        # 3. Loop dan lakukan semua kalkulasi secara lengkap dan andal
        for p in all_pendaftaran:
            # Kalkulasi untuk perjalanan pulang
            if p.status_pulang != 'Tidak Ikut' and p.rombongan_pulang:
                if current_user.role.name == 'Korpus' or p.rombongan_pulang_id in managed_rombongan_ids:
                    tarif_pulang = next((t for t in p.rombongan_pulang.tarifs if t.titik_turun == p.titik_turun), None)
                    if tarif_pulang:
                        biaya_pulang = tarif_pulang.harga_bus + tarif_pulang.fee_korda + 10000
                        data['pulang_total'] += biaya_pulang
                        
                        # Alokasi fee dihitung untuk semua status (lunas dan belum lunas)
                        data['alokasi_bus_pulang'] += tarif_pulang.harga_bus
                        data['alokasi_korda_pulang'] += tarif_pulang.fee_korda
                        data['alokasi_pondok_pulang'] += 10000
                        
                        if p.status_pulang == 'Lunas':
                            data['pulang_lunas'] += biaya_pulang
                            if p.metode_pembayaran_pulang == 'Cash': 
                                data['pulang_cash_lunas_rp'] += biaya_pulang
                            elif p.metode_pembayaran_pulang == 'Transfer': 
                                data['pulang_transfer_lunas_rp'] += biaya_pulang
                        else: # Belum Bayar
                            if p.metode_pembayaran_pulang == 'Cash': 
                                data['pulang_cash_belum_lunas_rp'] += biaya_pulang
                            elif p.metode_pembayaran_pulang == 'Transfer': 
                                data['pulang_transfer_belum_lunas_rp'] += biaya_pulang
            
            # Kalkulasi untuk perjalanan kembali
            rombongan_kembali = p.rombongan_kembali or p.rombongan_pulang
            titik_jemput = p.titik_jemput_kembali or p.titik_turun
            if p.status_kembali != 'Tidak Ikut' and rombongan_kembali and titik_jemput:
                if current_user.role.name == 'Korpus' or rombongan_kembali.id in managed_rombongan_ids:
                    tarif_kembali = next((t for t in rombongan_kembali.tarifs if t.titik_turun == titik_jemput), None)
                    if tarif_kembali:
                        biaya_kembali = tarif_kembali.harga_bus + tarif_kembali.fee_korda + 10000
                        data['kembali_total'] += biaya_kembali
                        
                        # Alokasi fee dihitung untuk semua status (lunas dan belum lunas)
                        data['alokasi_bus_kembali'] += tarif_kembali.harga_bus
                        data['alokasi_korda_kembali'] += tarif_kembali.fee_korda
                        data['alokasi_pondok_kembali'] += 10000
                        
                        if p.status_kembali == 'Lunas':
                            data['kembali_lunas'] += biaya_kembali
                            if p.metode_pembayaran_kembali == 'Cash': 
                                data['kembali_cash_lunas_rp'] += biaya_kembali
                            elif p.metode_pembayaran_kembali == 'Transfer': 
                                data['kembali_transfer_lunas_rp'] += biaya_kembali
                        else: # Belum Bayar
                            if p.metode_pembayaran_kembali == 'Cash': 
                                data['kembali_cash_belum_lunas_rp'] += biaya_kembali
                            elif p.metode_pembayaran_kembali == 'Transfer': 
                                data['kembali_transfer_belum_lunas_rp'] += biaya_kembali

        # Kalkulasi total turunan
        data['pulang_belum_lunas'] = data['pulang_total'] - data['pulang_lunas']
        data['kembali_belum_lunas'] = data['kembali_total'] - data['kembali_lunas']
        data['global_total'] = data['pulang_total'] + data['kembali_total']
        data['global_lunas'] = data['pulang_lunas'] + data['kembali_lunas']
        data['global_belum_lunas'] = data['global_total'] - data['global_lunas']

    # 4. Proses Pembuatan PDF dengan tampilan yang enhanced
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=A4, 
        rightMargin=1.5*cm, 
        leftMargin=1.5*cm, 
        topMargin=2*cm, 
        bottomMargin=2*cm
    )
    elements = []
    styles = getSampleStyleSheet()

    # Custom styles untuk tampilan yang lebih menarik
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=20,
        spaceAfter=30,
        textColor=colors.HexColor('#2c3e50'),
        alignment=1,  # Center alignment
        fontName='Helvetica-Bold'
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        spaceAfter=15,
        spaceBefore=20,
        textColor=colors.HexColor('#34495e'),
        fontName='Helvetica-Bold'
    )
    
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontSize=10,
        spaceAfter=8,
        alignment=1  # Center for subtitle
    )

    def format_rupiah(amount):
        return f"Rp {int(amount):,.0f}".replace(',', '.')

    # Header PDF dengan styling yang menarik
    title = Paragraph("LAPORAN MANAJEMEN KEUANGAN", title_style)
    elements.append(title)
    
    if active_edisi:
        subtitle_text = f"<b>Edisi: {active_edisi.nama}</b><br/>"
        subtitle_text += f"Tanggal Cetak: {datetime.now().strftime('%d %B %Y, %H:%M')} WIB<br/>"
        subtitle_text += f"Dicetak oleh: {current_user.username} ({current_user.role.name})"
        subtitle = Paragraph(subtitle_text, normal_style)
    else:
        subtitle = Paragraph("<b>Tidak ada edisi aktif</b>", normal_style)
    
    elements.append(subtitle)
    elements.append(Spacer(1, 1*cm))

    # Konten PDF yang lebih lengkap dan menarik
    if active_edisi and data['global_total'] > 0:
        
        # 1. RINGKASAN GLOBAL
        elements.append(Paragraph("RINGKASAN KEUANGAN GLOBAL", heading_style))
        
        global_data = [
            ['Kategori', 'Jumlah', 'Persentase'],
            ['Total Pemasukan', format_rupiah(data['global_total']), '100%'],
            ['Total Diterima (Lunas)', format_rupiah(data['global_lunas']), 
             f"{(data['global_lunas']/data['global_total']*100):.1f}%" if data['global_total'] > 0 else "0%"],
            ['Sisa Tagihan (Belum Lunas)', format_rupiah(data['global_belum_lunas']), 
             f"{(data['global_belum_lunas']/data['global_total']*100):.1f}%" if data['global_total'] > 0 else "0%"],
        ]
        
        global_table = Table(global_data, colWidths=[6*cm, 4*cm, 3*cm])
        global_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495e')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#ecf0f1')),
            ('BACKGROUND', (0, 2), (-1, 2), colors.HexColor('#d5f4e6')),
            ('BACKGROUND', (0, 3), (-1, 3), colors.HexColor('#fadbd8')),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(global_table)
        elements.append(Spacer(1, 0.8*cm))

        # 2. RINCIAN PERJALANAN PULANG
        elements.append(Paragraph("RINCIAN PERJALANAN PULANG", heading_style))
        
        pulang_summary_data = [
            ['Kategori', 'Jumlah'],
            ['Total Pemasukan', format_rupiah(data['pulang_total'])],
            ['Jumlah Lunas', format_rupiah(data['pulang_lunas'])],
            ['Jumlah Belum Lunas', format_rupiah(data['pulang_belum_lunas'])],
        ]
        
        pulang_summary_table = Table(pulang_summary_data, colWidths=[7*cm, 4*cm])
        pulang_summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#ebf3fd')),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(pulang_summary_table)
        elements.append(Spacer(1, 0.5*cm))

        # Detail metode pembayaran pulang
        pulang_detail_data = [
            ['Metode Pembayaran', 'Status', 'Jumlah'],
            ['Cash', 'Lunas', format_rupiah(data['pulang_cash_lunas_rp'])],
            ['Cash', 'Belum Lunas', format_rupiah(data['pulang_cash_belum_lunas_rp'])],
            ['Transfer', 'Lunas', format_rupiah(data['pulang_transfer_lunas_rp'])],
            ['Transfer', 'Belum Lunas', format_rupiah(data['pulang_transfer_belum_lunas_rp'])],
        ]
        
        pulang_detail_table = Table(pulang_detail_data, colWidths=[4*cm, 3*cm, 4*cm])
        pulang_detail_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2980b9')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (0, 2), colors.HexColor('#f8f9fa')),  # Cash rows
            ('BACKGROUND', (0, 3), (0, 4), colors.HexColor('#e3f2fd')),  # Transfer rows
            ('BACKGROUND', (1, 1), (1, 1), colors.HexColor('#d5f4e6')),  # Lunas
            ('BACKGROUND', (1, 2), (1, 2), colors.HexColor('#fadbd8')),  # Belum Lunas
            ('BACKGROUND', (1, 3), (1, 3), colors.HexColor('#d5f4e6')),  # Lunas
            ('BACKGROUND', (1, 4), (1, 4), colors.HexColor('#fadbd8')),  # Belum Lunas
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(pulang_detail_table)
        elements.append(Spacer(1, 0.8*cm))

        # 3. RINCIAN PERJALANAN KEMBALI
        elements.append(Paragraph("RINCIAN PERJALANAN KEMBALI", heading_style))
        
        kembali_summary_data = [
            ['Kategori', 'Jumlah'],
            ['Total Pemasukan', format_rupiah(data['kembali_total'])],
            ['Jumlah Lunas', format_rupiah(data['kembali_lunas'])],
            ['Jumlah Belum Lunas', format_rupiah(data['kembali_belum_lunas'])],
        ]
        
        kembali_summary_table = Table(kembali_summary_data, colWidths=[7*cm, 4*cm])
        kembali_summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e74c3c')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#fdeaea')),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(kembali_summary_table)
        elements.append(Spacer(1, 0.5*cm))

        # Detail metode pembayaran kembali
        kembali_detail_data = [
            ['Metode Pembayaran', 'Status', 'Jumlah'],
            ['Cash', 'Lunas', format_rupiah(data['kembali_cash_lunas_rp'])],
            ['Cash', 'Belum Lunas', format_rupiah(data['kembali_cash_belum_lunas_rp'])],
            ['Transfer', 'Lunas', format_rupiah(data['kembali_transfer_lunas_rp'])],
            ['Transfer', 'Belum Lunas', format_rupiah(data['kembali_transfer_belum_lunas_rp'])],
        ]
        
        kembali_detail_table = Table(kembali_detail_data, colWidths=[4*cm, 3*cm, 4*cm])
        kembali_detail_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#c0392b')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (0, 2), colors.HexColor('#f8f9fa')),  # Cash rows
            ('BACKGROUND', (0, 3), (0, 4), colors.HexColor('#ffe8e8')),  # Transfer rows
            ('BACKGROUND', (1, 1), (1, 1), colors.HexColor('#d5f4e6')),  # Lunas
            ('BACKGROUND', (1, 2), (1, 2), colors.HexColor('#fadbd8')),  # Belum Lunas
            ('BACKGROUND', (1, 3), (1, 3), colors.HexColor('#d5f4e6')),  # Lunas
            ('BACKGROUND', (1, 4), (1, 4), colors.HexColor('#fadbd8')),  # Belum Lunas
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(kembali_detail_table)
        elements.append(Spacer(1, 0.8*cm))

        # 4. ALOKASI FEE
        elements.append(Paragraph("ALOKASI FEE", heading_style))
        
        fee_data = [
            ['Kategori', 'Pulang', 'Kembali', 'Total'],
            ['Biaya Bus', 
             format_rupiah(data['alokasi_bus_pulang']), 
             format_rupiah(data['alokasi_bus_kembali']), 
             format_rupiah(data['alokasi_bus_pulang'] + data['alokasi_bus_kembali'])],
            ['Fee Korda', 
             format_rupiah(data['alokasi_korda_pulang']), 
             format_rupiah(data['alokasi_korda_kembali']), 
             format_rupiah(data['alokasi_korda_pulang'] + data['alokasi_korda_kembali'])],
            ['Fee Pondok', 
             format_rupiah(data['alokasi_pondok_pulang']), 
             format_rupiah(data['alokasi_pondok_kembali']), 
             format_rupiah(data['alokasi_pondok_pulang'] + data['alokasi_pondok_kembali'])],
            ['TOTAL ALOKASI', 
             format_rupiah(data['pulang_total']), 
             format_rupiah(data['kembali_total']), 
             format_rupiah(data['global_total'])],
        ]
        
        fee_table = Table(fee_data, colWidths=[4*cm, 3*cm, 3*cm, 3*cm])
        fee_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#9b59b6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, 3), colors.HexColor('#f4ecf7')),
            ('BACKGROUND', (0, 4), (-1, 4), colors.HexColor('#8e44ad')),
            ('TEXTCOLOR', (0, 4), (-1, 4), colors.whitesmoke),
            ('FONTNAME', (0, 4), (-1, 4), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(fee_table)

    else:
        # No data available
        no_data_style = ParagraphStyle(
            'NoData',
            parent=styles['Normal'],
            fontSize=12,
            alignment=1,
            spaceAfter=20,
            textColor=colors.HexColor('#e74c3c')
        )
        elements.append(Paragraph("Tidak ada data keuangan untuk ditampilkan.", no_data_style))
        elements.append(Paragraph("Pastikan ada edisi yang aktif dan terdapat data pendaftaran.", styles['Normal']))

    # Footer
    elements.append(Spacer(1, 1*cm))
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.grey,
        alignment=1,
        borderWidth=1,
        borderColor=colors.grey,
        borderPadding=5
    )
    
    footer_text = f"Laporan ini digenerate secara otomatis oleh sistem pada {datetime.now().strftime('%d %B %Y, %H:%M:%S')} WIB"
    elements.append(Paragraph(footer_text, footer_style))
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    filename = f"laporan_keuangan_{active_edisi.nama if active_edisi else 'no_edition'}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    
    return send_file(
        buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=filename
    )

@admin_bp.route('/api/rombongan-detail/<int:rombongan_id>')
@login_required
def api_rombongan_detail(rombongan_id):
    rombongan = Rombongan.query.get_or_404(rombongan_id)
    
    tarifs_data = [{'titik_turun': t.titik_turun} for t in rombongan.tarifs]
    
    buses_data = []
    for bus in rombongan.buses:
        terisi_kembali = Pendaftaran.query.filter_by(bus_kembali_id=bus.id).count()
        sisa_kuota_kembali = bus.kuota - terisi_kembali
        buses_data.append({
            'id': bus.id,
            'nama_armada': bus.nama_armada,
            'nomor_lambung': bus.nomor_lambung,
            'plat_nomor': bus.plat_nomor,
            'sisa_kuota_kembali': sisa_kuota_kembali
        })

    return jsonify({
        'tarifs': tarifs_data,
        'buses': buses_data
    })

# Hapus fungsi-fungsi korlapda lama dan ganti dengan 3 fungsi ini

@admin_bp.route('/petugas-lapangan')
@login_required
@role_required('Korpus', 'Korda')
def manajemen_petugas():
    """Halaman utama untuk melihat semua petugas lapangan (Korlapda & Sarpras)."""
    if current_user.role.name == 'Korda':
        # Korda hanya melihat petugas di rombongannya
        managed_rombongan_ids = [r.id for r in current_user.active_managed_rombongan]
        all_petugas = User.query.join(Role).filter(
            Role.name.in_(['Korlapda', 'Sarpras']),
            User.managed_rombongan.any(Rombongan.id.in_(managed_rombongan_ids))
        ).all()
    else: # Korpus melihat semua
        all_petugas = User.query.join(Role).filter(Role.name.in_(['Korlapda', 'Sarpras'])).all()

    return render_template('manajemen_petugas.html', all_petugas=all_petugas)


@admin_bp.route('/petugas-lapangan/tambah', methods=['GET', 'POST'])
@login_required
@role_required('Korpus', 'Korda')
def tambah_petugas():
    form = PetugasLapanganForm()

    # --- Logika untuk mengisi pilihan dropdown ---
    # Ambil role Korlapda dan Sarpras
    field_roles = Role.query.filter(Role.name.in_(['Korlapda', 'Sarpras'])).all()
    form.role.choices = [(r.id, r.name) for r in field_roles]

    if current_user.role.name == 'Korda':
        # Korda hanya bisa menugaskan ke bus & rombongan miliknya
        buses = Bus.query.filter(Bus.rombongan_id.in_([r.id for r in current_user.active_managed_rombongan])).all()
        form.bus.choices = [(b.id, f"{b.nama_armada} ({b.rombongan.nama_rombongan})") for b in buses]
        form.managed_rombongan.choices = [(r.id, r.nama_rombongan) for r in current_user.active_managed_rombongan]
    else: # Korpus
        buses = Bus.query.all()
        form.bus.choices = [(b.id, f"{b.nama_armada} ({b.rombongan.nama_rombongan})") for b in buses]
        form.managed_rombongan.choices = [(r.id, r.nama_rombongan) for r in Rombongan.query.all()]
    # --- Akhir logika dropdown ---

    if form.validate_on_submit():
        role = Role.query.get(form.role.data)

        new_user = User(username=form.username.data, role=role)
        new_user.set_password(form.password.data)

        # Tugaskan ke bus (jika Korlapda) atau rombongan (jika Sarpras)
        if role.name == 'Korlapda':
            bus = Bus.query.get(form.bus.data)
            if bus:
                new_user.bus = bus
                new_user.managed_rombongan.append(bus.rombongan)
        elif role.name == 'Sarpras':
            rombongan = Rombongan.query.get(form.managed_rombongan.data)
            if rombongan:
                new_user.managed_rombongan.append(rombongan)

        db.session.add(new_user)
        db.session.commit()
        flash(f'User petugas lapangan "{new_user.username}" berhasil dibuat.', 'success')
        return redirect(url_for('admin.manajemen_petugas'))

    return render_template('form_petugas.html', title="Tambah Petugas Lapangan", form=form)


@admin_bp.route('/petugas-lapangan/hapus/<int:user_id>', methods=['POST'])
@login_required
@role_required('Korpus', 'Korda')
def hapus_petugas(user_id):
    user = User.query.get_or_404(user_id)
    # Tambahkan validasi keamanan agar Korda tidak bisa hapus petugas di luar rombongannya
    db.session.delete(user)
    db.session.commit()
    flash(f'User "{user.username}" berhasil dihapus.', 'success')
    return redirect(url_for('admin.manajemen_petugas'))

# utils/audit.py
from flask import current_app, request
from flask_login import current_user
from sqlalchemy.exc import SQLAlchemyError
from app import db
from app.models import ActivityLog

def log_activity(action_type, feature, description=None):
    try:
        # user_id NOT NULL -> pastikan login
        if getattr(current_user, "id", None) is None:
            current_app.logger.warning("Skip log_activity: user belum login.")
            return None

        entry = ActivityLog(
            user_id=current_user.id,
            action_type=action_type,   # <- sesuai model
            feature=feature,
            description=description,   # <- sesuai model
            # timestamp pakai default=datetime.utcnow di model
        )
        db.session.add(entry)
        db.session.commit()
        return entry.id
    except SQLAlchemyError as e:
        db.session.rollback()
        current_app.logger.error("Gagal log_activity: %s", e, exc_info=True)
        return None 
    

@admin_bp.route('/log-aktivitas')
@login_required
@role_required('Korpus', 'Korda', 'Korwil', 'Keamanan', 'PJ Acara', 'Bendahara Pusat', 'Sekretaris', 'Korpuspi')
def log_aktivitas():
    page = request.args.get('page', 1, type=int)
    
    # Query dasar untuk semua log
    query = ActivityLog.query.order_by(ActivityLog.timestamp.desc())

    # --- FILTER BERDASARKAN HAK AKSES ---
    role = current_user.role.name
    if role == 'Korwil' or role == 'Korda':
        # Filter log yang deskripsinya mengandung nama rombongan yang dikelola
        managed_rombongan_names = [r.nama_rombongan for r in current_user.active_managed_rombongan]
        if managed_rombongan_names:
            query = query.filter(or_(*[ActivityLog.description.like(f'%{name}%') for name in managed_rombongan_names]))
        else:
            query = query.filter(db.false()) # Jika tidak kelola rombongan, jangan tampilkan apa-apa
    elif role == 'Keamanan':
        query = query.filter(ActivityLog.feature == 'Perizinan')
    elif role == 'PJ Acara':
        query = query.filter(ActivityLog.feature == 'Partisipan')
    
    # --- FILTER DARI INPUT USER ---
    f_user_id = request.args.get('user_id', type=int)
    f_action_type = request.args.get('action_type')
    f_start_date = request.args.get('start_date')
    
    if f_user_id:
        query = query.filter_by(user_id=f_user_id)
    if f_action_type:
        query = query.filter_by(action_type=f_action_type)
    if f_start_date:
        start_date = datetime.strptime(f_start_date, '%Y-%m-%d')
        query = query.filter(ActivityLog.timestamp >= start_date)

    logs = query.paginate(page=page, per_page=50, error_out=False)
    
    # Ambil data untuk mengisi dropdown filter
    all_users = User.query.order_by(User.username).all()
    action_types = db.session.query(ActivityLog.action_type).distinct().all()

    return render_template(
        'log_aktivitas.html', 
        logs=logs,
        all_users=all_users,
        action_types=[a[0] for a in action_types]
    )

@admin_bp.route('/santri-wilayah')
@login_required
@role_required('Korpus', 'Korwil', 'Korda', 'Bendahara Pusat', 'Sekretaris', 'Korpuspi')
def manajemen_santri_wilayah():
    """Menampilkan daftar santri yang relevan dengan wilayah Korda/Korwil dengan statistik dan filter detail."""
    active_edisi = get_active_edisi()
    if not active_edisi:
        flash('Tidak ada edisi yang aktif.', 'warning')
        return render_template('manajemen_santri_wilayah.html', pagination=None, stats={}, pendaftaran_map={})

    # 1. Tentukan santri dan rombongan yang relevan berdasarkan cakupan wilayah user
    santri_in_wilayah_ids = []
    managed_rombongan_ids = []
    
    if current_user.role.name in ['Korwil', 'Korda']:
        managed_rombongan_ids = [r.id for r in current_user.active_managed_rombongan]
        wilayah_kelolaan = set()
        for rombongan in current_user.active_managed_rombongan:
            if rombongan.cakupan_wilayah:
                for wilayah in rombongan.cakupan_wilayah:
                    wilayah_kelolaan.add(wilayah.get('label'))
        
        if wilayah_kelolaan:
            santri_in_wilayah_ids = [s.id for s in Santri.query.with_entities(Santri.id).filter(Santri.kabupaten.in_(list(wilayah_kelolaan))).all()]
    else: # Korpus melihat semua santri
        santri_in_wilayah_ids = [s.id for s in Santri.query.with_entities(Santri.id).all()]

    # 2. Hitung semua statistik yang dibutuhkan
    stats = {}
    if santri_in_wilayah_ids:
        # Statistik Total Santri di Wilayah
        stats['total_putra'] = Santri.query.filter(Santri.id.in_(santri_in_wilayah_ids), Santri.jenis_kelamin == 'PUTRA').count()
        stats['total_putri'] = Santri.query.filter(Santri.id.in_(santri_in_wilayah_ids), Santri.jenis_kelamin == 'PUTRI').count()
        
        # Logika perhitungan statistik pendaftaran yang presisi
        base_pendaftaran_query = Pendaftaran.query.filter(
            Pendaftaran.edisi_id == active_edisi.id,
            Pendaftaran.santri_id.in_(santri_in_wilayah_ids)
        )

        if current_user.role.name in ['Korwil', 'Korda'] and managed_rombongan_ids:
            # Untuk Korda/Korwil, hitung hanya yang terkait rombongan mereka
            pendaftar_ids_managed = db.session.query(Pendaftaran.santri_id).filter(
                Pendaftaran.edisi_id == active_edisi.id,
                Pendaftaran.santri_id.in_(santri_in_wilayah_ids),
                or_(
                    Pendaftaran.rombongan_pulang_id.in_(managed_rombongan_ids),
                    Pendaftaran.rombongan_kembali_id.in_(managed_rombongan_ids)
                )
            ).distinct().all()
            pendaftar_ids = [p[0] for p in pendaftar_ids_managed]

            # Detail Perjalanan (hanya dari rombongan yang dikelola)
            stats['total_pulang_putra'] = base_pendaftaran_query.join(Pendaftaran.santri).filter(Pendaftaran.rombongan_pulang_id.in_(managed_rombongan_ids), Pendaftaran.status_pulang != 'Tidak Ikut', Santri.jenis_kelamin == 'PUTRA').count()
            stats['total_pulang_putri'] = base_pendaftaran_query.join(Pendaftaran.santri).filter(Pendaftaran.rombongan_pulang_id.in_(managed_rombongan_ids), Pendaftaran.status_pulang != 'Tidak Ikut', Santri.jenis_kelamin == 'PUTRI').count()
            stats['total_kembali_putra'] = base_pendaftaran_query.join(Pendaftaran.santri).filter(Pendaftaran.rombongan_kembali_id.in_(managed_rombongan_ids), Pendaftaran.status_kembali != 'Tidak Ikut', Santri.jenis_kelamin == 'PUTRA').count()
            stats['total_kembali_putri'] = base_pendaftaran_query.join(Pendaftaran.santri).filter(Pendaftaran.rombongan_kembali_id.in_(managed_rombongan_ids), Pendaftaran.status_kembali != 'Tidak Ikut', Santri.jenis_kelamin == 'PUTRI').count()
        else: # Untuk Korpus, hitung semua pendaftaran di wilayah tersebut
            pendaftar_ids_all = db.session.query(Pendaftaran.santri_id).filter(Pendaftaran.edisi_id == active_edisi.id, Pendaftaran.santri_id.in_(santri_in_wilayah_ids)).distinct().all()
            pendaftar_ids = [p[0] for p in pendaftar_ids_all]
            stats['total_pulang_putra'] = base_pendaftaran_query.join(Pendaftaran.santri).filter(Pendaftaran.status_pulang != 'Tidak Ikut', Santri.jenis_kelamin == 'PUTRA').count()
            stats['total_pulang_putri'] = base_pendaftaran_query.join(Pendaftaran.santri).filter(Pendaftaran.status_pulang != 'Tidak Ikut', Santri.jenis_kelamin == 'PUTRI').count()
            stats['total_kembali_putra'] = base_pendaftaran_query.join(Pendaftaran.santri).filter(Pendaftaran.status_kembali != 'Tidak Ikut', Santri.jenis_kelamin == 'PUTRA').count()
            stats['total_kembali_putri'] = base_pendaftaran_query.join(Pendaftaran.santri).filter(Pendaftaran.status_kembali != 'Tidak Ikut', Santri.jenis_kelamin == 'PUTRI').count()
        if current_user.role.name in ['Korwil', 'Korda'] and managed_rombongan_ids:
    # Hitung yang sudah daftar pulang per jenis kelamin (dari rombongan yang dikelola)
            stats['putra_daftar_pulang'] = base_pendaftaran_query.join(Pendaftaran.santri).filter(
                Pendaftaran.rombongan_pulang_id.in_(managed_rombongan_ids), 
                Pendaftaran.status_pulang != 'Tidak Ikut', 
                Santri.jenis_kelamin == 'PUTRA'
            ).distinct(Pendaftaran.santri_id).count()
    
            stats['putri_daftar_pulang'] = base_pendaftaran_query.join(Pendaftaran.santri).filter(
                Pendaftaran.rombongan_pulang_id.in_(managed_rombongan_ids), 
                Pendaftaran.status_pulang != 'Tidak Ikut', 
                Santri.jenis_kelamin == 'PUTRI'
            ).distinct(Pendaftaran.santri_id).count()
    
    # Hitung yang sudah daftar kembali per jenis kelamin (dari rombongan yang dikelola)
            stats['putra_daftar_kembali'] = base_pendaftaran_query.join(Pendaftaran.santri).filter(
                Pendaftaran.rombongan_kembali_id.in_(managed_rombongan_ids), 
                Pendaftaran.status_kembali != 'Tidak Ikut', 
                Santri.jenis_kelamin == 'PUTRA'
            ).distinct(Pendaftaran.santri_id).count()
    
            stats['putri_daftar_kembali'] = base_pendaftaran_query.join(Pendaftaran.santri).filter(
                Pendaftaran.rombongan_kembali_id.in_(managed_rombongan_ids), 
                Pendaftaran.status_kembali != 'Tidak Ikut', 
                Santri.jenis_kelamin == 'PUTRI'
            ).distinct(Pendaftaran.santri_id).count()
    
        else:  # Untuk Korpus
            stats['putra_daftar_pulang'] = base_pendaftaran_query.join(Pendaftaran.santri).filter(
                Pendaftaran.status_pulang != 'Tidak Ikut', 
                Santri.jenis_kelamin == 'PUTRA'
            ).distinct(Pendaftaran.santri_id).count()
    
            stats['putri_daftar_pulang'] = base_pendaftaran_query.join(Pendaftaran.santri).filter(
                Pendaftaran.status_pulang != 'Tidak Ikut', 
                Santri.jenis_kelamin == 'PUTRI'
            ).distinct(Pendaftaran.santri_id).count()
    
            stats['putra_daftar_kembali'] = base_pendaftaran_query.join(Pendaftaran.santri).filter(
                Pendaftaran.status_kembali != 'Tidak Ikut', 
                Santri.jenis_kelamin == 'PUTRA'
            ).distinct(Pendaftaran.santri_id).count()
    
            stats['putri_daftar_kembali'] = base_pendaftaran_query.join(Pendaftaran.santri).filter(
                Pendaftaran.status_kembali != 'Tidak Ikut', 
                Santri.jenis_kelamin == 'PUTRI'
            ).distinct(Pendaftaran.santri_id).count()
        # Update stats turunan berdasarkan pendaftar yang relevan
        stats['putra_terdaftar'] = Santri.query.filter(Santri.id.in_(pendaftar_ids), Santri.jenis_kelamin == 'PUTRA').count()
        stats['putri_terdaftar'] = Santri.query.filter(Santri.id.in_(pendaftar_ids), Santri.jenis_kelamin == 'PUTRI').count()
        stats['putra_belum_daftar'] = stats['total_putra'] - stats['putra_terdaftar']
        stats['putri_belum_daftar'] = stats['total_putri'] - stats['putri_terdaftar']
        stats['total_pendaftaran'] = len(pendaftar_ids)

    # 3. Siapkan query dasar untuk daftar santri di tabel
    query = Santri.query.filter(Santri.id.in_(santri_in_wilayah_ids))

    # 4. Terapkan filter dari form
    nama_filter = request.args.get('nama')
    gender_filter = request.args.get('jenis_kelamin')
    status_santri_filter = request.args.get('status_santri')
    status_daftar_filter = request.args.get('status_daftar')

    if nama_filter:
        query = query.filter(Santri.nama.ilike(f'%{nama_filter}%'))
    if gender_filter:
        query = query.filter(Santri.jenis_kelamin == gender_filter)
    if status_santri_filter:
        query = query.filter(Santri.status_santri == status_santri_filter)
    
    # Blok logika filter pendaftaran baru yang andal
    if status_daftar_filter and santri_in_wilayah_ids:
        rombongan_filter_pulang = db.true()
        rombongan_filter_kembali = db.true()
        if current_user.role.name in ['Korwil', 'Korda'] and managed_rombongan_ids:
            rombongan_filter_pulang = Pendaftaran.rombongan_pulang_id.in_(managed_rombongan_ids)
            rombongan_filter_kembali = or_(
                Pendaftaran.rombongan_kembali_id.in_(managed_rombongan_ids),
                and_(Pendaftaran.rombongan_kembali_id.is_(None), Pendaftaran.rombongan_pulang_id.in_(managed_rombongan_ids))
            )

        if status_daftar_filter == 'sudah_pulang':
            santri_ids = [s_id for s_id, in db.session.query(Pendaftaran.santri_id).filter(Pendaftaran.edisi_id == active_edisi.id, Pendaftaran.santri_id.in_(santri_in_wilayah_ids), Pendaftaran.status_pulang != 'Tidak Ikut', rombongan_filter_pulang).distinct()]
            query = query.filter(Santri.id.in_(santri_ids))
        elif status_daftar_filter == 'belum_pulang':
            santri_ids_sudah = [s_id for s_id, in db.session.query(Pendaftaran.santri_id).filter(Pendaftaran.edisi_id == active_edisi.id, Pendaftaran.santri_id.in_(santri_in_wilayah_ids), Pendaftaran.status_pulang != 'Tidak Ikut', rombongan_filter_pulang).distinct()]
            query = query.filter(Santri.id.notin_(santri_ids_sudah))
        elif status_daftar_filter == 'sudah_kembali':
            santri_ids = [s_id for s_id, in db.session.query(Pendaftaran.santri_id).filter(Pendaftaran.edisi_id == active_edisi.id, Pendaftaran.santri_id.in_(santri_in_wilayah_ids), Pendaftaran.status_kembali != 'Tidak Ikut', rombongan_filter_kembali).distinct()]
            query = query.filter(Santri.id.in_(santri_ids))
        elif status_daftar_filter == 'belum_kembali':
            santri_ids_sudah = [s_id for s_id, in db.session.query(Pendaftaran.santri_id).filter(Pendaftaran.edisi_id == active_edisi.id, Pendaftaran.santri_id.in_(santri_in_wilayah_ids), Pendaftaran.status_kembali != 'Tidak Ikut', rombongan_filter_kembali).distinct()]
            query = query.filter(Santri.id.notin_(santri_ids_sudah))
    
    # 5. Lakukan paginasi
    page = request.args.get('page', 1, type=int)
    per_page = 50
    pagination = query.order_by(Santri.nama).paginate(page=page, per_page=per_page, error_out=False)
    
    # 6. Ambil data pendaftaran yang relevan untuk santri di halaman saat ini
    santri_ids_on_page = [s.id for s in pagination.items]
    pendaftaran_map = {}
    if santri_ids_on_page:
        pendaftarans_on_page = Pendaftaran.query.options(
            selectinload(Pendaftaran.bus_pulang),
            selectinload(Pendaftaran.bus_kembali)
        ).filter(
            Pendaftaran.edisi_id == active_edisi.id,
            Pendaftaran.santri_id.in_(santri_ids_on_page)
        ).all()
        pendaftaran_map = {p.santri_id: p for p in pendaftarans_on_page}

    # 7. Render template dengan semua data
    return render_template('manajemen_santri_wilayah.html',
                           pagination=pagination,
                           stats=stats,
                           pendaftaran_map=pendaftaran_map,
                           active_edisi=active_edisi)

@admin_bp.route('/laporan-wali')
@login_required
@role_required('Korda', 'Korwil')
def laporan_grup_wali():
    active_edisi = get_active_edisi()
    if not active_edisi:
        return jsonify({'error': 'Edisi tidak aktif'}), 404

    # 1. Ambil data dasar Korda/Korwil
    managed_rombongan = current_user.active_managed_rombongan
    if not managed_rombongan:
        return jsonify({'report_text': 'Anda tidak mengelola rombongan apapun.'})
    
    managed_rombongan_ids = [r.id for r in managed_rombongan]
    grup_url = managed_rombongan[0].grup_wali_url if managed_rombongan[0].grup_wali_url else None
    
    managed_regions = {
        wilayah.get('label')
        for rombongan in managed_rombongan
        if rombongan.cakupan_wilayah
        for wilayah in rombongan.cakupan_wilayah
    }

    # 2. Ambil semua pendaftaran yang relevan
    all_pendaftaran = Pendaftaran.query.options(
        joinedload(Pendaftaran.santri),
        joinedload(Pendaftaran.rombongan_pulang)
    ).filter(Pendaftaran.edisi_id == active_edisi.id).all()
    
    # 3. Ambil semua santri dari wilayah Korda (kecuali yang Izin)
    all_santri_in_region = Santri.query.filter(
        Santri.kabupaten.in_(list(managed_regions)),
        Santri.status_santri != 'Izin'
    ).all()
    
    # Buat map untuk akses cepat
    pendaftaran_map = {p.santri_id: p for p in all_pendaftaran}
    
    # 4. Kategorikan santri berdasarkan logika baru
    sudah_daftar_rombongan_ini = []
    sudah_daftar_rombongan_lain = []
    belum_daftar_partisipan = []
    belum_daftar_wisuda = []
    belum_daftar_aktif = []

    # Proses santri yang terdaftar di rombongan Korda
    for p in all_pendaftaran:
        is_in_managed_rombongan = p.rombongan_pulang_id in managed_rombongan_ids or p.rombongan_kembali_id in managed_rombongan_ids
        if is_in_managed_rombongan:
            is_from_region = p.santri.kabupaten in managed_regions
            tanda = "" if is_from_region else "(!)" # Tanda (*) untuk santri dari luar wilayah
            sudah_daftar_rombongan_ini.append((p, tanda))

    # Proses santri dari wilayah Korda
    santri_terdaftar_di_rombongan_ini_ids = {p.santri_id for p, tanda in sudah_daftar_rombongan_ini}
    
    for santri in all_santri_in_region:
        if santri.id in pendaftaran_map:
            # Jika santri ada di map pendaftaran TAPI tidak di rombongan ini, berarti dia ikut rombongan lain
            if santri.id not in santri_terdaftar_di_rombongan_ini_ids:
                 p_lain = pendaftaran_map[santri.id]
                 nama_romb_lain = p_lain.rombongan_pulang.nama_rombongan if p_lain.rombongan_pulang else "Lain"
                 sudah_daftar_rombongan_lain.append((santri, nama_romb_lain))
        else:
            # Santri dari wilayah ini dan belum terdaftar sama sekali
            if santri.status_santri == 'Partisipan':
                belum_daftar_partisipan.append(santri)
            elif santri.status_santri == 'Wisuda':
                belum_daftar_wisuda.append(santri)
            else: # Aktif
                belum_daftar_aktif.append(santri)
    
    # 5. Susun teks laporan
    today_date = datetime.now().strftime('%d %B %Y')
    nama_rombongan_str = ", ".join(r.nama_rombongan for r in managed_rombongan)
    
    report_lines = [
        f"Assalamualaikum wr. wb.",
        f"Berikut rekap pendaftaran santri *Rombongan {nama_rombongan_str}* per *{today_date}*:\n",
        "*SUDAH TERDAFTAR*",
        "-------------------"
    ]
    
    # Urutkan berdasarkan nama
    sudah_daftar_rombongan_ini.sort(key=lambda x: x[0].santri.nama)
    sudah_daftar_rombongan_lain.sort(key=lambda x: x[0].nama)

    for p, tanda in sudah_daftar_rombongan_ini:
        # (Logika emoji dan payment_str tetap sama)
        lunas_p = p.status_pulang == 'Lunas'
        lunas_k = p.status_kembali == 'Lunas'
        emoji = "✅✅" if lunas_p and lunas_k else "✅" if lunas_p or lunas_k else ""
        metode_p = 'C' if p.metode_pembayaran_pulang == 'Cash' else 'TF' if p.metode_pembayaran_pulang == 'Transfer' else '-'
        metode_k = 'C' if p.metode_pembayaran_kembali == 'Cash' else 'TF' if p.metode_pembayaran_kembali == 'Transfer' else '-'
        payment_str = f"({metode_p}, {metode_k})" if p.status_pulang != 'Tidak Ikut' and p.status_kembali != 'Tidak Ikut' else f"({metode_p})" if p.status_pulang != 'Tidak Ikut' else f"({metode_k})" if p.status_kembali != 'Tidak Ikut' else ""
        report_lines.append(f"{emoji} {p.santri.nama} {tanda} {payment_str}".strip())

    for santri, nama_romb_lain in sudah_daftar_rombongan_lain:
        report_lines.append(f"- {santri.nama} _(ikut rombongan {nama_romb_lain})_")
        
    # Tambahkan bagian BELUM DAFTAR jika ada isinya
    def add_section(title, santri_list):
        if santri_list:
            report_lines.append(f"\n*{title}*")
            report_lines.append("-------------------")
            santri_list.sort(key=lambda s: s.nama)
            for santri in santri_list:
                report_lines.append(f"- {santri.nama}")

    add_section("PARTISIPAN", belum_daftar_partisipan)
    add_section("WISUDA", belum_daftar_wisuda)
    add_section("BELUM DAFTAR", belum_daftar_aktif)

    report_lines.extend([
        "\nKeterangan:", "✅✅: Lunas Pulang & Kembali", "✅: Lunas salah satu",
        "(!): Santri dari luar daerah yang ikut rombongan", "(nama rombongan): Ikut rombongan lain",
        "(C): Cash", "(TF): Transfer",
        "\nTerima kasih.", "Wassalamualaikum wr. wb."
    ])
    
    final_report_text = "\n".join(report_lines)
    
    return jsonify({'report_text': final_report_text, 'grup_url': grup_url})

@admin_bp.route('/rombongan/<int:rombongan_id>/url-grup', methods=['GET', 'POST'])
@login_required
@role_required('Korpus', 'Korda', 'Korwil')
def atur_url_grup(rombongan_id):
    rombongan = Rombongan.query.get_or_404(rombongan_id)
    
    # Keamanan: Pastikan Korda hanya bisa mengedit rombongan mereka
    if current_user.role.name in ['Korda', 'Korwil'] and rombongan not in current_user.active_managed_rombongan:
        flash('Anda tidak memiliki hak akses untuk rombongan ini.', 'danger')
        return redirect(url_for('admin.manajemen_rombongan'))

    form = GrupWaliURLForm(obj=rombongan)
    if form.validate_on_submit():
        rombongan.grup_wali_url = form.grup_wali_url.data
        db.session.commit()
        flash(f'Link grup WhatsApp untuk {rombongan.nama_rombongan} berhasil diperbarui.', 'success')
        return redirect(url_for('admin.manajemen_rombongan'))
        
    return render_template('atur_url_grup.html', form=form, rombongan=rombongan, title="Atur Link Grup Wali")

@admin_bp.route('/export-santri-wilayah')
@login_required
@role_required('Korpus', 'Korwil', 'Korda', 'Sekretaris', 'Korpuspi')
def export_santri_wilayah():
    """Export data santri wilayah ke Excel dengan filter yang sama - SEMUA SANTRI AKAN TERMUAT"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from flask import make_response
    from io import BytesIO
    import datetime
    
    # Get active edition
    active_edisi = get_active_edisi()
    if not active_edisi:
        flash('Tidak ada edisi yang aktif.', 'warning')
        return redirect(url_for('admin.manajemen_santri_wilayah'))

    # Get export type (pulang atau kembali)
    export_type = request.args.get('type', 'pulang')  # default pulang
    
    # Get the same santri and rombongan data as the main view
    santri_in_wilayah_ids = []
    managed_rombongan_ids = []
    
    if current_user.role.name in ['Korwil', 'Korda']:
        managed_rombongan_ids = [r.id for r in current_user.active_managed_rombongan]
        wilayah_kelolaan = set()
        for rombongan in current_user.active_managed_rombongan:
            if rombongan.cakupan_wilayah:
                for wilayah in rombongan.cakupan_wilayah:
                    wilayah_kelolaan.add(wilayah.get('label'))
        
        if wilayah_kelolaan:
            santri_in_wilayah_ids = [s.id for s in Santri.query.with_entities(Santri.id).filter(Santri.kabupaten.in_(list(wilayah_kelolaan))).all()]
    else:
        santri_in_wilayah_ids = [s.id for s in Santri.query.with_entities(Santri.id).all()]

    # MODIFIED: Get ALL santri in wilayah first (no registration filter)
    base_query = Santri.query.filter(Santri.id.in_(santri_in_wilayah_ids))
    
    # Apply basic filters (nama, gender, status_santri)
    nama_filter = request.args.get('nama')
    gender_filter = request.args.get('jenis_kelamin')
    status_santri_filter = request.args.get('status_santri')
    status_daftar_filter = request.args.get('status_daftar')

    if nama_filter:
        base_query = base_query.filter(Santri.nama.ilike(f'%{nama_filter}%'))
    if gender_filter:
        base_query = base_query.filter(Santri.jenis_kelamin == gender_filter)
    if status_santri_filter:
        base_query = base_query.filter(Santri.status_santri == status_santri_filter)
    
    # Get all santri matching basic filters
    all_santri = base_query.order_by(Santri.nama).all()
    
    # MODIFIED: Apply registration status filter ONLY for filtering, but keep all santri for export
    filtered_santri_ids = set([s.id for s in all_santri])  # Start with all santri
    
    if status_daftar_filter and santri_in_wilayah_ids:
        rombongan_filter_pulang = db.true()
        rombongan_filter_kembali = db.true()
        if current_user.role.name in ['Korwil', 'Korda'] and managed_rombongan_ids:
            rombongan_filter_pulang = Pendaftaran.rombongan_pulang_id.in_(managed_rombongan_ids)
            rombongan_filter_kembali = or_(
                Pendaftaran.rombongan_kembali_id.in_(managed_rombongan_ids),
                and_(Pendaftaran.rombongan_kembali_id.is_(None), Pendaftaran.rombongan_pulang_id.in_(managed_rombongan_ids))
            )

        if status_daftar_filter == 'sudah_pulang':
            santri_ids = [s_id for s_id, in db.session.query(Pendaftaran.santri_id).filter(Pendaftaran.edisi_id == active_edisi.id, Pendaftaran.santri_id.in_(santri_in_wilayah_ids), Pendaftaran.status_pulang != 'Tidak Ikut', rombongan_filter_pulang).distinct()]
            filtered_santri_ids = set(santri_ids)
        elif status_daftar_filter == 'belum_pulang':
            santri_ids_sudah = [s_id for s_id, in db.session.query(Pendaftaran.santri_id).filter(Pendaftaran.edisi_id == active_edisi.id, Pendaftaran.santri_id.in_(santri_in_wilayah_ids), Pendaftaran.status_pulang != 'Tidak Ikut', rombongan_filter_pulang).distinct()]
            filtered_santri_ids = filtered_santri_ids - set(santri_ids_sudah)
        elif status_daftar_filter == 'sudah_kembali':
            santri_ids = [s_id for s_id, in db.session.query(Pendaftaran.santri_id).filter(Pendaftaran.edisi_id == active_edisi.id, Pendaftaran.santri_id.in_(santri_in_wilayah_ids), Pendaftaran.status_kembali != 'Tidak Ikut', rombongan_filter_kembali).distinct()]
            filtered_santri_ids = set(santri_ids)
        elif status_daftar_filter == 'belum_kembali':
            santri_ids_sudah = [s_id for s_id, in db.session.query(Pendaftaran.santri_id).filter(Pendaftaran.edisi_id == active_edisi.id, Pendaftaran.santri_id.in_(santri_in_wilayah_ids), Pendaftaran.status_kembali != 'Tidak Ikut', rombongan_filter_kembali).distinct()]
            filtered_santri_ids = filtered_santri_ids - set(santri_ids_sudah)
    
    # MODIFIED: For export, show all santri but apply additional rombongan filter if needed
    export_santri_ids = filtered_santri_ids.copy()
    
    # ADDITIONAL FILTER: Only include santri registered in managed rombongan for specific export type
    # BUT keep unregistered santri as well
    if current_user.role.name in ['Korwil', 'Korda'] and managed_rombongan_ids:
        if export_type == 'pulang':
            # Get santri registered in managed rombongan for pulang
            santri_ids_in_managed_rombongan = [
                s_id for s_id, in db.session.query(Pendaftaran.santri_id).filter(
                    Pendaftaran.edisi_id == active_edisi.id,
                    Pendaftaran.santri_id.in_(santri_in_wilayah_ids),
                    Pendaftaran.rombongan_pulang_id.in_(managed_rombongan_ids)
                ).distinct()
            ]
        else:  # kembali
            # Get santri registered in managed rombongan for kembali
            santri_ids_in_managed_rombongan = [
                s_id for s_id, in db.session.query(Pendaftaran.santri_id).filter(
                    Pendaftaran.edisi_id == active_edisi.id,
                    Pendaftaran.santri_id.in_(santri_in_wilayah_ids),
                    or_(
                        Pendaftaran.rombongan_kembali_id.in_(managed_rombongan_ids),
                        and_(
                            Pendaftaran.rombongan_kembali_id.is_(None), 
                            Pendaftaran.rombongan_pulang_id.in_(managed_rombongan_ids)
                        )
                    )
                ).distinct()
            ]
        
        # Get unregistered santri in the filtered list
        all_registered_santri = [
            s_id for s_id, in db.session.query(Pendaftaran.santri_id).filter(
                Pendaftaran.edisi_id == active_edisi.id,
                Pendaftaran.santri_id.in_(list(filtered_santri_ids))
            ).distinct()
        ]
        unregistered_santri_ids = filtered_santri_ids - set(all_registered_santri)
        
        # Combine registered santri in managed rombongan + all unregistered santri
        export_santri_ids = set(santri_ids_in_managed_rombongan) | unregistered_santri_ids
    
    # Get final santri list for export
    santri_list = [s for s in all_santri if s.id in export_santri_ids]
    
    # Get pendaftaran data for all santri with tarif data
    santri_ids = [s.id for s in santri_list]
    pendaftaran_map = {}
    tarif_map = {}
    
    if santri_ids:
        pendaftarans = Pendaftaran.query.options(
            selectinload(Pendaftaran.bus_pulang),
            selectinload(Pendaftaran.bus_kembali),
            selectinload(Pendaftaran.rombongan_pulang),
            selectinload(Pendaftaran.rombongan_kembali)
        ).filter(
            Pendaftaran.edisi_id == active_edisi.id,
            Pendaftaran.santri_id.in_(santri_ids)
        ).all()
        pendaftaran_map = {p.santri_id: p for p in pendaftarans}
        
        # Get tarif data untuk kalkulasi biaya
        for p in pendaftarans:
            if p.rombongan_pulang_id and p.titik_turun:
                tarif = Tarif.query.filter_by(
                    rombongan_id=p.rombongan_pulang_id, 
                    titik_turun=p.titik_turun
                ).first()
                if tarif:
                    tarif_map[f"{p.rombongan_pulang_id}_{p.titik_turun}"] = tarif

    # Create workbook
    wb = Workbook()
    ws = wb.active
    
    # Set worksheet title based on export type
    if export_type == 'pulang':
        ws.title = "Data Perjalanan Pulang"
    else:
        ws.title = "Data Perjalanan Kembali"

    # Define styles
    header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Style for unpaid santri (red background)
    unpaid_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
    unpaid_font = Font(name='Arial', size=10, bold=True, color='CC0000')
    
    # ADDED: Style for unregistered santri (yellow background)
    unregistered_fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
    unregistered_font = Font(name='Arial', size=10, bold=False, color='CC6600')

    # Set headers based on export type
    if export_type == 'pulang':
        headers = [
            'No', 'Nama Santri', 'Jenis Kelamin', 'Kabupaten', 'Status Santri', 
            'Jabatan', 'Status Pulang', 'Metode Pembayaran', 'Rombongan', 
            'Nama Bus', 'Biaya Pulang'
        ]
    else:  # kembali
        headers = [
            'No', 'Nama Santri', 'Jenis Kelamin', 'Kabupaten', 'Status Santri',
            'Jabatan', 'Status Kembali', 'Metode Pembayaran', 'Rombongan',
            'Nama Bus', 'Biaya Kembali'
        ]

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    # Initialize counters for summary
    total_transfer = 0
    total_cash = 0
    total_biaya_pulang = 0
    total_biaya_kembali = 0
    count_transfer = 0
    count_cash = 0
    count_belum_lunas = 0
    count_belum_terdaftar = 0  # ADDED: Counter for unregistered santri

    # Write data
    for idx, santri in enumerate(santri_list, 2):
        pendaftaran = pendaftaran_map.get(santri.id)
        
        # Base data
        row_data = [
            idx - 1,  # No
            santri.nama,
            santri.jenis_kelamin,
            santri.kabupaten,
            santri.status_santri,
            santri.nama_jabatan or 'Santri'
        ]
        
        # Determine santri status
        is_unpaid = False
        is_unregistered = False
        
        if not pendaftaran:
            is_unregistered = True
            count_belum_terdaftar += 1
        
        # Add specific data based on export type
        if export_type == 'pulang':
            if pendaftaran:
                # Calculate biaya pulang - sesuai logika di pendaftaran_rombongan
                biaya_pulang = 0
                if pendaftaran.status_pulang and pendaftaran.status_pulang != 'Tidak Ikut':
                    if pendaftaran.rombongan_pulang_id and pendaftaran.titik_turun:
                        tarif_key = f"{pendaftaran.rombongan_pulang_id}_{pendaftaran.titik_turun}"
                        tarif = tarif_map.get(tarif_key)
                        if not tarif:
                            # Fetch tarif if not in map
                            tarif = Tarif.query.filter_by(
                                rombongan_id=pendaftaran.rombongan_pulang_id, 
                                titik_turun=pendaftaran.titik_turun
                            ).first()
                        if tarif:
                            biaya_pulang = tarif.harga_bus + tarif.fee_korda + 10000
                            total_biaya_pulang += biaya_pulang
                
                # Determine if unpaid based on status_pulang = 'Belum Bayar'
                if pendaftaran.status_pulang == 'Belum Bayar':
                    is_unpaid = True
                    count_belum_lunas += 1
                elif pendaftaran.status_pulang and pendaftaran.status_pulang not in ['Tidak Ikut', 'Belum Bayar']:
                    # Count payment methods for paid santri
                    if pendaftaran.metode_pembayaran_pulang:
                        if 'transfer' in pendaftaran.metode_pembayaran_pulang.lower():
                            count_transfer += 1
                            total_transfer += biaya_pulang
                        elif 'cash' in pendaftaran.metode_pembayaran_pulang.lower():
                            count_cash += 1
                            total_cash += biaya_pulang
                
                row_data.extend([
                    pendaftaran.status_pulang,
                    pendaftaran.metode_pembayaran_pulang or '-',
                    pendaftaran.rombongan_pulang.nama_rombongan if pendaftaran.rombongan_pulang else '-',
                    f"{pendaftaran.bus_pulang.nama_armada} - {pendaftaran.bus_pulang.nomor_lambung}" if pendaftaran.bus_pulang else '-',
                    f"Rp {biaya_pulang:,.0f}".replace(',', '.') if biaya_pulang > 0 else 'Rp 0',
                ])
            else:
                row_data.extend(['Belum Terdaftar', '-', '-', '-', 'Rp 0'])
                
        else:  # kembali
            if pendaftaran:
                # Calculate biaya kembali - sesuai logika di pendaftaran_rombongan
                biaya_kembali = 0
                if pendaftaran.status_kembali and pendaftaran.status_kembali != 'Tidak Ikut':
                    # Tentukan rombongan dan titik untuk kembali
                    rombongan_kembali_id = pendaftaran.rombongan_kembali_id or pendaftaran.rombongan_pulang_id
                    titik_jemput_kembali = pendaftaran.titik_jemput_kembali or pendaftaran.titik_turun
                    
                    if rombongan_kembali_id and titik_jemput_kembali:
                        tarif_key = f"{rombongan_kembali_id}_{titik_jemput_kembali}"
                        tarif = tarif_map.get(tarif_key)
                        if not tarif:
                            # Fetch tarif if not in map
                            tarif = Tarif.query.filter_by(
                                rombongan_id=rombongan_kembali_id, 
                                titik_turun=titik_jemput_kembali
                            ).first()
                            if tarif:
                                tarif_map[tarif_key] = tarif
                        if tarif:
                            biaya_kembali = tarif.harga_bus + tarif.fee_korda + 10000
                            total_biaya_kembali += biaya_kembali
                
                # Determine payment status for kembali
                if pendaftaran.status_kembali and pendaftaran.status_kembali != 'Tidak Ikut':
                    if not pendaftaran.metode_pembayaran_kembali:
                        is_unpaid = True
                        count_belum_lunas += 1
                    else:
                        # Count payment methods for kembali
                        if 'transfer' in pendaftaran.metode_pembayaran_kembali.lower():
                            count_transfer += 1
                            total_transfer += biaya_kembali
                        elif 'cash' in pendaftaran.metode_pembayaran_kembali.lower():
                            count_cash += 1
                            total_cash += biaya_kembali
                
                row_data.extend([
                    pendaftaran.status_kembali,
                    pendaftaran.metode_pembayaran_kembali or '-',
                    pendaftaran.rombongan_kembali.nama_rombongan if pendaftaran.rombongan_kembali else 
                    (pendaftaran.rombongan_pulang.nama_rombongan if pendaftaran.rombongan_pulang else '-'),
                    f"{pendaftaran.bus_kembali.nama_armada} - {pendaftaran.bus_kembali.nomor_lambung}" if pendaftaran.bus_kembali else '-',
                    f"Rp {biaya_kembali:,.0f}".replace(',', '.') if biaya_kembali > 0 else 'Rp 0',
                ])
            else:
                row_data.extend(['Belum Terdaftar', '-', '-', '-', 'Rp 0'])
        
        # Write row data with conditional styling
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=idx, column=col, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical='center')
            
            # MODIFIED: Apply styling based on status
            if is_unregistered:
                cell.fill = unregistered_fill
                cell.font = unregistered_font
            elif is_unpaid:
                cell.fill = unpaid_fill
                cell.font = unpaid_font

    # Add summary section at the bottom
    summary_row_start = len(santri_list) + 4  # Start 3 rows after the last data
    
    # Summary styles
    summary_header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    summary_header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    summary_value_font = Font(name='Arial', size=11, bold=True, color='000000')
    summary_label_font = Font(name='Arial', size=11, bold=False, color='000000')
    
    # MODIFIED: Enhanced summary data with unregistered count
    summary_data = [
        ('RINGKASAN PEMBAYARAN', '', True),
        ('Total Pembayaran Transfer:', f"{count_transfer} orang - Rp {total_transfer:,.0f}".replace(',', '.'), False),
        ('Total Pembayaran Cash:', f"{count_cash} orang - Rp {total_cash:,.0f}".replace(',', '.'), False),
        ('Total Belum Lunas:', f"{count_belum_lunas} orang", False),
        ('Total Belum Terdaftar:', f"{count_belum_terdaftar} orang", False),  # ADDED
        ('', '', False),
        ('RINGKASAN BIAYA', '', True),
    ]
    
    if export_type == 'pulang':
        summary_data.append(('Total Biaya Pulang:', f"Rp {total_biaya_pulang:,.0f}".replace(',', '.'), False))
    else:
        summary_data.append(('Total Biaya Kembali:', f"Rp {total_biaya_kembali:,.0f}".replace(',', '.'), False))
    
    # ADDED: Legend section
    summary_data.extend([
        ('', '', False),
        ('KETERANGAN WARNA', '', True),
        ('Kuning:', 'Santri Belum Terdaftar', False),
        ('Merah:', 'Santri Belum Lunas Pembayaran', False),
        ('Putih:', 'Santri Sudah Lunas', False),
    ])
    
    # Write summary
    for i, (label, value, is_header) in enumerate(summary_data):
        row_idx = summary_row_start + i
        
        # Write label
        if label:  # Only write if label is not empty
            label_cell = ws.cell(row=row_idx, column=1, value=label)
            if is_header:
                label_cell.font = summary_header_font
                label_cell.fill = summary_header_fill
                # Merge cells for headers
                ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=4)
            else:
                label_cell.font = summary_label_font
        
        # Write value
        if value and not is_header:
            value_cell = ws.cell(row=row_idx, column=2, value=value)
            value_cell.font = summary_value_font

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Create response
    response = make_response(output.read())
    
    # Generate filename
    filename = f"Santri_Wilayah_Lengkap_{export_type.title()}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    
    return response

@admin_bp.route('/rombongan/salin-dari-sebelumnya', methods=['POST'])
@login_required
@role_required('Korpus')
def salin_rombongan():
    active_edisi = get_active_edisi()
    if not active_edisi or Rombongan.query.filter_by(edisi_id=active_edisi.id).first():
        flash("Aksi tidak diizinkan.", "danger")
        return redirect(url_for('admin.manajemen_rombongan'))

    sumber_edisi = Edisi.query.filter_by(is_active=False).order_by(Edisi.tahun.desc()).first()
    if not sumber_edisi:
        flash("Tidak ditemukan edisi sebelumnya untuk menyalin data.", "warning")
        return redirect(url_for('admin.manajemen_rombongan'))

    # Kamus untuk memetakan ID rombongan lama ke objek rombongan baru
    old_to_new_rombongan_map = {}

    for rombongan_lama in sumber_edisi.rombongans:
        # Salin data utama rombongan
        rombongan_baru = Rombongan(
            edisi_id=active_edisi.id,
            nama_rombongan=rombongan_lama.nama_rombongan,
            penanggung_jawab_putra=rombongan_lama.penanggung_jawab_putra,
            kontak_person_putra=rombongan_lama.kontak_person_putra,
            penanggung_jawab_putri=rombongan_lama.penanggung_jawab_putri,
            kontak_person_putri=rombongan_lama.kontak_person_putri,
            nomor_rekening=rombongan_lama.nomor_rekening,
            cakupan_wilayah=rombongan_lama.cakupan_wilayah
            # Jadwal dan batas bayar sengaja dikosongkan
        )
        for tarif_lama in rombongan_lama.tarifs:
            rombongan_baru.tarifs.append(Tarif(
                titik_turun=tarif_lama.titik_turun,
                harga_bus=tarif_lama.harga_bus,
                fee_korda=tarif_lama.fee_korda
            ))
        
        db.session.add(rombongan_baru)
        db.session.flush() # Penting: untuk mendapatkan ID rombongan_baru
        old_to_new_rombongan_map[rombongan_lama.id] = rombongan_baru

    # --- LOGIKA BARU: PERBARUI RELASI KORDA/KORWIL ---
    # Cari semua user yang merupakan Korda atau Korwil
    coordinators = User.query.join(Role).filter(Role.name.in_(['Korda', 'Korwil'])).all()
    for user in coordinators:
        # Cek rombongan lama mana saja yang dikelola oleh user ini
        old_managed_ids = {r.id for r in user.managed_rombongan}
        
        # Cari padanan rombongan baru dari peta yang sudah kita buat
        new_rombongan_to_assign = [
            old_to_new_rombongan_map[old_id] 
            for old_id in old_managed_ids 
            if old_id in old_to_new_rombongan_map
        ]
        
        # Tambahkan relasi baru ke user
        if new_rombongan_to_assign:
            user.managed_rombongan.extend(new_rombongan_to_assign)
    # --- AKHIR LOGIKA BARU ---

    log_activity('Salin Data', 'Rombongan', f"Menyalin {len(old_to_new_rombongan_map)} rombongan dari edisi '{sumber_edisi.nama}'")
    db.session.commit()
    
    flash(f"Berhasil menyalin {len(old_to_new_rombongan_map)} rombongan dari edisi sebelumnya.", "success")
    return redirect(url_for('admin.manajemen_rombongan'))

# Di dalam file app/admin/routes.py

@admin_bp.route('/cetak-kartu')
@login_required
@role_required('Korpus', 'Korda', 'Korpuspi', 'Sekretaris')
def cetak_kartu():
    active_edisi = get_active_edisi()
    if not active_edisi:
        flash("Tidak ada edisi aktif.", "warning")
        return render_template('cetak_kartu.html', semua_pendaftar=[], rombongan_for_filter=[])

    # Ambil semua parameter filter dari URL
    nama_filter = request.args.get('nama', '')
    rombongan_id_filter = request.args.get('rombongan_id', type=int)
    perjalanan_filter = request.args.get('perjalanan', '')
    jenis_kelamin_filter = request.args.get('jenis_kelamin', '') # <-- BARIS BARU

    # Siapkan query dasar
    query = Pendaftaran.query.join(Santri).filter(Pendaftaran.edisi_id == active_edisi.id)

    # Filter berdasarkan hak akses
    managed_rombongan_ids = []
    rombongan_for_filter = Rombongan.query.filter_by(edisi_id=active_edisi.id).order_by(Rombongan.nama_rombongan).all()

    if current_user.role.name in ['Korwil', 'Korda']:
        managed_rombongan_ids = [r.id for r in current_user.active_managed_rombongan]
        if not managed_rombongan_ids:
            query = query.filter(db.false())
        else:
            query = query.filter(
                or_(
                    Pendaftaran.rombongan_pulang_id.in_(managed_rombongan_ids),
                    Pendaftaran.rombongan_kembali_id.in_(managed_rombongan_ids)
                )
            )
        # Sesuaikan daftar rombongan untuk dropdown filter
        rombongan_for_filter = [r for r in rombongan_for_filter if r.id in managed_rombongan_ids]
    
    # Terapkan filter ke query
    if nama_filter:
        query = query.filter(Santri.nama.ilike(f'%{nama_filter}%'))
    
    if rombongan_id_filter:
        query = query.filter(
            or_(
                Pendaftaran.rombongan_pulang_id == rombongan_id_filter,
                Pendaftaran.rombongan_kembali_id == rombongan_id_filter
            )
        )
    
    if perjalanan_filter:
        if perjalanan_filter == 'pulang':
            query = query.filter(Pendaftaran.status_pulang != 'Tidak Ikut')
        elif perjalanan_filter == 'kembali':
            query = query.filter(Pendaftaran.status_kembali != 'Tidak Ikut')
            
    if jenis_kelamin_filter: # <-- BLOK LOGIKA BARU
        query = query.filter(Santri.jenis_kelamin == jenis_kelamin_filter)

    # Ambil semua pendaftar yang relevan dan urutkan
    semua_pendaftar = query.options(
        joinedload(Pendaftaran.santri),
        joinedload(Pendaftaran.rombongan_pulang),
        joinedload(Pendaftaran.bus_pulang)
    ).order_by(Santri.nama).all()

    # Hilangkan duplikat santri jika perlu
    unique_pendaftar = list({p.santri_id: p for p in semua_pendaftar}.values())

    return render_template('cetak_kartu.html', 
                           semua_pendaftar=unique_pendaftar,
                           rombongan_for_filter=rombongan_for_filter)

@admin_bp.route('/cetak-tiket')
@login_required
@role_required('Korpus', 'Korda', 'Korpuspi') # Pastikan role sesuai
def cetak_tiket():
    active_edisi = get_active_edisi()
    if not active_edisi:
        flash("Tidak ada edisi aktif.", "warning")
        return render_template('cetak_tiket.html', semua_pendaftar=[], rombongan_for_filter=[])

    # BARU: Ambil parameter filter dari URL
    nama_filter = request.args.get('nama', '')
    rombongan_id_filter = request.args.get('rombongan_id', type=int)
    perjalanan_filter = request.args.get('perjalanan', '')

    # Siapkan query dasar
    query = Pendaftaran.query.join(Santri).filter(Pendaftaran.edisi_id == active_edisi.id)

    # Filter berdasarkan hak akses dan siapkan daftar rombongan untuk dropdown
    managed_rombongan_ids = []
    rombongan_for_filter = Rombongan.query.filter_by(edisi_id=active_edisi.id).order_by(Rombongan.nama_rombongan).all()

    if current_user.role.name in ['Korwil', 'Korda']:
        managed_rombongan_ids = [r.id for r in current_user.active_managed_rombongan]
        if not managed_rombongan_ids:
            query = query.filter(db.false())
        else:
            query = query.filter(
                or_(
                    Pendaftaran.rombongan_pulang_id.in_(managed_rombongan_ids),
                    Pendaftaran.rombongan_kembali_id.in_(managed_rombongan_ids)
                )
            )
        rombongan_for_filter = [r for r in rombongan_for_filter if r.id in managed_rombongan_ids]
    
    # BARU: Terapkan filter dari form ke query
    if nama_filter:
        query = query.filter(Santri.nama.ilike(f'%{nama_filter}%'))
    
    if rombongan_id_filter:
        query = query.filter(
            or_(
                Pendaftaran.rombongan_pulang_id == rombongan_id_filter,
                Pendaftaran.rombongan_kembali_id == rombongan_id_filter
            )
        )
    
    if perjalanan_filter:
        if perjalanan_filter == 'pulang':
            query = query.filter(Pendaftaran.status_pulang != 'Tidak Ikut')
        elif perjalanan_filter == 'kembali':
            query = query.filter(Pendaftaran.status_kembali != 'Tidak Ikut')

    # Ambil semua pendaftar yang relevan
    semua_pendaftar = query.options(
        joinedload(Pendaftaran.santri),
        joinedload(Pendaftaran.rombongan_pulang),
        joinedload(Pendaftaran.rombongan_kembali),
        joinedload(Pendaftaran.bus_pulang),
        joinedload(Pendaftaran.bus_kembali)
    ).order_by(Santri.nama).all()

    unique_pendaftar = list({p.santri_id: p for p in semua_pendaftar}.values())

    return render_template('cetak_tiket.html', 
                           semua_pendaftar=unique_pendaftar,
                           rombongan_for_filter=rombongan_for_filter)

# 1. Halaman Utama Manajemen Wisuda (termasuk logika impor)
# Di dalam file app/admin/routes.py

@admin_bp.route('/manajemen-wisuda', methods=['GET', 'POST'])
@login_required
@role_required('Korpus', 'PJ Acara', 'Korwil', 'Korda', 'Bendahara Pusat', 'Sekretaris', 'Korpuspi')
def manajemen_wisuda():
    active_edisi = get_active_edisi()
    import_form = ImportWisudaForm()

    if import_form.validate_on_submit():
        file = import_form.file.data
        kategori = import_form.kategori_wisuda.data

        try:
            df = pd.read_excel(file)
            if 'NIS' not in df.columns:
                flash('File Excel harus memiliki kolom bernama "NIS".', 'danger')
                return redirect(url_for('admin.manajemen_wisuda'))

            all_nis = df['NIS'].dropna().astype(str).tolist()

            # Cari semua santri yang cocok dalam satu query
            santri_map = {s.nis: s for s in Santri.query.filter(Santri.nis.in_(all_nis)).all()}

            newly_graduated = []
            for nis in all_nis:
                if nis in santri_map:
                    santri = santri_map[nis]
                    # Cek apakah sudah ada data wisuda
                    if not santri.wisuda_info:
                        santri.status_santri = 'Wisuda'
                        new_wisuda = Wisuda(
                            santri_nis=santri.nis,
                            edisi_id=active_edisi.id,
                            kategori_wisuda=kategori
                        )
                        newly_graduated.append(new_wisuda)

            if newly_graduated:
                db.session.add_all(newly_graduated)
                db.session.commit()
                flash(f'{len(newly_graduated)} santri berhasil diimpor dan statusnya diubah menjadi Wisuda.', 'success')
            else:
                flash('Tidak ada santri baru yang diimpor. Mungkin semua sudah terdata.', 'info')

        except Exception as e:
            flash(f'Terjadi error saat memproses file: {e}', 'danger')

        return redirect(url_for('admin.manajemen_wisuda'))

    # --- AWAL BLOK LOGIKA BARU ---
    wisudawan_list = []
    if active_edisi:
        # Mulai query dasar untuk wisudawan di edisi aktif
        query = Wisuda.query.filter_by(edisi_id=active_edisi.id).options(joinedload(Wisuda.santri))

        # Terapkan filter jika rolenya Korda atau Korwil
        if current_user.role.name in ['Korda', 'Korwil']:
            # Kumpulkan cakupan wilayah yang dikelola
            managed_regions = {
                wilayah.get('label')
                for rombongan in current_user.active_managed_rombongan
                if rombongan.cakupan_wilayah
                for wilayah in rombongan.cakupan_wilayah
            }

            if managed_regions:
                # Lakukan join dengan tabel Santri dan filter berdasarkan 'kabupaten'
                # Note: Relasi Wisuda -> Santri via 'santri_nis' ke 'nis'
                query = query.join(Santri, Wisuda.santri_nis == Santri.nis).filter(Santri.kabupaten.in_(list(managed_regions)))
            else:
                # Jika tidak punya wilayah, jangan tampilkan data
                query = query.filter(db.false())
        
        wisudawan_list = query.all()
    # --- AKHIR BLOK LOGIKA BARU ---

    # Kelompokkan data berdasarkan kategori (logika ini tidak berubah)
    grouped_wisudawan = defaultdict(list)
    for w in wisudawan_list:
        kategori = w.kategori_wisuda
        if kategori not in grouped_wisudawan:
            grouped_wisudawan[kategori] = []
        grouped_wisudawan[kategori].append(w)

    return render_template('manajemen_wisuda.html', 
                           import_form=import_form, 
                           grouped_wisudawan=dict(grouped_wisudawan), # Konversi kembali ke dict
                           active_edisi=active_edisi)

# 2. Halaman untuk menambah wisudawan manual
# Di dalam file app/admin/routes.py

@admin_bp.route('/wisuda/tambah', methods=['GET', 'POST'])
@login_required
@role_required('Korpus', 'PJ Acara', 'Sekretaris', 'Korpuspi')
def tambah_wisudawan():
    form = WisudaForm()
    active_edisi = get_active_edisi()

    if not active_edisi:
        flash("Tidak ada edisi aktif untuk menambahkan wisudawan.", "warning")
        return redirect(url_for('admin.manajemen_wisuda'))

    if form.validate_on_submit():
        # 1. Ambil string NIS dan ubah menjadi list
        nis_list_str = form.santri.data
        all_nis = [nis.strip() for nis in nis_list_str.split(',') if nis.strip()]
        
        if not all_nis:
            flash("Anda belum memilih santri.", "danger")
            return redirect(url_for('admin.tambah_wisudawan'))

        kategori = form.kategori_wisuda.data
        
        # 2. Cari semua santri yang relevan dalam satu query
        santri_to_process = Santri.query.filter(Santri.nis.in_(all_nis)).all()
        
        newly_graduated_count = 0
        for santri in santri_to_process:
            # 3. Cek duplikat dan proses setiap santri
            if not santri.wisuda_info and santri.status_santri == 'Aktif':
                santri.status_santri = 'Wisuda' # Update status santri
                
                new_wisuda = Wisuda(
                    santri_nis=santri.nis,
                    edisi_id=active_edisi.id,
                    kategori_wisuda=kategori
                )
                db.session.add(new_wisuda)
                newly_graduated_count += 1
        
        # 4. Simpan semua perubahan ke database sekaligus
        if newly_graduated_count > 0:
            db.session.commit()
            flash(f'{newly_graduated_count} santri berhasil ditetapkan sebagai wisudawan.', 'success')
        else:
            flash('Tidak ada santri baru yang ditambahkan. Mungkin semua sudah terdata sebagai wisudawan atau statusnya tidak aktif.', 'info')

        return redirect(url_for('admin.manajemen_wisuda'))
        
    return render_template('tambah_wisudawan.html', form=form, title="Tambah Wisudawan")

# 3. Route untuk menghapus status wisuda
@admin_bp.route('/hapus-wisuda/<int:wisuda_id>', methods=['POST'])
@login_required
@role_required('Korpus', 'Sekretaris', 'Korpuspi')
def hapus_wisuda(wisuda_id):
    wisuda = Wisuda.query.get_or_404(wisuda_id)
    santri = wisuda.santri
    santri.status_santri = 'Aktif' # Kembalikan statusnya
    db.session.delete(wisuda)
    db.session.commit()
    flash(f'Status wisuda untuk {santri.nama} berhasil dihapus.', 'success')
    return redirect(url_for('admin.manajemen_wisuda'))

# 4. API endpoint baru untuk pencarian santri di form tambah
@admin_bp.route('/api/search-santri-for-wisuda')
@login_required
def api_search_santri_for_wisuda():
    q = request.args.get('q', '')
    query = Santri.query.filter(
        Santri.status_santri == 'Aktif',
        Santri.wisuda_info == None,
        Santri.nama.ilike(f'%{q}%')
    ).limit(20)

    results = [{
        'id': s.id,
        'nis': s.nis,
        'nama': s.nama,
        'asrama': s.asrama, 
        'kabupaten': s.kabupaten
    } for s in query.all()]
    return jsonify({'results': results})

@admin_bp.route('/santri/update-wali/<int:santri_id>', methods=['POST'])
@login_required
@role_required('Korpus', 'Korda', 'Korwil', 'Sekretaris', 'Korpuspi')
def update_nomor_wali(santri_id):
    santri = Santri.query.get_or_404(santri_id)
    new_phone_number = request.json.get('no_hp_wali')

    if not new_phone_number:
        return jsonify({'success': False, 'message': 'Nomor HP tidak boleh kosong.'}), 400

    # --- Bagian Sinkronisasi ke API Induk (Sudah Disederhanakan) ---
    try:
        if not santri.nis:
            return jsonify({'success': False, 'message': f'Santri {santri.nama} tidak memiliki NIS untuk sinkronisasi.'}), 400

        api_url = f"{current_app.config['API_INDUK_URL']}/{santri.nis}"
        payload = json.dumps({"parrentPhone": new_phone_number})
        headers = {
            'Content-Type': 'application/json' # Header yang lebih sederhana
        }

        print(api_url, payload)  # Debugging output
        
        response = requests.request("PATCH", api_url, headers=headers, data=payload, timeout=15)
        response.raise_for_status()

        print(response.text)

    except requests.exceptions.RequestException as e:
        error_message = f"Gagal sinkronisasi ke server induk. Respon: {e.response.text if e.response else 'Tidak ada respon'}"
        return jsonify({'success': False, 'message': error_message}), 500
    # --- Akhir Bagian Sinkronisasi ---

    # Jika sinkronisasi berhasil, update database lokal kita
    santri.no_hp_wali = new_phone_number
    log_activity('Edit', 'Santri', f"Mengubah nomor HP wali untuk santri: '{santri.nama}'")
    db.session.commit()

    return jsonify({'success': True, 'message': 'Nomor HP wali berhasil diperbarui dan disinkronkan.'})

# Ganti fungsi bendahara_dashboard dan konfirmasi_setor dengan ini

@admin_bp.route('/bendahara')
@login_required
@role_required('Bendahara Pusat', 'Korpus')
def bendahara_dashboard():
    active_edisi = get_active_edisi()
    if not active_edisi:
        return render_template('bendahara_dashboard.html', active_edisi=None)

    # Hitung saldo rekening virtual dari tabel Transaksi
    def get_saldo(rekening):
        pemasukan = db.session.query(func.sum(Transaksi.jumlah)).filter_by(edisi_id=active_edisi.id, tipe='PEMASUKAN', rekening=rekening).scalar() or 0
        pengeluaran = db.session.query(func.sum(Transaksi.jumlah)).filter_by(edisi_id=active_edisi.id, tipe='PENGELUARAN', rekening=rekening).scalar() or 0
        return pemasukan - pengeluaran

    saldo_rekening_saya = get_saldo('REKENING_SAYA')
    saldo_bus_pulang = get_saldo('REKENING_BUS_PULANG')
    saldo_bus_kembali = get_saldo('REKENING_BUS_KEMBALI')

    # Siapkan data untuk setiap rombongan
    all_rombongan = Rombongan.query.filter_by(edisi_id=active_edisi.id).all()
    rombongan_data_list = []
    for rombongan in all_rombongan:
        # --- BLOK PERHITUNGAN BARU: Menghitung dari SEMUA PENDAFTAR ---

        # 1. Hitung Total Biaya Bus Seharusnya (dari semua yang statusnya bukan 'Tidak Ikut')
        total_bus_pulang_seharusnya = db.session.query(func.sum(Tarif.harga_bus)).select_from(Pendaftaran).join(
            Tarif, and_(
                Pendaftaran.rombongan_pulang_id == Tarif.rombongan_id,
                Pendaftaran.titik_turun == Tarif.titik_turun
            )
        ).filter(Pendaftaran.rombongan_pulang_id == rombongan.id, Pendaftaran.status_pulang != 'Tidak Ikut').scalar() or 0
        
        total_bus_kembali_seharusnya = db.session.query(func.sum(Tarif.harga_bus)).select_from(Pendaftaran).join(
            Tarif, and_(
                (Pendaftaran.rombongan_kembali_id == Tarif.rombongan_id),
                (Pendaftaran.titik_jemput_kembali == Tarif.titik_turun)
            )
        ).filter(Pendaftaran.rombongan_kembali_id == rombongan.id, Pendaftaran.status_kembali != 'Tidak Ikut').scalar() or 0
        
        # 2. Hitung Total Fee Pondok Seharusnya (dari semua yang statusnya bukan 'Tidak Ikut')
        total_pondok_pulang_seharusnya = Pendaftaran.query.filter_by(rombongan_pulang_id=rombongan.id).filter(Pendaftaran.status_pulang != 'Tidak Ikut').count() * 10000
        total_pondok_kembali_seharusnya = Pendaftaran.query.filter_by(rombongan_kembali_id=rombongan.id).filter(Pendaftaran.status_kembali != 'Tidak Ikut').count() * 10000
        
        # 3. Hitung Sisa Setoran berdasarkan total yang seharusnya
        sisa_setoran_bus_pulang = total_bus_pulang_seharusnya - (rombongan.total_setoran_bus_pulang or 0)
        sisa_setoran_bus_kembali = total_bus_kembali_seharusnya - (rombongan.total_setoran_bus_kembali or 0)
        sisa_setoran_pondok_pulang = total_pondok_pulang_seharusnya - (rombongan.total_setoran_pondok_pulang or 0)
        sisa_setoran_pondok_kembali = total_pondok_kembali_seharusnya - (rombongan.total_setoran_pondok_kembali or 0)
        
        # 4. Hitung peserta belum lunas (logika ini tetap sama)
        peserta_belum_lunas = Pendaftaran.query.filter(
            or_(Pendaftaran.rombongan_pulang_id == rombongan.id, Pendaftaran.rombongan_kembali_id == rombongan.id),
            or_(Pendaftaran.status_pulang == 'Belum Bayar', Pendaftaran.status_kembali == 'Belum Bayar')
        ).count()

        rombongan_data_list.append({
            'rombongan': rombongan,
            'total_bus_pulang_seharusnya': total_bus_pulang_seharusnya,
            'total_bus_kembali_seharusnya': total_bus_kembali_seharusnya,
            'total_pondok_pulang_seharusnya': total_pondok_pulang_seharusnya,
            'total_pondok_kembali_seharusnya': total_pondok_kembali_seharusnya,
            'sisa_setoran_bus_pulang': sisa_setoran_bus_pulang,
            'sisa_setoran_bus_kembali': sisa_setoran_bus_kembali,
            'sisa_setoran_pondok_pulang': sisa_setoran_pondok_pulang,
            'sisa_setoran_pondok_kembali': sisa_setoran_pondok_kembali,
            'peserta_belum_lunas': peserta_belum_lunas
        })

    form_setoran = KonfirmasiSetoranForm()
    return render_template('bendahara_dashboard.html', 
                           saldo_rekening_saya=saldo_rekening_saya,
                           saldo_bus_pulang=saldo_bus_pulang,
                           saldo_bus_kembali=saldo_bus_kembali,
                           rombongan_data_list=rombongan_data_list,
                           active_edisi=active_edisi,
                           form_setoran=form_setoran)

@admin_bp.route('/bendahara/konfirmasi-setor/<int:rombongan_id>', methods=['POST'])
@login_required
@role_required('Bendahara Pusat', 'Korpus')
def konfirmasi_setor(rombongan_id):
    tipe = request.args.get('tipe')
    rombongan = Rombongan.query.get_or_404(rombongan_id)
    form = KonfirmasiSetoranForm()

    if form.validate_on_submit():
        jumlah = form.jumlah_disetor.data
        
        rekening_map = {
            'bus_pulang': ('REKENING_BUS_PULANG', 'Biaya Bus Pulang'),
            'bus_kembali': ('REKENING_BUS_KEMBALI', 'Biaya Bus Kembali'),
            'pondok_pulang': ('REKENING_SAYA', 'Fee Pondok Pulang'),
            'pondok_kembali': ('REKENING_SAYA', 'Fee Pondok Kembali')
        }
        
        if tipe in rekening_map:
            rekening, deskripsi_prefix = rekening_map[tipe]
            
            # --- BLOK KODE BARU UNTUK MENCATAT PEMASUKAN ---
            # 1. Buat record transaksi baru
            transaksi = Transaksi(
                edisi_id=get_active_edisi().id,
                deskripsi=f"Setoran {deskripsi_prefix} dari {rombongan.nama_rombongan}",
                jumlah=jumlah,
                tipe='PEMASUKAN', # Menandakan ini adalah uang masuk
                rekening=rekening,
                user_id=current_user.id,
                rombongan_id=rombongan.id
            )
            db.session.add(transaksi)
            # --- AKHIR BLOK KODE BARU ---
            
            # 2. Update total setoran di rombongan (logika ini tetap sama)
            if tipe == 'bus_pulang':
                if rombongan.total_setoran_bus_pulang is None:
                    rombongan.total_setoran_bus_pulang = 0
                rombongan.total_setoran_bus_pulang += jumlah
    
            elif tipe == 'bus_kembali':
                if rombongan.total_setoran_bus_kembali is None:
                    rombongan.total_setoran_bus_kembali = 0
                rombongan.total_setoran_bus_kembali += jumlah
    
            elif tipe == 'pondok_pulang':
                if rombongan.total_setoran_pondok_pulang is None:
                    rombongan.total_setoran_pondok_pulang = 0
                rombongan.total_setoran_pondok_pulang += jumlah
    
            elif tipe == 'pondok_kembali':
                if rombongan.total_setoran_pondok_kembali is None:
                    rombongan.total_setoran_pondok_kembali = 0
                rombongan.total_setoran_pondok_kembali += jumlah
            
            db.session.commit()
            flash(f"Setoran sebesar Rp {jumlah:,} berhasil dikonfirmasi dan dicatat sebagai pemasukan.", 'success')
        else:
            flash("Tipe setoran tidak valid.", "danger")
    else:
        # Menampilkan error validasi form jika ada
        for field, errors in form.errors.items():
            for error in errors:
                flash(f"Error pada input '{getattr(form, field).label.text}': {error}", 'danger')

    return redirect(url_for('admin.bendahara_dashboard'))

# 3. Rute Halaman Pengeluaran
# Ganti fungsi bendahara_pengeluaran dengan ini
@admin_bp.route('/bendahara/buku-kas', methods=['GET', 'POST'])
@login_required
@role_required('Bendahara Pusat', 'Korpus')
def bendahara_buku_kas():
    active_edisi = get_active_edisi()
    form = TransaksiForm()
    
    if form.validate_on_submit():
        pengeluaran = Transaksi(
            edisi_id=active_edisi.id,
            deskripsi=form.deskripsi.data,
            jumlah=form.jumlah.data,
            tipe='PENGELUARAN',
            rekening='REKENING_SAYA',
            user_id=current_user.id
        )
        db.session.add(pengeluaran)
        db.session.commit()
        flash('Data pengeluaran berhasil dicatat.', 'success')
        return redirect(url_for('admin.bendahara_buku_kas'))

    # Hitung saldo untuk ditampilkan
    pemasukan_saya = db.session.query(func.sum(Transaksi.jumlah)).filter_by(edisi_id=active_edisi.id, tipe='PEMASUKAN', rekening='REKENING_SAYA').scalar() or 0
    pengeluaran_saya = db.session.query(func.sum(Transaksi.jumlah)).filter_by(edisi_id=active_edisi.id, tipe='PENGELUARAN', rekening='REKENING_SAYA').scalar() or 0
    saldo_rekening_saya = pemasukan_saya - pengeluaran_saya
    
    # --- PERUBAHAN DI SINI ---
    # Ambil riwayat SEMUA transaksi (Pemasukan dan Pengeluaran)
    riwayat_transaksi = Transaksi.query.filter(
        Transaksi.edisi_id == active_edisi.id,
        Transaksi.rekening == 'REKENING_SAYA' # Hanya dari Rekening Saya
    ).order_by(Transaksi.tanggal.desc()).all()

    return render_template('bendahara_buku_kas.html', 
                           form=form, 
                           saldo_rekening_saya=saldo_rekening_saya,
                           riwayat_transaksi=riwayat_transaksi)


# Ganti fungsi bendahara_buku_kas_bus dengan ini
@admin_bp.route('/bendahara/buku-kas-bus', methods=['GET', 'POST'])
@login_required
@role_required('Bendahara Pusat', 'Korpus')
def bendahara_buku_kas_bus():
    active_edisi = get_active_edisi()
    form = PengeluaranBusForm() # Gunakan form baru

    if form.validate_on_submit():
        pengeluaran = Transaksi(
            edisi_id=active_edisi.id,
            deskripsi=form.deskripsi.data,
            jumlah=form.jumlah.data,
            tipe='PENGELUARAN', # Tandai sebagai pengeluaran
            rekening=form.rekening.data, # Ambil rekening dari pilihan dropdown
            user_id=current_user.id
        )
        db.session.add(pengeluaran)
        db.session.commit()
        flash('Data pengeluaran bus berhasil dicatat.', 'success')
        return redirect(url_for('admin.bendahara_buku_kas_bus'))

    # Logika untuk menghitung saldo dan mengambil riwayat (tidak berubah)
    def get_saldo(rekening):
        pemasukan = db.session.query(func.sum(Transaksi.jumlah)).filter_by(edisi_id=active_edisi.id, tipe='PEMASUKAN', rekening=rekening).scalar() or 0
        pengeluaran = db.session.query(func.sum(Transaksi.jumlah)).filter_by(edisi_id=active_edisi.id, tipe='PENGELUARAN', rekening=rekening).scalar() or 0
        return pemasukan - pengeluaran

    saldo_bus_pulang = get_saldo('REKENING_BUS_PULANG')
    saldo_bus_kembali = get_saldo('REKENING_BUS_KEMBALI')

    riwayat_transaksi_bus = Transaksi.query.filter(
        Transaksi.edisi_id == active_edisi.id,
        Transaksi.rekening.in_(['REKENING_BUS_PULANG', 'REKENING_BUS_KEMBALI'])
    ).order_by(Transaksi.tanggal.desc()).all()
    total_saldo_bus = saldo_bus_pulang + saldo_bus_kembali

    return render_template('bendahara_buku_kas_bus.html', 
                           saldo_bus_pulang=saldo_bus_pulang,
                           saldo_bus_kembali=saldo_bus_kembali,
                           riwayat_transaksi_bus=riwayat_transaksi_bus,
                           form=form,
                           total_saldo_bus=total_saldo_bus) # Kirim form ke template


# Di dalam file app/admin/routes.py

@admin_bp.route('/ganti-password', methods=['GET', 'POST'])
@login_required # Hanya untuk user yang sudah login
def ganti_password():
    form = ChangePasswordForm()
    if form.validate_on_submit():
        user = current_user # Ambil data user yang sedang login

        # 1. Verifikasi password saat ini
        if not user.check_password(form.current_password.data):
            flash('Password saat ini yang Anda masukkan salah.', 'danger')
            return redirect(url_for('admin.ganti_password'))

        # 2. Pastikan password baru tidak sama dengan password lama
        if user.check_password(form.new_password.data):
            flash('Password baru tidak boleh sama dengan password lama.', 'warning')
            return redirect(url_for('admin.ganti_password'))

        # 3. Jika semua valid, ganti password
        user.set_password(form.new_password.data)
        db.session.commit()
        
        log_activity('Keamanan', 'Akun', f"User '{user.username}' berhasil mengganti passwordnya sendiri.")
        flash('Password Anda telah berhasil diperbarui!', 'success')
        return redirect(url_for('admin.dashboard'))

    return render_template('ganti_password.html', title="Ganti Password", form=form)

# Di dalam file app/admin/routes.py

@admin_bp.route('/profil-saya')
@login_required
def profil_saya():
    # Logika untuk panduan akan ditambahkan di sini nanti
    # Untuk sekarang, kita hanya menampilkan data user
    return render_template('profil_saya.html', title="Profil Saya")



# Ganti seluruh fungsi tambah_santri_manual dengan ini

# Ganti/tambahkan route ini di app/admin/routes.py

@admin_bp.route('/santri/tambah-manual', methods=['GET', 'POST'])
@login_required
@role_required('Korpus', 'Korpuspi')
def tambah_santri_manual():
    form = SantriManualForm()
    
    # Validasi tambahan untuk memastikan form berfungsi
    if request.method == 'POST':
        print(f"Form submitted. Valid: {form.validate()}")
        print(f"Form errors: {form.errors}")
        print(f"Form data: {form.data}")
    
    if form.validate_on_submit():
        try:
            # --- LOGIKA PEMBUATAN NIS OTOMATIS ---
            tahun_sekarang = str(datetime.now().year)
            # Cari santri terakhir yang dibuat manual di tahun ini
            santri_terakhir = Santri.query.filter(
                Santri.nis.like(f'M{tahun_sekarang}%')
            ).order_by(Santri.nis.desc()).first()
            
            if santri_terakhir:
                # Ambil 4 digit terakhir dan convert ke int
                try:
                    nomor_urut_terakhir = int(santri_terakhir.nis[-4:])
                    nomor_urut_baru = nomor_urut_terakhir + 1
                except ValueError:
                    # Jika gagal parsing, mulai dari 1
                    nomor_urut_baru = 1
            else:
                nomor_urut_baru = 1
            
            nis_baru = f"M{tahun_sekarang}{nomor_urut_baru:04d}"
            # ------------------------------------

            # Buat api_student_id unik sementara
            api_id_manual = f"manual_{nis_baru}_{int(datetime.utcnow().timestamp())}"

            # Validasi data sebelum disimpan
            if not form.nama.data or len(form.nama.data.strip()) < 2:
                flash('Nama harus diisi minimal 2 karakter.', 'error')
                return render_template('tambah_santri_manual.html', title="Tambah Santri Manual", form=form)
            
            if not form.provinsi.data or not form.kabupaten.data:
                flash('Provinsi dan Kabupaten harus dipilih.', 'error')
                return render_template('tambah_santri_manual.html', title="Tambah Santri Manual", form=form)

            new_santri = Santri(
                api_student_id=api_id_manual,
                nis=nis_baru,
                nama=form.nama.data.strip(),
                provinsi=form.provinsi.data,
                kabupaten=form.kabupaten.data,
                jenis_kelamin=form.jenis_kelamin.data,
                status_santri=form.status_santri.data,
                asrama=form.asrama.data.strip() if form.asrama.data else None,
                no_hp_wali=form.no_hp_wali.data.strip() if form.no_hp_wali.data else None,
                kelas_formal=form.kelas_formal.data.strip() if form.kelas_formal.data else None,
                kelas_ngaji=form.kelas_ngaji.data.strip() if form.kelas_ngaji.data else None
            )
            
            db.session.add(new_santri)
            db.session.commit()
            
            log_activity('Tambah', 'Santri', f"Menambah santri manual: '{new_santri.nama}' dengan NIS '{new_santri.nis}'")
            flash(f'Santri "{new_santri.nama}" berhasil ditambahkan dengan NIS: {nis_baru}.', 'success')
            return redirect(url_for('admin.manajemen_santri'))
            
        except Exception as e:
            db.session.rollback()
            print(f"Error saat menambah santri: {str(e)}")
            flash(f'Terjadi kesalahan: {str(e)}', 'error')
    
    return render_template('tambah_santri_manual.html', title="Tambah Santri Manual", form=form)

# Di dalam file app/admin/routes.py

@admin_bp.route('/santri/update-keterangan/<int:santri_id>', methods=['POST'])
@login_required
def update_keterangan_santri(santri_id):
    santri = Santri.query.get_or_404(santri_id)
    
    # Langsung ambil data dari request form yang dikirim oleh JavaScript
    keterangan = request.form.get('keterangan_manual')
    
    # Update data di database
    santri.keterangan_manual = keterangan
    db.session.commit()
    
    # Kirim respons sukses
    return jsonify({'success': True, 'message': 'Keterangan berhasil disimpan.'})

# API endpoints untuk provinsi dan kabupaten
@admin_bp.route('/api/get-provinsi')
@login_required
def api_get_provinsi():
    try:
        api_url = "https://backapp.amtsilatipusat.com/api/provinces"
        response = requests.get(api_url, timeout=10)
        response.raise_for_status()
        return jsonify(response.json())
    except requests.exceptions.RequestException as e:
        print(f"Error fetching provinsi: {str(e)}")
        return jsonify({'data': [], 'error': 'Gagal mengambil data provinsi'})


@admin_bp.route('/api/get-kabupaten/<int:provinsi_id>')
@login_required
def api_get_kabupaten(provinsi_id):
    try:
        api_url = f"https://backapp.amtsilatipusat.com/api/regencies/{provinsi_id}"
        response = requests.get(api_url, timeout=10)
        response.raise_for_status()
        return jsonify(response.json())
    except requests.exceptions.RequestException as e:
        print(f"Error fetching kabupaten: {str(e)}")
        return jsonify({'data': [], 'error': 'Gagal mengambil data kabupaten'})
    
@admin_bp.route('/rekap-absen')
@login_required
@role_required('Korpus', 'Korwil', 'Korda', 'Sekretaris', 'Korpuspi')
def rekap_absen():
    active_edisi = get_active_edisi()
    if not active_edisi:
        flash("Tidak ada edisi aktif.", "warning")
        return render_template('rekap_absen.html', active_edisi=None, rekap_data={})

    # 1. Ambil semua rombongan yang relevan berdasarkan peran
    base_query = Rombongan.query.filter_by(edisi_id=active_edisi.id)
    if current_user.role.name in ['Korwil', 'Korda']:
        managed_rombongan_ids = [r.id for r in current_user.active_managed_rombongan]
        if not managed_rombongan_ids:
            base_query = base_query.filter(db.false())
        else:
            base_query = base_query.filter(Rombongan.id.in_(managed_rombongan_ids))
    
    rombongan_list = base_query.order_by(Rombongan.nama_rombongan).all()

    # 2. Ambil semua data absensi yang relevan dalam satu query besar
    all_absensi = db.session.query(
        Absen, Pendaftaran, Santri, Bus, Rombongan
    ).join(
        Pendaftaran, Absen.pendaftaran_id == Pendaftaran.id
    ).join(
        Santri, Pendaftaran.santri_id == Santri.id
    ).join(
        Rombongan, or_(Pendaftaran.rombongan_pulang_id == Rombongan.id, Pendaftaran.rombongan_kembali_id == Rombongan.id)
    ).outerjoin(
        Bus, or_(Pendaftaran.bus_pulang_id == Bus.id, Pendaftaran.bus_kembali_id == Bus.id)
    ).filter(
        Rombongan.id.in_([r.id for r in rombongan_list])
    ).all()

    # 3. Proses dan strukturkan data untuk ditampilkan di template
    rekap_data = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: {'hadir': [], 'tidak_hadir': []})))
    
    for absen, pendaftaran, santri, bus, rombongan in all_absensi:
        bus_key = bus if bus else 'Tanpa Bus'
        
        if absen.status == 'Hadir':
            rekap_data[rombongan][bus_key][absen.nama_absen]['hadir'].append(santri.nama)
        else:
            rekap_data[rombongan][bus_key][absen.nama_absen]['tidak_hadir'].append(santri.nama)

    return render_template('rekap_absen.html', 
                           rekap_data=rekap_data, 
                           active_edisi=active_edisi,
                           rombongan_list=rombongan_list)

@admin_bp.route('/manajemen-bus')
@login_required
@role_required('Korpus', 'Korda', 'Korwil', 'Korpuspi', 'Sekretaris', 'Bendahara Pusat')
def manajemen_peserta_bus():
    active_edisi = get_active_edisi()
    bus_form = BusForm() # Siapkan form untuk modal

    if not active_edisi:
        flash("Tidak ada edisi yang aktif.", "warning")
        return render_template('manajemen_peserta_bus.html', semua_rombongan=[], bus_form=bus_form)

    # Ambil semua rombongan yang relevan berdasarkan peran
    if current_user.role.name in ['Korda', 'Korwil']:
        # Korda/Korwil hanya melihat rombongan yang mereka kelola dari edisi aktif
        semua_rombongan = current_user.active_managed_rombongan
    else: # Korpus melihat semua
        semua_rombongan = Rombongan.query.options(
            joinedload(Rombongan.buses)
        ).filter_by(edisi_id=active_edisi.id).order_by(Rombongan.nama_rombongan).all()

    return render_template('manajemen_peserta_bus.html', 
                           semua_rombongan=semua_rombongan,
                           active_edisi=active_edisi,
                           bus_form=bus_form) # Kirim bus_form ke template

@admin_bp.route('/alokasi-bus/<int:bus_id>', methods=['GET', 'POST'])
@login_required
@role_required('Korpus', 'Korda', 'Korpuspi')
def alokasi_bus(bus_id):
    bus = Bus.query.get_or_404(bus_id)
    rombongan = bus.rombongan
    
    # Keamanan: Pastikan Korda/Korwil hanya bisa mengakses bus di rombongannya
    if current_user.role.name in ['Korda', 'Korwil']:
        if rombongan not in current_user.active_managed_rombongan:
            abort(403)

    perjalanan = request.args.get('perjalanan', 'pulang') # Default ke 'pulang'

    if request.method == 'POST':
        # Ambil daftar ID pendaftaran dari form yang disubmit
        peserta_ids_in_bus = request.form.getlist('peserta_in_bus[]', type=int)
        
        # Ambil semua pendaftar di rombongan ini
        all_pendaftar_in_rombongan = Pendaftaran.query.filter(
            or_(Pendaftaran.rombongan_pulang_id == rombongan.id, Pendaftaran.rombongan_kembali_id == rombongan.id)
        ).all()
        
        for pendaftar in all_pendaftar_in_rombongan:
            if perjalanan == 'pulang':
                # Jika ID ada di daftar, set bus_id. Jika tidak, set ke None.
                if pendaftar.id in peserta_ids_in_bus:
                    pendaftar.bus_pulang_id = bus.id
                elif pendaftar.bus_pulang_id == bus.id: # Hanya hapus jika sebelumnya di bus ini
                    pendaftar.bus_pulang_id = None
            elif perjalanan == 'kembali':
                if pendaftar.id in peserta_ids_in_bus:
                    pendaftar.bus_kembali_id = bus.id
                elif pendaftar.bus_kembali_id == bus.id:
                    pendaftar.bus_kembali_id = None
        
        db.session.commit()
        flash(f'Alokasi peserta untuk perjalanan {perjalanan} di bus {bus.nama_armada} berhasil disimpan.', 'success')
        return redirect(url_for('admin.alokasi_bus', bus_id=bus_id, perjalanan=perjalanan))

    # Logika untuk GET request (menampilkan data)
    if perjalanan == 'pulang':
        peserta_in_bus = Pendaftaran.query.filter_by(bus_pulang_id=bus.id).join(Santri).order_by(Santri.nama).all()
        peserta_no_bus = Pendaftaran.query.filter(
            Pendaftaran.rombongan_pulang_id == rombongan.id,
            Pendaftaran.status_pulang != 'Tidak Ikut',
            Pendaftaran.bus_pulang_id == None
        ).join(Santri).order_by(Santri.nama).all()
    else: # Perjalanan kembali
        peserta_in_bus = Pendaftaran.query.filter_by(bus_kembali_id=bus.id).join(Santri).order_by(Santri.nama).all()
        peserta_no_bus = Pendaftaran.query.filter(
            or_(Pendaftaran.rombongan_kembali_id == rombongan.id, and_(Pendaftaran.rombongan_kembali_id == None, Pendaftaran.rombongan_pulang_id == rombongan.id)),
            Pendaftaran.status_kembali != 'Tidak Ikut',
            Pendaftaran.bus_kembali_id == None
        ).join(Santri).order_by(Santri.nama).all()
        
    return render_template('alokasi_bus.html', 
                           bus=bus, 
                           rombongan=rombongan,
                           peserta_in_bus=peserta_in_bus,
                           peserta_no_bus=peserta_no_bus,
                           perjalanan=perjalanan)

# Di dalam file app/admin/routes.py

# Di dalam file app/admin/routes.py

@admin_bp.route('/rekapitulasi-global')
@login_required
@role_required('Korpus', 'Korpuspi', 'Korwil')
def rekapitulasi_global():
    active_edisi = get_active_edisi()
    if not active_edisi:
        flash("Tidak ada edisi aktif untuk menampilkan rekapitulasi.", "warning")
        return render_template('rekapitulasi_global.html', title="Rekapitulasi Global", global_stats={}, rekap_rombongan_g1=[], rekap_rombongan_g2=[], terdaftar_tidak_ikut=[], belum_terdaftar=[])

    # 1. Ambil semua data pendaftaran
    semua_pendaftaran = Pendaftaran.query.filter_by(edisi_id=active_edisi.id).options(joinedload(Pendaftaran.santri)).all()
    semua_rombongan = Rombongan.query.filter_by(edisi=active_edisi).order_by(Rombongan.nama_rombongan).all()

    # 2. Pisahkan pendaftar berdasarkan gelombang
    pendaftar_g1 = [p for p in semua_pendaftaran if p.gelombang_pulang == 1 and p.status_pulang != 'Tidak Ikut']
    pendaftar_g2 = [p for p in semua_pendaftaran if p.gelombang_pulang == 2 and p.status_pulang != 'Tidak Ikut']
    
    # Hitung statistik global untuk kembali (ini tidak berubah)
    peserta_ikut_kembali = sum(1 for p in semua_pendaftaran if p.status_kembali != 'Tidak Ikut')
    peserta_tidak_ikut_kembali = sum(1 for p in semua_pendaftaran if p.status_kembali == 'Tidak Ikut')

    global_stats = {
        'pulang_g1': len(pendaftar_g1),
        'pulang_g2': len(pendaftar_g2),
        'kembali': peserta_ikut_kembali,
        'tidak_kembali': peserta_tidak_ikut_kembali
    }

    # 3. Buat rekap per rombongan untuk setiap gelombang
    def buat_rekap_rombongan(pendaftar_gelombang):
        rekap = []
        for rombongan in semua_rombongan:
            count = sum(1 for p in pendaftar_gelombang if p.rombongan_pulang_id == rombongan.id)
            if count > 0: # Hanya tampilkan rombongan yang punya peserta di gelombang ini
                rekap.append({'nama': rombongan.nama_rombongan, 'peserta_pulang': count})
        return rekap

    rekap_rombongan_g1 = buat_rekap_rombongan(pendaftar_g1)
    rekap_rombongan_g2 = buat_rekap_rombongan(pendaftar_g2)

    # 4. Buat format pesan WA untuk setiap gelombang
    def format_pesan_wa(rekap_list, judul):
        pesan_lines = [f"*{judul}:*"]
        if not rekap_list:
            pesan_lines.append("_Tidak ada peserta terdaftar._")
        else:
            for i, data in enumerate(rekap_list, 1):
                pesan_lines.append(f"{i}. {data['nama']} = {data['peserta_pulang']} peserta")
        return "\n".join(pesan_lines)
    
    pesan_wa_pulang_g1 = format_pesan_wa(rekap_rombongan_g1, "Laporan Peserta Pulang Gelombang 1 (26 Sept)")
    pesan_wa_pulang_g2 = format_pesan_wa(rekap_rombongan_g2, "Laporan Peserta Pulang Gelombang 2 (27 Sept)")
    
    # Logika untuk tabel bawah (tidak berubah)
    terdaftar_tidak_ikut = [p for p in semua_pendaftaran if p.status_pulang == 'Tidak Ikut' or p.status_kembali == 'Tidak Ikut']
    santri_terdaftar_ids = {p.santri_id for p in semua_pendaftaran}
    belum_terdaftar = Santri.query.filter(Santri.status_santri != 'Izin', Santri.id.notin_(santri_terdaftar_ids)).order_by(Santri.nama).all()

    return render_template('rekapitulasi_global.html',
                           title="Rekapitulasi Global",
                           global_stats=global_stats,
                           rekap_rombongan_g1=rekap_rombongan_g1,
                           rekap_rombongan_g2=rekap_rombongan_g2,
                           terdaftar_tidak_ikut=terdaftar_tidak_ikut,
                           belum_terdaftar=belum_terdaftar,
                           pesan_wa_pulang_g1=pesan_wa_pulang_g1,
                           pesan_wa_pulang_g2=pesan_wa_pulang_g2,
                           pesan_wa_kembali="Laporan Total Peserta Kembali: " + str(global_stats['kembali']) + " santri.")
                           
def calculate_and_update_biaya(pendaftaran_obj):
    """Menghitung dan memperbarui total biaya pada objek pendaftaran."""
    biaya_pulang = 0
    biaya_kembali = 0
    FEE_PONDOK = 10000

    # Hitung biaya pulang
    if pendaftaran_obj.status_pulang != 'Tidak Ikut' and pendaftaran_obj.rombongan_pulang and pendaftaran_obj.titik_turun:
        tarif_pulang = Tarif.query.filter_by(
            rombongan_id=pendaftaran_obj.rombongan_pulang_id,
            titik_turun=pendaftaran_obj.titik_turun
        ).first()
        if tarif_pulang:
            biaya_pulang = tarif_pulang.harga_bus + tarif_pulang.fee_korda + FEE_PONDOK

    # Hitung biaya kembali
    rombongan_utk_kembali = pendaftaran_obj.rombongan_kembali or pendaftaran_obj.rombongan_pulang
    titik_jemput = pendaftaran_obj.titik_jemput_kembali or pendaftaran_obj.titik_turun
    
    if pendaftaran_obj.status_kembali != 'Tidak Ikut' and rombongan_utk_kembali and titik_jemput:
        tarif_kembali = Tarif.query.filter_by(
            rombongan_id=rombongan_utk_kembali.id,
            titik_turun=titik_jemput
        ).first()
        if tarif_kembali:
            biaya_kembali = tarif_kembali.harga_bus + tarif_kembali.fee_korda + FEE_PONDOK
            
    pendaftaran_obj.total_biaya = biaya_pulang + biaya_kembali

@admin_bp.route('/recalculate-all-fees')
@login_required
@role_required('Korpus') # Hanya Korpus yang boleh menjalankan ini
def recalculate_all_fees():
    active_edisi = get_active_edisi()
    if not active_edisi:
        flash('Tidak ada edisi aktif.', 'warning')
        return redirect(url_for('admin.dashboard'))

    all_pendaftaran = Pendaftaran.query.filter_by(edisi_id=active_edisi.id).all()
    count = 0
    for p in all_pendaftaran:
        old_total = p.total_biaya
        calculate_and_update_biaya(p)
        if p.total_biaya != old_total:
            count += 1
    
    db.session.commit()
    
    flash(f'Selesai! {count} data pendaftaran telah dihitung ulang dan disinkronkan.', 'success')
    return redirect(url_for('admin.daftar_peserta_global'))
